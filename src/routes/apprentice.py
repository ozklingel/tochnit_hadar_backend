import os
import uuid
import datetime
from urllib.parse import urlparse

import boto3
from botocore.exceptions import NoCredentialsError
from flask import Blueprint, request, jsonify
from http import HTTPStatus
from openpyxl.reader.excel import load_workbook

from config import AWS_access_key_id, AWS_secret_access_key, BUCKET, BUCKET_PATH
from src.services import db
from src.models.apprentice_model import Apprentice
from src.models.base_model import Base
from src.models.city_model import City
from src.models.institution_model import Institution
from src.models.task_model_v2 import Task
from src.models.report_model import Report
from src.logic.apprentices import maps_apprentices as maps
from src.logic import apprentices as apprentices_service
from src.logic import update_apprentice

apprentice_profile_form_blueprint = Blueprint(
    "apprentices", __name__, url_prefix="/apprentices"
)


@apprentice_profile_form_blueprint.route("/maps_apprentices", methods=["GET"])
def maps_apprentices():
    try:
        return maps(request.args.get("userId"))
    except Exception as e:
        raise e


@apprentice_profile_form_blueprint.route("/delete", methods=["POST"])
def delete():
    try:

        data = request.json
        apprenticetId = data["apprenticetId"]
        db.session.query(Task).filter(Task.subject == apprenticetId).delete()
        db.session.query(Report).filter(
            Report.ent_reported == apprenticetId,
        ).delete()

        db.session.query(Apprentice).filter(Apprentice.id == apprenticetId).delete()
        db.session.commit()
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.OK
    return jsonify({"result": "success"}), HTTPStatus.OK


@apprentice_profile_form_blueprint.route("/update", methods=["put"])
def update():
    return update_apprentice.update(request=request)


@apprentice_profile_form_blueprint.route("/get_all", methods=["get"])
def get_all_apprentices():
    return apprentices_service.get_all_apprentices()


@apprentice_profile_form_blueprint.route("/add_apprentice_manual", methods=["post"])
def add_apprentice_manual():
    try:
        data = request.json
        first_name = data["first_name"]
        last_name = data["last_name"]
        phone = data["phone"]
        institution_id = data["institution_id"]
        accompany_id = data["accompany_id"]
        city_id = data["city_id"]
        maritalstatus = data["maritalstatus"]
        marriage_date = data["marriage_date"]
        unit_name = data["unit_name"]
        serve_type = data["serve_type"]  # סדיר
        release_date = data["release_date"]
        recruitment_date = data["recruitment_date"]
        onlinestatus = data["onlinestatus"]  # אונלין?
        matsber = data["matsber"]  # מצב רוחני
        hadar_plan_session = data["hadar_plan_session"]  # מחזןר בהדר thperiod
        email = data["email"]
        birthday = data["birthday"]
        address = data["address"]  # רחוב ובית
        identity_id = data["identity_id"]
        institution_mahzor = data["institution_mahzor"]
        militarycompound_id = data["militarycompound_id"]
        cluster_id = (
            db.session.query(Institution.cluster_id)
            .filter(Institution.id == institution_id)
            .first()
        )

        Apprentice1 = Apprentice(
            id=int(phone[1:]),
            name=first_name,
            last_name=last_name,
            phone=phone,
            marriage_status=maritalstatus,
            marriage_date=marriage_date,
            unit_name=unit_name,
            serve_type=serve_type,
            release_date=release_date,
            recruitment_date=recruitment_date,
            accompany_connect_status=onlinestatus,
            spirit_status=matsber,
            hadar_plan_session=hadar_plan_session,
            email=email,
            address=address,
            teudatZehut=identity_id,
            institution_mahzor=institution_mahzor,
            institution_id=institution_id,
            accompany_id=int(accompany_id[1:]),
            base_address=str(militarycompound_id),
            city_id=city_id,
            photo_path="https://www.gravatar.com/avatar",
            birthday=birthday,
            cluster_id=cluster_id.cluster_id,
            association_date=datetime.date.today()
        )
        db.session.add(Apprentice1)
        db.session.commit()
    except Exception as e:
        return (
            jsonify({"result": "error while inserting" + str(e)}),
            HTTPStatus.BAD_REQUEST,
        )

    if Apprentice1:
        # TODO: add contact form to DB
        return jsonify({"result": "success"}), HTTPStatus.OK


@apprentice_profile_form_blueprint.route("/add_apprentice_excel", methods=["put"])
def add_apprentice_excel():
    try:
        file = request.files["file"]
        wb = load_workbook(file)

        sheet = wb.active
        columns = [cell.value for cell in sheet[1]]

        commited_ids = []
        not_commited = []

        unknown_value = "לא ידוע"
        index_ = 2

        for row in sheet.iter_rows(min_row=2):

            def column_value(column_name):
                # Use the Hebrew labels, to handle files with mixing column orders
                try:
                    return row[columns.index(column_name)].value
                except ValueError:
                    return None

            def strip_or_none(label, default=""):
                value = column_value(label)
                return str(value).strip() if not value is None else default

            try:
                name = strip_or_none("שם")
                last_name = strip_or_none("שם משפחה")
                phone = column_value("פלאפון חניך")
                phone = int(str(phone).replace("-", "").replace(" ", "")[:9])
                city = strip_or_none("עיר", unknown_value)
                base_name = strip_or_none("בסיס", unknown_value)
                institution_name = strip_or_none("מוסד")
                if not name or not last_name or not phone or phone < 500000000:
                    not_commited.append(
                        str(index_)
                        + ":"
                        + "שדה חובה לא חוקי-"
                        + name
                        + " "
                        + last_name
                        + " "
                        + str(phone)
                    )
                    index_ = index_ + 1
                    continue
                city_id = db.session.query(City.id).filter(City.name == city).first()
                if city_id is None:
                    not_commited.append(str(index_) + ":ערך עיר לא חוקי- " + str(city))
                    index_ = index_ + 1
                    continue
                military_compound_id = (
                    db.session.query(Base.id).filter(Base.name == base_name).first()
                )
                if (
                    military_compound_id is None
                    and base_name != "משוחרר"
                    and base_name != "טרום גיוס"
                ):
                    not_commited.append(
                        str(index_) + ": ערך בסיס לא חוקי- " + str(base_name)
                    )
                    index_ = index_ + 1
                    continue

                institution_id = (
                    db.session.query(Institution.id)
                    .filter(Institution.name == institution_name)
                    .first()
                )
                if institution_id is None:
                    not_commited.append(
                        str(index_) + ": מוסד לא חוקי- " + institution_name
                    )
                    index_ = index_ + 1
                    continue

                cluster_id = (
                    db.session.query(Institution.cluster_id)
                    .filter(Institution.id == institution_id.id)
                    .first()
                )
                if cluster_id is None:
                    not_commited.append(
                        str(index_) + " אשכול לא חוקי עבור מוסד- " + institution_name
                    )
                    index_ = index_ + 1
                    continue

                Apprentice_ = (
                    db.session.query(Apprentice)
                    .filter(Apprentice.id == str(phone))
                    .first()
                )
                birthday_ivry = None
                if strip_or_none("יום הולדת עברי - יום") and strip_or_none(
                    "יום הולדת עברי חודש"
                ):
                    birthday_ivry = (
                        strip_or_none("יום הולדת עברי - יום")
                        + "' "
                        + strip_or_none("יום הולדת עברי חודש")
                    )
                marriage_date_ivry = None
                if strip_or_none("יום נישואין תאריך עברי - יום") and strip_or_none(
                    "יום נישואין תאריך עברי - חודש"
                ):
                    marriage_date_ivry = (
                        strip_or_none("יום נישואין תאריך עברי - יום")
                        + "' "
                        + strip_or_none("יום נישואין תאריך עברי - חודש")
                    )
                if Apprentice_:

                    Apprentice_.phone = (str(phone),)
                    Apprentice_.city_id = city_id.id
                    Apprentice_.cluster_id = (cluster_id.cluster_id,)
                    Apprentice_.base_address = (
                        (military_compound_id.id,) if military_compound_id else None
                    )
                    Apprentice_.institution_id = (
                        (institution_id.id if institution_id is not None else None),
                    )
                    Apprentice_.name = (name,)
                    Apprentice_.last_name = (last_name,)
                    Apprentice_.address = (strip_or_none("רחוב"),)
                    Apprentice_.teudatZehut = (strip_or_none("תעודת זהות"),)
                    Apprentice_.birthday_ivry = (birthday_ivry,)
                    Apprentice_.birthday = (column_value("יום הולדת לועזי"),)
                    Apprentice_.email = (strip_or_none("מייל חניך"),)
                    Apprentice_.marriage_status = (strip_or_none("סטאטוס משפחתי"),)
                    Apprentice_.serve_type = (strip_or_none("סטטוס"),)
                    Apprentice_.hadar_plan_session = (
                        strip_or_none("מחזור תוכנית הדר"),
                    )
                    Apprentice_.contact1_relation = (
                        strip_or_none("איש קשר - 1 - קירבה"),
                    )
                    Apprentice_.contact1_first_name = (
                        strip_or_none("איש קשר - 1 - שם "),
                    )
                    Apprentice_.contact1_phone = (strip_or_none("איש קשר - 1 -פלאפון"),)
                    Apprentice_.contact1_email = (strip_or_none("איש קשר - 1 - מייל "),)
                    Apprentice_.contact2_relation = (
                        strip_or_none("איש קשר - 2 - קירבה"),
                    )
                    Apprentice_.contact2_first_name = (
                        strip_or_none("איש קשר - 2 - שם "),
                    )
                    Apprentice_.contact2_phone = (
                        strip_or_none("איש קשר - 2 - פלאפון "),
                    )
                    Apprentice_.contact2_email = (strip_or_none("איש קשר - 2 - מייל "),)
                    Apprentice_.contact3_relation = (
                        strip_or_none("איש קשר - 3 - קירבה "),
                    )
                    Apprentice_.contact3_first_name = (
                        strip_or_none("איש קשר - 3 - שם "),
                    )
                    Apprentice_.contact3_phone = (
                        strip_or_none("איש קשר - 3 - טלפון "),
                    )
                    Apprentice_.contact3_email = (strip_or_none("איש קשר - 3 - מייל "),)
                    Apprentice_.marriage_date_ivry = (marriage_date_ivry,)
                    Apprentice_.marriage_date = (
                        strip_or_none("יום נישואין תאריך לועזי", None),
                    )
                    Apprentice_.institution_mahzor = (
                        strip_or_none("מחזור במכינה/ישיבה"),
                    )
                    Apprentice_.teacher_grade_a = (strip_or_none('ר"מ שנה א'),)
                    Apprentice_.teacher_grade_a_phone = (
                        strip_or_none("פלאפון ר״מ שנה א"),
                    )
                    Apprentice_.teacher_grade_b = (strip_or_none('ר"מ שנה ב'),)
                    Apprentice_.teacher_grade_b_phone = (
                        strip_or_none("פלאפון ר״מ שנה ב"),
                    )
                    Apprentice_.paying = column_value("משלם או לא") == "משלם"
                    Apprentice_.spirit_status = (strip_or_none("מצב״ר - מצב רוחני"),)
                    Apprentice_.high_school_name = (strip_or_none("שם תיכון"),)
                    Apprentice_.high_school_teacher = (strip_or_none("שם מחנך תיכון"),)
                    Apprentice_.high_school_teacher_phone = (
                        strip_or_none("פלאפון מחנך תיכון"),
                    )
                    Apprentice_.workstatus = (strip_or_none("עיסוק "),)
                    Apprentice_.workplace = (strip_or_none("מקום עבודה"),)
                    Apprentice_.educationfaculty = (strip_or_none("פקולטה"),)
                    Apprentice_.workoccupation = (strip_or_none("סוג עבודה"),)
                    Apprentice_.army_role = (strip_or_none("סוג שירות"),)  # סיירות
                    Apprentice_.unit_name = (strip_or_none("חטיבה - שיוך יחידתי"),)
                    Apprentice_.militaryPositionNew = (strip_or_none("תפקיד נוכחי"),)
                    Apprentice_.militaryPositionOld = (strip_or_none("תפקיד קודם"),)
                    Apprentice_.recruitment_date = (column_value("תאריך גיוס"),)
                    Apprentice_.release_date = (column_value("תאריך שחרור"),)
                    Apprentice_.accompany_id = (column_value("מלווה אחראי - טלפון"),)
                    Apprentice_.worktype = (column_value("סוג עבודה"),)

                    commited_ids.append(phone)
                    index_ = index_ + 1
                    db.session.add(Apprentice_)
                    continue
                apprentice = Apprentice(
                    id=phone,
                    phone=str(phone),
                    city_id=city_id.id,
                    cluster_id=cluster_id.cluster_id,
                    base_address=(
                        military_compound_id.id if military_compound_id else None
                    ),
                    institution_id=(
                        institution_id.id if institution_id is not None else 0
                    ),
                    name=name,
                    last_name=last_name,
                    address=strip_or_none("רחוב"),
                    teudatZehut=strip_or_none("תעודת זהות"),
                    birthday_ivry=birthday_ivry,
                    birthday=column_value("יום הולדת לועזי"),
                    email=strip_or_none("מייל חניך"),
                    marriage_status=strip_or_none("סטאטוס משפחתי"),
                    serve_type=strip_or_none("סטאטוס"),
                    hadar_plan_session=strip_or_none("מחזור תוכנית הדר"),
                    contact1_relation=strip_or_none("איש קשר - 1 - קירבה"),
                    contact1_first_name=strip_or_none("איש קשר - 1 - שם "),
                    contact1_phone=strip_or_none("איש קשר - 1 -פלאפון"),
                    contact1_email=strip_or_none("איש קשר - 1 - מייל "),
                    contact2_relation=strip_or_none("איש קשר - 2 - קירבה"),
                    contact2_first_name=strip_or_none("איש קשר - 2 - שם "),
                    contact2_phone=strip_or_none("איש קשר - 2 -פלאפון"),
                    contact2_email=strip_or_none("איש קשר - 2 - מייל "),
                    contact3_relation=strip_or_none("איש קשר - 3 - קירבה"),
                    contact3_first_name=strip_or_none("איש קשר - 3 - שם "),
                    contact3_phone=strip_or_none("איש קשר - 3 -פלאפון"),
                    contact3_email=strip_or_none("איש קשר - 3 - מייל "),
                    marriage_date_ivry=marriage_date_ivry,
                    marriage_date=(strip_or_none("יום נישואין תאריך לועזי", None),),
                    institution_mahzor=strip_or_none("מחזור במכינה/ישיבה"),
                    teacher_grade_a=strip_or_none('ר"מ שנה א'),
                    teacher_grade_a_phone=strip_or_none("פלאפון ר״מ שנה א"),
                    teacher_grade_b=strip_or_none('ר"מ שנה ב'),
                    teacher_grade_b_phone=strip_or_none("פלאפון ר״מ שנה ב"),
                    paying=column_value("משלם או לא") == "משלם",
                    spirit_status=strip_or_none("מצב״ר - מצב רוחני"),
                    high_school_name=strip_or_none("שם תיכון"),
                    high_school_teacher=strip_or_none("שם מחנך תיכון"),
                    high_school_teacher_phone=strip_or_none("פלאפון מחנך תיכון"),
                    workstatus=strip_or_none("עיסוק "),
                    workplace=strip_or_none("מקום עבודה"),
                    educationfaculty=strip_or_none("פקולטה"),
                    workoccupation=strip_or_none("סוג עבודה"),
                    army_role=strip_or_none("סוג שירות"),  # סיירות
                    unit_name=strip_or_none("חטיבה - שיוך יחידתי"),
                    militaryPositionNew=strip_or_none("תפקיד נוכחי"),
                    militaryPositionOld=strip_or_none("תפקיד קודם"),
                    recruitment_date=column_value("תאריך גיוס"),
                    release_date=column_value("תאריך שחרור"),
                    accompany_id=column_value("מלווה אחראי - טלפון"),
                    worktype=column_value("סוג עבודה"),
                    association_date=datetime.date.today()

                )

                db.session.add(apprentice)
                commited_ids.append(phone)
                index_ = index_ + 1
            except Exception as e:
                not_commited.append(str(index_) + ":" + str(e))
                index_ = index_ + 1
        db.session.commit()
    except Exception as e:
        return (
            jsonify({"result": "Error while inserting " + str(e)}),
            HTTPStatus.BAD_REQUEST,
        )
    return jsonify(
        {
            "result": "success",
            "not_commited": [x for x in not_commited if x is not None],
        }
    )
@apprentice_profile_form_blueprint.route("/upload_and_create", methods=["put"])
def upload_and_create():
    # Assign directory
    directory = r"C:\Users\OZKL\Downloads\picture_test"
    session = boto3.Session()
    s3 = session.client(
        "s3",
        aws_access_key_id=AWS_access_key_id,
        aws_secret_access_key=AWS_secret_access_key,
    )

    for filename in os.listdir(directory):
        if filename.endswith('.jpg'):  # Only process PNG files
            file_path = os.path.join(directory, filename)
            new_filename = (
                    uuid.uuid4().hex + filename
            )
            try:
                # Upload file to S3
                apprentice_id = filename.split(".")[0][3:]
                updated_ent = Apprentice.query.get(apprentice_id)
                if updated_ent:
                    if updated_ent.photo_path:
                        delete_s3_object_by_url(updated_ent.photo_path)
                    s3.upload_file(file_path, BUCKET, new_filename)
                    url_path_to_image = BUCKET_PATH + new_filename
                    setattr(updated_ent, "photo_path", url_path_to_image)
                    db.session.add(updated_ent)
                print(f'{filename} uploaded successfully!')
            except FileNotFoundError:
                return jsonify({"result": f'The file {filename} was not found!'}), HTTPStatus.OK
            except NoCredentialsError:
                return jsonify({"result": 'Credentials not available!'}), HTTPStatus.OK
            except Exception as e:
                return jsonify({"result":f'An error occurred: {e}'}), HTTPStatus.OK

    db.session.commit()
    return jsonify({"result": "success"}), HTTPStatus.OK


def delete_s3_object_by_url(s3_url):
    # Parse the S3 URL
    parsed_url = urlparse(s3_url)

    # Extract bucket name and key
    bucket_name = parsed_url.netloc.split('.')[0]  # Assumes standard AWS S3 URL format
    object_key = parsed_url.path.lstrip('/')  # Remove leading "/"

    # Initialize an S3 client
    s3_client = boto3.client('s3')

    # Delete the object
    response = s3_client.delete_object(Bucket=bucket_name, Key=object_key)

    return response