
import uuid
import boto3
from flask import Blueprint, request, jsonify
from http import HTTPStatus
from datetime import datetime
from openpyxl.reader.excel import load_workbook
from datetime import datetime,date

import config
from app import db, red
from config import AWS_secret_access_key, AWS_access_key_id
from src.models.apprentice_model import Apprentice, front_end_dict
from src.models.base_model import Base
from src.models.city_model import City
from src.models.cluster_model import Cluster
from src.models.institution_model import Institution
from src.models.notification_model import notifications
from src.models.user_model import user1
from src.models.visit_model import Visit
from src.routes.setEntityDetails_form_routes import validate_email, validate_date

apprentice_Profile_form_blueprint = Blueprint('apprentice_Profile_form', __name__, url_prefix='/apprentice_Profile_form')
@apprentice_Profile_form_blueprint.route('/delete', methods=['POST'])
def delete():
    try:
        data = request.json
        apprenticetId = data['apprenticetId']
        print(apprenticetId)
        res = db.session.query(notifications).filter(notifications.subject == apprenticetId, ).delete()
        res = db.session.query(Visit).filter(Visit.ent_reported == apprenticetId, ).delete()
        res = db.session.query(Apprentice).filter(Apprentice.id == apprenticetId).delete()
        db.session.commit()
    except Exception as e:
        return jsonify({"result": str(e)}),HTTPStatus.OK
    return jsonify({"result":"success"}), HTTPStatus.OK
        # return jsonify([{'id':str(noti.id),'result': 'success',"apprenticeId":str(noti.apprenticeid),"date":str(noti.date),"timeFromNow":str(noti.timefromnow),"event":str(noti.event),"allreadyread":str(noti.allreadyread)}]), HTTPStatus.OK

@apprentice_Profile_form_blueprint.route("/update", methods=['put'])
def update():
    try:
        # get tasksAndEvents
        apprenticetId = request.args.get("apprenticetId")
        data = request.json
        updatedEnt = Apprentice.query.get(apprenticetId)

        for key in data:

            if key == "email" :
                if validate_email(data[key]):
                    setattr(updatedEnt, key, data[key])
                else:
                    return jsonify({'result': "email or date -wrong format"}), 401
            elif  key == "birthday":
                if validate_date(data[key]):
                    setattr(updatedEnt, key, data[key])
                else:
                    return jsonify({'result': "email or date -wrong format"}), 401
            else:
                print(front_end_dict[key])
                setattr(updatedEnt, front_end_dict[key], data[key])
        db.session.commit()
        if updatedEnt:
            # print(f'setWasRead form: subject: [{subject}, notiId: {notiId}]')
            # TODO: add contact form to DB
            return jsonify({'result': 'success'}), HTTPStatus.OK
        return jsonify({'result': 'error'}), 401
    except Exception as e:
        return jsonify({'result': str(e)}), 401

@apprentice_Profile_form_blueprint.route("/add_apprentice_manual", methods=['post'])
def add_apprentice():
    data = request.json
    try:
        first_name = data['first_name']
        last_name = data['last_name']
        phone = data['phone']
        institution_name = data['institution_name']
        accompany_id = data['accompany_id']
        birthday = data['birthday']
        city_name = data['city_name']
        maritalstatus= data['maritalstatus']
        marriage_date= data['marriage_date']
        unit_name= data['unit_name']
        serve_type= data['serve_type']#סדיר
        release_date= data['release_date']
        recruitment_date= data['recruitment_date']
        onlinestatus= data['onlinestatus']#אונלין?
        matsber= data['matsber']#מצב רוחני
        hadar_plan_session= data['hadar_plan_session']#מחזןר בהדר thperiod
        email= data['email']
        birthday= data['birthday']
        address= data['address']#רחוב ובית
        teudatzehut= data['teudatzehut']
        institution_mahzor= data['institution_mahzor']
        militarycompound_name= data['militarycompound_name']

        militarycompoundid = db.session.query(Base).filter(Base.name==militarycompound_name).first()
        CityId = db.session.query(City).filter(City.name==city_name).first()
        institution_id = db.session.query(Institution.id).filter(Institution.name==institution_name).first()
        institution_id=institution_id[0] if institution_id is not None else 0
        Apprentice1 = Apprentice(
            id=int(phone[1:]),
            name=first_name,
            last_name=last_name,
            phone=phone,
            marriage_status=maritalstatus,
            marriage_date=marriage_date,
            unit_name=unit_name,
            serve_type=serve_type,
            release_date=release_date,
            recruitment_date=recruitment_date,
            accompany_connect_status=onlinestatus,
            spirit_status=matsber,
            hadar_plan_session=hadar_plan_session,
            email=email,
            address=address,
            teudatZehut=teudatzehut,
            institution_mahzor=institution_mahzor,
            institution_id=institution_id,
            accompany_id=int(accompany_id[1:]),
            base_address=str(militarycompoundid.id),
            city_id=CityId.id,
            photo_path="https://www.gravatar.com/avatar",
            birthday=birthday
        )
        db.session.add(Apprentice1)
        db.session.commit()
    except Exception as e:
        return jsonify({'result': 'error while inserting'+str(e)}), HTTPStatus.BAD_REQUEST

    if Apprentice1:
        # TODO: add contact form to DB
        return jsonify({'result': 'success'}), HTTPStatus.OK


@apprentice_Profile_form_blueprint.route("/add_apprentice_excel", methods=['put'])
def add_apprentice_excel():
    #/home/ubuntu/flaskapp/
    file = request.files['file']
    wb = load_workbook(file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2):
        first_name = row[0].value.strip()
        last_name = row[1].value.strip()
        phone = row[2].value
        city = row[3].value.strip()
        address = row[4].value.strip()
        teudatZehut = row[5].value#מפקד?
        birthday_Ivry_month = row[6].value.strip()#מפקד?
        birthday_Ivry_day = row[7].value.strip()#מפקד?
        birthday_Ivry=birthday_Ivry_day+" "+birthday_Ivry_month
        birthday_loazi = row[8].value.strip()#מפקד?
        mail = row[9].value.strip() if row[9].value else ""#מפקד?

        marriage_status = row[10].value.strip()
        serve_type = row[11].value.strip()
        hadar_plan_session = row[12].value
        contact1_first_relation = row[13].value.strip()
        contact1_first_name = row[14].value.strip()
        contact1_phone = row[15].value
        contact1_email = row[16].value.strip()
        contact2_first_relation = row[17].value.strip()
        contact2_first_name = row[18].value.strip()
        contact2_phone = row[19].value
        contact2_email = row[20].value.strip()
        contact3_first_relation = row[21].value.strip()
        contact3_first_name = row[22].value.strip()
        contact3_phone = row[23].value
        contact3_email = row[24].value.strip()
        marriage_date_month =row[25].value
        marriage_date_day =row[26].value
        marriage_date=marriage_date_day+" "+marriage_date_month
        marriage_date_loazi =row[27].value
        institution_mahzor=row[28].value
        teacher_grade_a = row[29].value.strip()
        teacher_grade_a_phone = row[30].value
        teacher_grade_b = row[31].value.strip()
        teacher_grade_b_phone = row[32].value

        paying=row[33].value.strip()
        matzbar=row[34].value.strip()
        high_school_name=row[35].value.strip()
        high_school_teacher=row[36].value
        high_school_teacher_phone=row[37].value
        workstatus=row[38].value.strip()#isuk
        workplace=row[39].value.strip()#city
        educationfaculty=row[40].value.strip()#ariel
        workoccupation=row[41].value.strip()#Anaf
        worktype=row[42].value.strip()#
        unit_name = row[43].value.strip() #חיל שריון
        army_role=row[44].value.strip()#צנחנים
        militaryPositionNew = row[46].value.strip()#מפקצ
        militaryPositionOld=row[47].value.strip()#צנחנים
        recruitment_date=row[48].value.strip()#צנחנים
        release_date=row[49].value.strip()#צנחנים
        base_name=row[50].value.strip()
        institution_name = row[51].value.strip()
        accompany_id = row[52].value#מפקד?

        CityId = db.session.query(City.id).filter(City.name==city).first()
        militaryCompoundId = db.session.query(Base.id).filter(Base.name == base_name).first()
        print(institution_name)
        institution_id = db.session.query(Institution.id).filter(Institution.name == institution_name).first()
        eshcol=db.session.query(Institution.eshcol_id).filter(Institution.id == institution_id.id).first()

        try:
            institution_id = db.session.query(Institution.id).filter(Institution.name == str(institution_name)).first()
            Apprentice1 = Apprentice(
                email=mail,
                high_school_teacher=high_school_teacher,
                #release_date=release_date,
                #recruitment_date=recruitment_date,
                militaryPositionOld=militaryPositionOld,
                militaryPositionNew=militaryPositionNew,
                institution_mahzor=institution_mahzor,
                teacher_grade_a_phone=teacher_grade_a_phone,
                teacher_grade_b_phone=teacher_grade_b_phone,
                city_id=CityId.id,
                id=phone,
                eshcol=eshcol.eshcol_id,
                base_address=militaryCompoundId.id,
                institution_id=institution_id.id if institution_id is not None else 0,
                address=address,
                serve_type=serve_type,
                name=first_name,
                last_name=last_name,
                phone=str(phone),
                army_role=army_role,
                marriage_status=marriage_status,
                contact1_email=contact1_email,
                teacher_grade_b=teacher_grade_b,
                teacher_grade_a=teacher_grade_a,
                hadar_plan_session=hadar_plan_session,
                contact2_phone=contact2_phone,
                contact2_first_name=contact2_first_name,
                contact1_phone=contact1_phone,
                contact1_first_name=contact1_first_name,
                teudatZehut=teudatZehut,
                birthday_ivry=birthday_Ivry,
                #birthday=birthday_loazi,
                unit_name=unit_name,
                accompany_id=accompany_id,
            contact3_email=contact3_email,
            contact3_first_name=contact3_first_name,
            contact3_phone=contact3_phone,
            #release_date=release_date,
            #recruitment_date=recruitment_date,
            marriage_date_ivry=marriage_date,
            #marriage_date=marriage_date_loazi,
            spirit_status=matzbar,
            contact2_email=contact2_email,
            worktype=worktype,
            workoccupation=workoccupation,
            educationfaculty=educationfaculty,
            workplace=workplace,
            workstatus=workstatus,
            high_school_teacher_phone=high_school_teacher_phone,
            high_school_name=high_school_name,
            paying=paying,
            contact1_relation=contact1_first_relation,
            contact2_relation=contact2_first_relation,
            contact3_relation=contact3_first_relation

            )
            db.session.add(Apprentice1)
            db.session.commit()
        except Exception as e:
            return jsonify({'result': 'error while inserting' + str(e)}), HTTPStatus.BAD_REQUEST

    return jsonify({'result': 'success'}), HTTPStatus.OK

@apprentice_Profile_form_blueprint.route('/myApprentices', methods=['GET'])
def getmyApprentices_form():
    try:
        created_by_id = request.args.get('userId')
        print(created_by_id)
        apprenticeList=[]
        user1ent = db.session.query(user1.role_ids,user1.institution_id,user1.eshcol).filter(user1.id==created_by_id).first()
        if "0" in user1ent.role_ids:
            apprenticeList = db.session.query(Apprentice).filter(Apprentice.accompany_id == created_by_id).all()
        if "1" in user1ent.role_ids:
            apprenticeList = db.session.query(Apprentice).filter(Apprentice.institution_id == user1ent.institution_id).all()
        if "2" in user1ent.role_ids:
            apprenticeList = db.session.query(Apprentice).filter(Apprentice.eshcol == user1ent.eshcol).all()

        my_dict = []

        for noti in apprenticeList:
            call_status=visit_gap_color(config.call_report, noti, 30, 15)
            personalMeet_status=visit_gap_color(config.personalMeet_report, noti, 100, 80)
            Horim_status=visit_gap_color(config.HorimCall_report, noti, 365, 350)
            print(noti.city_id)
            city = db.session.query(City).filter(City.id == noti.city_id).first()
            print(city)
            reportList = db.session.query(Visit.id).filter(Visit.ent_reported == noti.id).all()
            eventlist = db.session.query(notifications.id,notifications.event,notifications.details,notifications.date).filter(
                                                                               notifications.subject == str(noti.id),
                                                                               notifications.numoflinesdisplay == 3).all()
            base_id = db.session.query(Base.id).filter(Base.id == int(noti.base_address)).first()
            base_id = base_id[0] if base_id else 0
            my_dict.append(
                {"Horim_status":Horim_status,
                 "personalMeet_status":personalMeet_status,
                 "call_status":call_status,
                 "highSchoolRavMelamed_phone": noti.high_school_teacher_phone
                         ,"highSchoolRavMelamed_name": noti.high_school_teacher,
                      "highSchoolRavMelamed_email": noti.high_school_teacher_email,

                     "thRavMelamedYearA_name": noti.teacher_grade_a,
                     "thRavMelamedYearA_phone": noti.teacher_grade_a_phone,
                     "thRavMelamedYearA_email": noti.teacher_grade_a_email,

                     "thRavMelamedYearB_name": noti.teacher_grade_b,
                     "thRavMelamedYearB_phone": noti.teacher_grade_b_phone,
                     "thRavMelamedYearB_email": noti.teacher_grade_b_email,
                    "address": {
                        "country": "IL",
                        "city": city.name if city else "",
                        "cityId": str(noti.city_id),
                        "street": noti.address,
                        "houseNumber": "1",
                        "apartment": "1",
                        "region": str(city.cluster_id) if city else "",
                        "entrance": "a",
                        "floor": "1",
                        "postalCode": "12131",
                        "lat": 32.04282620026557,#no need city cord
                        "lng": 34.75186193813887
                    },
                 "contact1_first_name": noti.contact1_first_name,
                 "contact1_last_name": noti.contact1_last_name,
                 "contact1_phone": noti.contact1_phone,
                 "contact1_email": noti.contact1_email,
                 "contact1_relation": noti.contact1_relation,
                 "contact2_first_name": noti.contact2_first_name,
                 "contact2_last_name": noti.contact2_last_name,
                 "contact2_phone": noti.contact2_phone,
                 "contact2_email": noti.contact2_email,
                 "contact2_relation": noti.contact2_relation,
                 "contact3_first_name": noti.contact3_first_name,
                 "contact3_last_name": noti.contact3_last_name,
                 "contact3_phone": noti.contact3_phone,
                 "contact3_email": noti.contact3_email,
                 "contact3_relation": noti.contact3_relation,
                 "activity_score": len(reportList),

                 "reports":
                     [str(i[0]) for i in [tuple(row) for row in reportList]]
                 ,
                 "events":

                      [{"id" :str(row[0]),"title":row[1],"description":row[2],"date" : toISO(row[3])} for row in eventlist]

                    , "id": str(noti.id), "thMentor": str(noti.accompany_id),
                 "militaryPositionNew": str(noti.militaryPositionNew)
                    , "avatar": noti.photo_path if noti.photo_path is not None else 'https://www.gravatar.com/avatar' , "name": str(noti.name), "last_name": str(noti.last_name),
                 "institution_id": str(noti.institution_id), "thPeriod": str(noti.hadar_plan_session),
                 "serve_type": noti.serve_type,
                 "marriage_status": str(noti.marriage_status), "militaryCompoundId": str(base_id),
                 "phone": noti.phone, "email": noti.email, "teudatZehut": noti.teudatZehut,
                 "birthday": toISO(noti.birthday),  "marriage_date": toISO(noti.marriage_date),
                 "highSchoolInstitution": noti.highSchoolInstitution, "army_role": noti.army_role,
                 "unit_name": noti.unit_name,
                  "matsber": str(noti.spirit_status),
                 "militaryDateOfDischarge": toISO(noti.release_date),
                 "militaryDateOfEnlistment": toISO(noti.recruitment_date)
                    , "militaryUpdatedDateTime": toISO(noti.militaryupdateddatetime),
                 "militaryPositionOld": noti.militaryPositionOld, "educationalInstitution": noti.educationalinstitution
                    , "educationFaculty": noti.educationfaculty, "workOccupation": noti.workoccupation,
                 "workType": noti.worktype, "workPlace": noti.workplace, "workStatus": noti.workstatus, "paying": noti.paying

                 })

        if apprenticeList is None:
            # acount not found
            return jsonify({"result":"Wrong id"})
        if apprenticeList ==[]:
            # acount not found
            return jsonify([])
        else:
            # print(f' notifications: {my_dict}]')
            # TODO: get Noti form to DB
            return jsonify(my_dict), HTTPStatus.OK
            # return jsonify([{'id':str(noti.id),'result': 'success',"apprenticeId":str(noti.apprenticeid),"date":str(noti.date),"timeFromNow":str(noti.timefromnow),"event":str(noti.event),"allreadyread":str(noti.allreadyread)}]), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.OK


def toISO(d):
    if d:
        return datetime(d.year, d.month, d.day).isoformat()
    else:
        return None
@apprentice_Profile_form_blueprint.route('/maps_apprentices', methods=['GET'])
def maps_apprentices():
    try:
        created_by_id = request.args.get('userId')
        print(created_by_id)
        apprenticeList=[]
        user1ent = db.session.query(user1.role_ids,user1.institution_id,user1.eshcol).filter(user1.id==created_by_id).first()
        if "0" in user1ent.role_ids:
            apprenticeList = db.session.query(Apprentice).filter(Apprentice.institution_id == user1ent.institution_id).all()
        if "1" in user1ent.role_ids:
            apprenticeList = db.session.query(Apprentice).filter(Apprentice.eshcol == user1ent.eshcol).all()
        if "2" in user1ent.role_ids or "3" in user1ent.role_ids:
            apprenticeList = db.session.query(Apprentice).all()

        my_dict = []

        for noti in apprenticeList:
            call_status=visit_gap_color(config.call_report, noti, 30, 15)
            personalMeet_status=visit_gap_color(config.personalMeet_report, noti, 100, 80)
            Horim_status=visit_gap_color(config.HorimCall_report, noti, 365, 350)
            print(noti.city_id)
            city = db.session.query(City).filter(City.id == noti.city_id).first()
            print(city)
            reportList = db.session.query(Visit.id).filter(Visit.ent_reported == noti.id).all()
            eventlist = db.session.query(notifications.id,notifications.event,notifications.details,notifications.date).filter(
                                                                               notifications.subject == str(noti.id),
                                                                               notifications.numoflinesdisplay == 3).all()
            base_id = db.session.query(Base.id).filter(Base.id == int(noti.base_address)).first()
            base_id = base_id[0] if base_id else 0
            my_dict.append(
                {
                    "accompany_id":str(noti.accompany_id),"Horim_status":Horim_status,
                 "personalMeet_status":personalMeet_status,
                 "call_status":call_status,
                 "highSchoolRavMelamed_phone": noti.high_school_teacher_phone
                         ,"highSchoolRavMelamed_name": noti.high_school_teacher,
                      "highSchoolRavMelamed_email": noti.high_school_teacher_email,
                     "thRavMelamedYearA_name": noti.teacher_grade_a,
                     "thRavMelamedYearA_phone": noti.teacher_grade_a_phone,
                     "thRavMelamedYearA_email": noti.teacher_grade_a_email,
                     "thRavMelamedYearB_name": noti.teacher_grade_b,
                     "thRavMelamedYearB_phone": noti.teacher_grade_b_phone,
                     "thRavMelamedYearB_email": noti.teacher_grade_b_email,
                    "address": {
                        "country": "IL",
                        "city": city.name if city else "",
                        "cityId": str(noti.city_id),
                        "street": noti.address,
                        "houseNumber": "1",
                        "apartment": "1",
                        "region": str(city.cluster_id) if city else "",
                        "entrance": "a",
                        "floor": "1",
                        "postalCode": "12131",
                        "lat": 32.04282620026557,#no need city cord
                        "lng": 34.75186193813887
                    },
                 "contact1_first_name": noti.contact1_first_name,
                 "contact1_last_name": noti.contact1_last_name,
                 "contact1_phone": noti.contact1_phone,
                 "contact1_email": noti.contact1_email,
                 "contact1_relation": noti.contact1_relation,
                 "contact2_first_name": noti.contact2_first_name,
                 "contact2_last_name": noti.contact2_last_name,
                 "contact2_phone": noti.contact2_phone,
                 "contact2_email": noti.contact2_email,
                 "contact2_relation": noti.contact2_relation,
                 "contact3_first_name": noti.contact3_first_name,
                 "contact3_last_name": noti.contact3_last_name,
                 "contact3_phone": noti.contact3_phone,
                 "contact3_email": noti.contact3_email,
                 "contact3_relation": noti.contact3_relation,
                 "activity_score": len(reportList),

                 "reports":
                     [str(i[0]) for i in [tuple(row) for row in reportList]]
                 ,
                 "events":

                      [{"id" :str(row[0]),"title":row[1],"description":row[2],"date" : toISO(row[3])} for row in eventlist]

                    , "id": str(noti.id), "thMentor": str(noti.accompany_id),
                 "militaryPositionNew": str(noti.militaryPositionNew)
                    , "avatar": noti.photo_path if noti.photo_path is not None else 'https://www.gravatar.com/avatar' , "name": str(noti.name), "last_name": str(noti.last_name),
                 "institution_id": str(noti.institution_id), "thPeriod": str(noti.hadar_plan_session),
                 "serve_type": noti.serve_type,
                 "marriage_status": str(noti.marriage_status), "militaryCompoundId": str(base_id),
                 "phone": noti.phone, "email": noti.email, "teudatZehut": noti.teudatZehut,
                 "birthday": toISO(noti.birthday),  "marriage_date": toISO(noti.marriage_date),
                 "highSchoolInstitution": noti.highSchoolInstitution, "army_role": noti.army_role,
                 "unit_name": noti.unit_name,
                  "matsber": str(noti.spirit_status),
                 "militaryDateOfDischarge": toISO(noti.release_date),
                 "militaryDateOfEnlistment": toISO(noti.recruitment_date)
                    , "militaryUpdatedDateTime": toISO(noti.militaryupdateddatetime),
                 "militaryPositionOld": noti.militaryPositionOld, "educationalInstitution": noti.educationalinstitution
                    , "educationFaculty": noti.educationfaculty, "workOccupation": noti.workoccupation,
                 "workType": noti.worktype, "workPlace": noti.workplace, "workStatus": noti.workstatus, "paying": noti.paying

                 })

        if apprenticeList is None:
            # acount not found
            return jsonify({"result":"Wrong id"})
        if apprenticeList ==[]:
            # acount not found
            return jsonify([])
        else:
            # print(f' notifications: {my_dict}]')
            # TODO: get Noti form to DB
            return jsonify(my_dict), HTTPStatus.OK
            # return jsonify([{'id':str(noti.id),'result': 'success',"apprenticeId":str(noti.apprenticeid),"date":str(noti.date),"timeFromNow":str(noti.timefromnow),"event":str(noti.event),"allreadyread":str(noti.allreadyread)}]), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.OK
def visit_gap_color(type, apprentice, redLine, greenLine):
    # Apprentice_call_status
    visitEvent = db.session.query(Visit).filter(Visit.ent_reported == apprentice.id,
                                                Visit.title == type).order_by(Visit.visit_date.desc()).first()
    # handle no row so insert need a call
    if visitEvent is None:
        return "red"
    gap = (date.today() - visitEvent.visit_date).days if visitEvent is not None else 0
    if gap > redLine:
        return "red"
    if redLine >= gap > greenLine:
        return "orange"
    if greenLine >= gap:
        return "green"