import boto3
from flask import Blueprint, request, jsonify
from http import HTTPStatus
from openpyxl.reader.excel import load_workbook
from datetime import date
import config

from src.services import db
from config import AWS_secret_access_key, AWS_access_key_id
from src.models.models_utils import to_iso
from src.models.apprentice_model import Apprentice, front_end_dict
from src.models.base_model import Base
from src.models.city_model import City
from src.models.institution_model import Institution
from src.models.task_model import Task
from src.models.user_model import User
from src.models.report_model import Report
from src.routes.set_entity_details_form_routes import validate_email, validate_date
from src.logic.apprentices import maps_apprentices as maps

apprentice_profile_form_blueprint = Blueprint(
    "apprentice_Profile_form", __name__, url_prefix="/apprentice_Profile_form"
)


@apprentice_profile_form_blueprint.route("/delete", methods=["POST"])
def delete():
    try:

        data = request.json
        apprenticetId = data["apprenticetId"]
        res = db.session.query(Task).filter(
            Task.subject == apprenticetId).delete()
        res = (
            db.session.query(Report)
            .filter(
                Report.ent_reported == apprenticetId,
            )
            .delete()
        )
        res = (
            db.session.query(Apprentice).filter(
                Apprentice.id == apprenticetId).delete()
        )
        db.session.commit()
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.OK
    return jsonify({"result": "success"}), HTTPStatus.OK
    # return jsonify([{'id':str(noti.id),'result': 'success',"apprenticeId":str(noti.apprenticeid),"date":str(noti.date),"timeFromNow":str(noti.timefromnow),"event":str(noti.event),"allreadyread":str(noti.allreadyread)}]), HTTPStatus.OK


@apprentice_profile_form_blueprint.route("/update", methods=["put"])
def update():
    try:

        # get tasksAndEvents
        apprenticetId = request.args.get("apprenticetId")
        data = request.json
        updatedEnt = Apprentice.query.get(apprenticetId)

        for key in data:
            if key == "avatar":
                s3 = boto3.resource(
                    "s3",
                    aws_access_key_id=AWS_access_key_id,
                    aws_secret_access_key=AWS_secret_access_key,
                )
                s3.Object("th01-s3", data[key]).delete()
            if key == "email":
                if validate_email(data[key]):
                    setattr(updatedEnt, key, data[key])
                else:
                    return jsonify({"result": "email -wrong format"}), 401
            elif key == "birthday":
                if validate_date(data[key][:-9]):
                    setattr(updatedEnt, key, data[key])
                else:
                    return jsonify({"result": "email or date -wrong format"}), 401
            else:
                setattr(updatedEnt, front_end_dict[key], data[key])
        db.session.commit()
        if updatedEnt:
            # TODO: add contact form to DB
            return jsonify({"result": "success"}), HTTPStatus.OK
        return jsonify({"result": "error"}), 401
    except Exception as e:
        return jsonify({"result": str(e)}), 401


@apprentice_profile_form_blueprint.route("/add_apprentice_manual", methods=["post"])
def add_apprentice_manual():
    try:

        data = request.json
        first_name = data["first_name"]
        last_name = data["last_name"]
        phone = data["phone"]
        institution_name = data["institution_name"]
        accompany_id = data["accompany_id"]
        birthday = data["birthday"]
        city_name = data["city_name"]
        maritalstatus = data["maritalstatus"]
        marriage_date = data["marriage_date"]
        unit_name = data["unit_name"]
        serve_type = data["serve_type"]  # סדיר
        release_date = data["release_date"]
        recruitment_date = data["recruitment_date"]
        onlinestatus = data["onlinestatus"]  # אונלין?
        matsber = data["matsber"]  # מצב רוחני
        hadar_plan_session = data["hadar_plan_session"]  # מחזןר בהדר thperiod
        email = data["email"]
        birthday = data["birthday"]
        address = data["address"]  # רחוב ובית
        teudatzehut = data["teudatzehut"]
        institution_mahzor = data["institution_mahzor"]
        militarycompound_name = data["militarycompound_name"]

        militarycompoundid = (
            db.session.query(Base).filter(
                Base.name == militarycompound_name).first()
        )
        CityId = db.session.query(City).filter(City.name == city_name).first()
        institution_id = (
            db.session.query(Institution.id)
            .filter(Institution.name == institution_name)
            .first()
        )
        institution_id = institution_id[0] if institution_id is not None else 0
        Apprentice1 = Apprentice(
            id=int(phone[1:]),
            name=first_name,
            last_name=last_name,
            phone=phone,
            marriage_status=maritalstatus,
            marriage_date=marriage_date,
            unit_name=unit_name,
            serve_type=serve_type,
            release_date=release_date,
            recruitment_date=recruitment_date,
            accompany_connect_status=onlinestatus,
            spirit_status=matsber,
            hadar_plan_session=hadar_plan_session,
            email=email,
            address=address,
            teudatZehut=teudatzehut,
            institution_mahzor=institution_mahzor,
            institution_id=institution_id,
            accompany_id=int(accompany_id[1:]),
            base_address=str(militarycompoundid.id),
            city_id=CityId.id,
            photo_path="https://www.gravatar.com/avatar",
            birthday=birthday,
        )
        db.session.add(Apprentice1)
        db.session.commit()
    except Exception as e:
        return (
            jsonify({"result": "error while inserting" + str(e)}),
            HTTPStatus.BAD_REQUEST,
        )

    if Apprentice1:
        # TODO: add contact form to DB
        return jsonify({"result": "success"}), HTTPStatus.OK


@apprentice_profile_form_blueprint.route("/add_apprentice_excel", methods=["put"])
def add_apprentice_excel():
    try:
        file = request.files["file"]
        wb = load_workbook(file)

        sheet = wb.active
        columns = [cell.value for cell in sheet[1]]

        commited_ids = []
        uncommited_ids = []

        unknown_value = "לא ידוע"

        for row in sheet.iter_rows(min_row=2):

            def column_value(column_name):
                # Use the Hebrew labels, to handle files with mixing column orders
                try:
                    return row[columns.index(column_name)].value
                except ValueError:
                    return None

            def strip_or_none(label, default=""):
                value = column_value(label)
                return str(value).strip() if not value is None else default

            try:
                name = strip_or_none("שם")
                last_name = strip_or_none("שם משפחה")
                phone = column_value("פלאפון חניך")
                if type(phone) is str:
                    phone = int("".join(filter(str.isdigit, phone)))

                if phone in commited_ids:
                    continue

                if not name or not last_name or not phone:
                    print("mandatory ", name, last_name, phone)

                    uncommited_ids.append(phone)
                    continue

                city = strip_or_none("עיר", unknown_value)
                base_name = strip_or_none("בסיס", unknown_value)
                institution_name = strip_or_none("מוסד")

                city_id = db.session.query(City.id).filter(
                    City.name == city).first()
                if city_id is None:
                    city_id = (
                        db.session.query(City.id)
                        .filter(City.name == unknown_value)
                        .first()
                    )

                military_compound_id = (
                    db.session.query(Base.id).filter(
                        Base.name == base_name).first()
                )
                institution_id = (
                    db.session.query(Institution.id)
                    .filter(Institution.name == institution_name)
                    .first()
                )

                if institution_id is None or military_compound_id is None:
                    print("base,institution_id ",
                          military_compound_id, institution_id)

                    uncommited_ids.append(phone)
                    continue

                cluster_id = (
                    db.session.query(Institution.cluster_id)
                    .filter(Institution.id == institution_id.id)
                    .first()
                )
                if cluster_id is None:
                    uncommited_ids.append(phone)
                    print("cluster")
                    continue

                institution_id = (
                    db.session.query(Institution.id)
                    .filter(Institution.name == str(institution_name))
                    .first()
                )
                apprentice = Apprentice(
                    id=phone,
                    phone=str(phone),
                    city_id=city_id.id,
                    cluster_id=cluster_id.cluster_id,
                    base_address=military_compound_id.id,
                    institution_id=(
                        institution_id.id if institution_id is not None else 0
                    ),
                    name=name,
                    last_name=last_name,
                    address=strip_or_none("רחוב"),
                    teudatZehut=strip_or_none("תעודת זהות"),
                    birthday_ivry=strip_or_none("יום הולדת עברי - יום")
                    + "' "
                    + strip_or_none("יום הולדת עברי חודש")
                    + strip_or_none("יום הולדת עברי - שנה"),
                    birthday=column_value("יום הולדת לועזי"),
                    email=strip_or_none("מייל חניך"),
                    marriage_status=strip_or_none("סטאטוס משפחתי"),
                    serve_type=strip_or_none("סטאטוס"),
                    hadar_plan_session=strip_or_none("מחזור תוכנית הדר"),
                    contact1_relation=strip_or_none("איש קשר - 1 - קירבה"),
                    contact1_first_name=strip_or_none("איש קשר - 1 - שם "),
                    contact1_phone=strip_or_none("איש קשר - 1 -פלאפון"),
                    contact1_email=strip_or_none("איש קשר - 1 - מייל "),
                    contact2_relation=strip_or_none("איש קשר - 2 - קירבה"),
                    contact2_first_name=strip_or_none("איש קשר - 2 - שם "),
                    contact2_phone=strip_or_none("איש קשר - 2 -פלאפון"),
                    contact2_email=strip_or_none("איש קשר - 2 - מייל "),
                    contact3_relation=strip_or_none("איש קשר - 3 - קירבה"),
                    contact3_first_name=strip_or_none("איש קשר - 3 - שם "),
                    contact3_phone=strip_or_none("איש קשר - 3 -פלאפון"),
                    contact3_email=strip_or_none("איש קשר - 3 - מייל "),
                    marriage_date_ivry=strip_or_none(
                        "יום נישואין תאריך עברי - יום")
                    + "' "
                    + strip_or_none("יום נישואין תאריך עברי - חודש"),
                    # marriage_date=column_value("יום נישואין תאריך לועזי"),
                    # Temporary, until we fix front-end to handle a null value for this
                    marriage_date="2000-01-01",
                    institution_mahzor=strip_or_none("מחזור במכינה/ישיבה"),
                    teacher_grade_a=strip_or_none('ר"מ שנה א'),
                    teacher_grade_a_phone=strip_or_none("פלאפון ר״מ שנה א"),
                    teacher_grade_b=strip_or_none('ר"מ שנה ב'),
                    teacher_grade_b_phone=strip_or_none("פלאפון ר״מ שנה ב"),
                    paying=column_value("משלם או לא") == "משלם",
                    spirit_status=strip_or_none("מצב״ר - מצב רוחני"),
                    high_school_name=strip_or_none("שם תיכון"),
                    high_school_teacher=strip_or_none("שם מחנך תיכון"),
                    high_school_teacher_phone=strip_or_none(
                        "פלאפון מחנך תיכון"),
                    workstatus=strip_or_none("עיסוק"),
                    workplace=strip_or_none("מקום עבודה"),
                    educationfaculty=strip_or_none("פקולטה"),
                    workoccupation=strip_or_none("סוג עבודה"),
                    army_role=strip_or_none("סוג שירות"),  # סיירות
                    unit_name=strip_or_none("חטיבה - שיוך יחידתי"),
                    militaryPositionNew=strip_or_none("תפקיד נוכחי"),
                    militaryPositionOld=strip_or_none("תפקיד קודם"),
                    recruitment_date=column_value("תאריך גיוס"),
                    release_date=column_value("תאריך שחרור"),
                    accompany_id=column_value("מלווה אחראי - טלפון"),
                    # worktype=worktype,
                )
                db.session.add(apprentice)
                print("commted", phone)
                commited_ids.append(phone)
            except Exception as e:
                print(e)
                uncommited_ids.append(phone)
    except Exception as e:
        return (
            jsonify({"result": "Error while inserting " + str(e)}),
            HTTPStatus.BAD_REQUEST,
        )
    db.session.commit()
    return jsonify(
        {
            "result": "success",
            "uncommited_ids": [x for x in uncommited_ids if x is not None],
        }
    )


@apprentice_profile_form_blueprint.route("/maps_apprentices", methods=["GET"])
def maps_apprentices():
    try:
        return maps(request.args.get("userId"))
    except Exception as e:
        raise e
