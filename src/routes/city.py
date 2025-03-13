import uuid

from flask import Blueprint, request, jsonify
from http import HTTPStatus

from openpyxl.reader.excel import load_workbook

from src.services import db
from src.models.city_model import City

city_blueprint = Blueprint("cities", __name__, url_prefix="/cities")

base_dir = ""


@city_blueprint.route("", methods=["get"])
def get_all():
    try:

        CityList = db.session.query(City).all()
        if CityList:
            return [
                {"id": str(row.id), "name": row.name, "cluster_id": row.region_id}
                for row in CityList
            ]
        return jsonify({"result": "error"}), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST


@city_blueprint.route("/add", methods=["POST"])
def add():
    try:

        data = request.json
        name = data["name"]
        region_id = data["cluster"]
        city = City(
            id=str(uuid.uuid1().int)[
                :5
            ],  # if ent_group_name!="" else str(uuid.uuid1().int)[:5],
            name=name,
            region_id=region_id,
        )
        db.session.add(city)
        db.session.commit()
        return jsonify({"result": "success"}), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST


@city_blueprint.route("/upload_CitiesDB", methods=["PUT"])
def upload_cities():
    try:
        file = request.files["file"]
        wb = load_workbook(file)
        sheet = wb.active
        my_list = []
        for row in sheet.iter_rows(min_row=2):
            my_list.append(City(row[2].value, row[1].value.strip(), row[0].value))
        for ent in my_list:
            db.session.add(ent)
        try:
            db.session.commit()
            return jsonify({"result": "success"}), HTTPStatus.OK
        except Exception as e:
            return jsonify({"result": str(e)}), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.OK
