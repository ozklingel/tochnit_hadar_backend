import csv
import datetime
import uuid

import boto3
from openpyxl.reader.excel import load_workbook
from flask import Blueprint, request, jsonify, send_file
from http import HTTPStatus
from datetime import datetime, date, timedelta

from sqlalchemy import func, or_

import config
from src.routes.user_profile import correct_auth
from src.services import db
from config import AWS_access_key_id, AWS_secret_access_key, melave_Score, visitcalls_melave_avg, visitmeets_melave_avg, \
    proffesionalMeet_presence, forgotenApprentice_cnt, cenes_presence, horim_meeting, call_report, groupMeet_report, \
    personalMeet_report, professional_report, HorimCall_report
from src.models.apprentice_model import Apprentice
from src.models.base_model import Base
from src.models.city_model import City
from src.models.gift_model import Gift
from src.models.institution_model import Institution
from src.models.system_report_model import SystemReport
from src.models.user_model import User
from src.models.report_model import Report
import src.routes.madadim as md
from src.routes.homepage import get_Eshcol_corrdintors_score, get_mosad_Coordinators_score, get_melave_score

export_import_blueprint = Blueprint('export_import', __name__, url_prefix='/export_import')
base_dir = ""  # "/home/ubuntu/flaskapp/"


@export_import_blueprint.route("lowScoreApprentice_mosad", methods=['post'])
def import_lowScoreApprentice_mosad(type="extenal"):
    try:
        if type=="extenal" and correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        export_date = request.args.get("export_date")
        if export_date != str(date.today()) and export_date is not None:
            data = db.session.query(SystemReport.value).filter(SystemReport.type == "lowScoreApprentice_mosad",
                                                                SystemReport.creation_date == export_date).first()
            if data is None:
                return jsonify({'result': "no such export was done"}), HTTPStatus.BAD_REQUEST
            return data.value
        lowScoreApprentice_dict = md.lowScoreApprentice(False)[0].json
        missingCalleApprentice_dict = md.missingCalleApprentice(False)[0].json
        missingMeetingApprentice_dict = md.missingMeetingApprentice(False)[0].json
        res = []
        lowScoreApprentice_List = lowScoreApprentice_dict["lowScoreApprentice_List"]
        callApprentice_List = missingCalleApprentice_dict["missingCalleApprentice_count"]
        meetingApprentice_List = missingMeetingApprentice_dict["missingmeetApprentice_count"]

        for i in range(0, len(lowScoreApprentice_List)):
            score_ent = lowScoreApprentice_List[i]
            call_ent = callApprentice_List[i]
            meet_ent = meetingApprentice_List[i]
            all_Apprentices_cnt = db.session.query(func.count(Apprentice.id)).filter(
                Apprentice.institution_id == Institution.id, Institution.name == score_ent["name"]).first()
            res.append(
                [score_ent["name"], all_Apprentices_cnt[0], score_ent["value"], call_ent["value"], meet_ent["value"]])
        fields = ['callgap_count', 'meetgap_count', 'lowScore_count', 'apprentice_count', 'Mosad_name']
        with open(base_dir + 'system_export', 'w', encoding="utf-8") as f:
            write = csv.writer(f)
            write.writerow(fields)
            write.writerows(res)
        if type == "local":
            with open(base_dir + 'system_export', 'r', encoding="utf-8") as file:
                return file.read()
        return send_file(base_dir + "system_export", as_attachment=True, download_name="dict2.csv")
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@export_import_blueprint.route("lowScoreApprentice_tohnit", methods=['post'])
def import_lowScoreApprentice_tohnit(type="extenal"):
    try:
        if type=="extenal" and correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        export_date = request.args.get("export_date")
        if export_date != str(date.today()) and export_date is not None:
            data = db.session.query(SystemReport.value).filter(SystemReport.type == "lowScoreApprentice_tohnit",
                                                                SystemReport.creation_date == export_date).first()
            if data is None:
                return jsonify({'result': "no such export was done"}), HTTPStatus.BAD_REQUEST

            return data.value
        lowScoreApprentice_dict = md.lowScoreApprentice()[0].json
        missingCalleApprentice_dict = md.missingCalleApprentice()[0].json
        missingMeetingApprentice_dict = md.missingMeetingApprentice()[0].json
        res = []
        lowScoreApprentice_List = lowScoreApprentice_dict["lowScoreApprentice_List"]
        callApprentice_List = missingCalleApprentice_dict["missingCalleApprentice_count"]
        meetingApprentice_List = missingMeetingApprentice_dict["missingmeetApprentice_count"]

        for i in range(0, len(lowScoreApprentice_List)):
            score_ent = lowScoreApprentice_List[i]
            call_ent = callApprentice_List[i]
            meet_ent = meetingApprentice_List[i]
            all_Apprentices_cnt = db.session.query(func.count(Apprentice.id)).filter(
                Apprentice.institution_id == Institution.id, Institution.name == score_ent["name"]).first()
            res.append(
                [score_ent["name"], all_Apprentices_cnt[0], score_ent["value"], call_ent["value"], meet_ent["value"]])
        fields = ['callgap_count', 'meetgap_count', 'lowScore_count', 'apprentice_count', 'Mosad_name']
        with open(base_dir + 'system_export', 'w', encoding="utf-8") as f:
            write = csv.writer(f)
            write.writerow(fields)
            write.writerows(res)
        if type == "local":
            with open(base_dir + 'system_export', 'r', encoding="utf-8") as file:
                return file.read()
        return send_file(base_dir + "system_export", as_attachment=True, download_name="dict2.csv")
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@export_import_blueprint.route("/melave_corrdintors_score", methods=['post'])
def import_melave_corrdintors_score(type="extenal"):
    try:
        if type=="extenal" and correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        export_date = request.args.get("export_date")
        if export_date != str(date.today()) and export_date is not None:
            data = db.session.query(SystemReport.value).filter(SystemReport.type == "melave_corrdintors_score",
                                                                SystemReport.creation_date == export_date).first()
            if data is None:
                return jsonify({'result': "no such export was done"}), HTTPStatus.BAD_REQUEST
            return data.value
        score_dict, score_melaveProfile = get_melave_score()
        rows = score_melaveProfile
        for dict in rows:
            user_mosad = db.session.query(User.name, Institution.name).filter(
                User.id == dict["melaveId"], Institution.id == User.institution_id).first()
            if user_mosad is None:
                continue
            dict["name"] = user_mosad[0]
            dict["mosad"] = user_mosad[1]
        with open(base_dir + 'system_export', 'w', encoding="utf-8") as f:
            title = "melave_score1,melaveId,mosad,name".split(",")  # quick hack
            cw = csv.DictWriter(f, title, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
            cw.writeheader()
            cw.writerows(rows)
        if type == "local":
            with open(base_dir + 'system_export', 'r', encoding="utf-8") as file:
                return file.read()
        return send_file(base_dir + "system_export", as_attachment=True, download_name="dict2.csv")
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@export_import_blueprint.route("/mosad_melavim_cnt", methods=['post'])
def import_mosad_melavim_cnt(type="extenal"):
    try:
        if type=="extenal" and correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        export_date = request.args.get("export_date")
        if export_date != str(date.today()) and export_date is not None:
            data = db.session.query(SystemReport.value).filter(SystemReport.type == "mosad_melavim_cnt",
                                                                SystemReport.creation_date == export_date).first()
            return data.value
        mosad_melavim_cnt = db.session.query(Institution.name, func.count(User.name)).filter(
            User.institution_id == Institution.id, User.role_ids.contains("0")).group_by(
            Institution.id).all()
        mosad_melavim_cnt = dict(mosad_melavim_cnt)
        rows = mosad_melavim_cnt.items()
        fields = ['count', 'name']

        with open(base_dir + 'system_export', 'w', encoding="utf-8") as f:
            write = csv.writer(f)
            write.writerow(fields)
            write.writerows(rows)
        if type == "local":
            with open(base_dir + 'system_export', 'r', encoding="utf-8") as file:
                return file.read()
        return send_file(base_dir + "system_export", as_attachment=True, download_name="dict2.csv")
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@export_import_blueprint.route("/mosad_corrdintors_score", methods=['post'])
def import_mosad_corrdintors_score(type="extenal"):
    try:
        if type=="extenal" and correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        export_date = request.args.get("export_date")
        if export_date != str(date.today()) and export_date is not None:
            data = db.session.query(SystemReport.value).filter(SystemReport.type == "mosad_corrdintors_score",
                                                                SystemReport.creation_date == export_date).first()
            if data is None:
                return jsonify({'result': "no such export was done"}), HTTPStatus.BAD_REQUEST

            return data.value
        score_dict = get_mosad_Coordinators_score()
        rows = score_dict[1]
        for dict in rows:
            user_name = db.session.query(User.name).filter(
                User.id == dict["id"]).first()
            dict["name"] = user_name[0]
        with open(base_dir + 'system_export', 'w', encoding="utf-8") as f:
            title = "score,id,name".split(",")  # quick hack
            cw = csv.DictWriter(f, title, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
            cw.writeheader()
            cw.writerows(rows)
        if type == "local":
            with open(base_dir + 'system_export', 'r', encoding="utf-8") as file:
                return file.read()
        return send_file(base_dir + "system_export", as_attachment=True, download_name="dict2.csv")
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@export_import_blueprint.route("/Eshcol_corrdintors_score", methods=['post'])
def import_Eshcol_corrdintors_score(type="extenal"):
    try:
        if type=="extenal" and correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        export_date = request.args.get("export_date")
        if export_date != str(date.today()) and export_date is not None:
            data = db.session.query(SystemReport.value).filter(SystemReport.type == "Eshcol_corrdintors_score",
                                                                SystemReport.creation_date == export_date).first()
            if data is None:
                return jsonify({'result': "no such export was done"}), HTTPStatus.BAD_REQUEST

            return data.value
        score_dict = get_Eshcol_corrdintors_score()
        rows = score_dict[1]
        for dict in rows:
            user_name = db.session.query(User.name).filter(
                User.id == dict["id"]).first()
            dict["name"] = user_name[0]
        with open(base_dir + 'system_export', 'w', encoding="utf-8") as f:
            title = "score,id,name".split(",")  # quick hack
            cw = csv.DictWriter(f, title, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
            cw.writeheader()
            cw.writerows(rows)
        if type == "local":
            with open(base_dir + 'system_export', 'r', encoding="utf-8") as file:
                return file.read()
        return send_file(base_dir + "system_export", as_attachment=True, download_name="dict2.csv")
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@export_import_blueprint.route("/forgoten_Tohnit", methods=['post'])
def import_forgoten_Tohnit(type="extenal"):
    try:
        if type=="extenal" and correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        export_date = request.args.get("export_date")
        if export_date != str(date.today()) and export_date is not None:
            data = db.session.query(SystemReport.value).filter(SystemReport.type == "forgoten_Tohnit",
                                                                SystemReport.creation_date == export_date).first()
            if data is None:
                return jsonify({'result': "no such export was done"}), HTTPStatus.BAD_REQUEST

            return data.value
        all_Apprentices = db.session.query(Apprentice.id, Institution.name).filter(
            Apprentice.institution_id == Institution.id).all()
        # update apprentices meet
        visitcalls = db.session.query(Report.ent_reported, func.max(Report.visit_date).label("visit_date"),
                                      Institution.name).filter(Apprentice.id == Report.ent_reported,
                                                               Institution.id == Apprentice.institution_id,
                                                               Report.title.in_(config.reports_as_call)).group_by(
            Report.ent_reported,
            Institution.name).all()
        ids_have_visit = [r[0] for r in visitcalls]
        ids_no_visit = []
        # handle no record
        for ent in all_Apprentices:
            if ent.id not in ids_have_visit:
                ids_no_visit.append([ent[0], ent[1]])
        counts = dict()
        forgotenApprentice_total = 0
        for i in visitcalls:
            vIsDate = i.visit_date
            now = date.today()
            gap = (now - vIsDate).days if vIsDate is not None else 0
            if gap > 100:
                forgotenApprentice_total += 1
                counts[i[2]] = counts.get(i[2], 0) + 1
        for i in ids_no_visit:
            forgotenApprentice_total += 1
            counts[i[1]] = counts.get(i[1], 0) + 1
        fields = ['count', 'name']
        rows = counts.items()
        with open(base_dir + 'system_export', 'w', encoding="utf-8") as f:
            write = csv.writer(f)
            write.writerow(fields)
            write.writerows(rows)
        if type == "local":
            with open(base_dir + 'system_export', 'r', encoding="utf-8") as file:
                return file.read()
        return send_file(base_dir + "system_export", as_attachment=True, download_name="dict2.csv")
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@export_import_blueprint.route("/forgoten_mosad", methods=['post'])
def import_forgoten_mosad(type="extenal"):
    try:
        if type=="extenal" and correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        export_date = request.args.get("export_date")
        if export_date != str(date.today()) and export_date is not None:
            data = db.session.query(SystemReport.value).filter(SystemReport.type == "forgoten_mosad",
                                                                SystemReport.creation_date == export_date).first()
            if data is None:
                return jsonify({'result': "no such export was done"}), HTTPStatus.BAD_REQUEST

            return data.value
        all_Apprentices = db.session.query(Apprentice.id, Apprentice.name, Institution.name).filter(
            Apprentice.institution_id == Institution.id).all()
        # update apprentices meet
        visitcalls = db.session.query(Report.ent_reported, func.max(Report.visit_date).label("visit_date"),
                                      Institution.name).filter(Apprentice.id == Report.ent_reported,
                                                               Institution.id == Apprentice.institution_id,
                                                               Report.title.in_(config.reports_as_call)).group_by(
            Report.ent_reported,
            Institution.name).all()
        ids_have_visit = [r[0] for r in visitcalls]
        ids_no_visit = []
        # handle no record
        for ent in all_Apprentices:
            if ent.id not in ids_have_visit:
                ids_no_visit.append([ent[0], ent[1], 100])
        for i in visitcalls:
            vIsDate = i.visit_date
            now = date.today()
            gap = (now - vIsDate).days if vIsDate is not None else 0
            ids_no_visit.append([i, ent[1], gap])
        fields = ['id', 'gap', 'Name']
        rows = ids_no_visit
        with open(base_dir + 'system_export', 'w', encoding="utf-8") as f:
            write = csv.writer(f)
            write.writerow(fields)
            write.writerows(rows)
        if type == "local":
            with open(base_dir + 'system_export', 'r', encoding="utf-8") as file:
                return file.read()
        return send_file(base_dir + "system_export", as_attachment=True, download_name="dict2.csv")
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@export_import_blueprint.route('/upload_CitiesDB', methods=['PUT'])
def upload_CitiesDB():
    try:
        import csv
        my_list = []
        # /home/ubuntu/flaskapp/
        with open(base_dir + 'data/cities_add.csv', 'r', encoding="utf8") as f:
            reader = csv.reader(f)
            for row in reader:
                my_list.append(City(row[0].strip(), row[1].strip(), row[2].strip()))
        for ent in my_list:
            db.session.add(ent)
        try:
            db.session.commit()
            return jsonify({"result": "success"}), HTTPStatus.OK
        except Exception as e:
            return jsonify({"result": str(e)}), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.OK


@export_import_blueprint.route('/upload_baseDB', methods=['PUT'])
def upload_baseDB():
    try:
        import csv
        my_list = []
        # /home/ubuntu/flaskapp/
        with open(base_dir + 'data/base_add.csv', 'r', encoding="utf8") as f:
            reader = csv.reader(f)
            for row in reader:
                ent = Base(int(str(uuid.uuid4().int)[:5]), row[0].strip(), row[1].strip())
                db.session.add(ent)
        db.session.commit()
        return jsonify({"result": "success"}), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.OK


@export_import_blueprint.route("/export_dict", methods=['post'])
def ZNOTINUSE_export_dict():
    try:
        data = request.json
        to_csv = data[list(data.keys())[0]]
        keys = to_csv[0].keys()
        with open('to_csv.csv', 'w', newline='', encoding="utf-8") as output_file:
            dict_writer = csv.DictWriter(output_file, keys)
            dict_writer.writeheader()
            dict_writer.writerows(to_csv)
        return send_file("to_csv.csv", as_attachment=True, download_name="dict2.csv")
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@export_import_blueprint.route("/add_giftCode_excel", methods=['put'])
def add_giftCode_excel():
    try:
        file = request.files['file']

        wb = load_workbook(file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2):
            code = row[0].value
            was_used = row[1].value

            gift1 = Gift(code=code, was_used=was_used)
            db.session.add(gift1)

        try:
            db.session.commit()
        except Exception as e:
            return jsonify({'result': 'error while inserting' + str(e)}), HTTPStatus.OK
        return jsonify({'result': 'success'}), HTTPStatus.BAD_REQUEST
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST



@export_import_blueprint.route('/monthly', methods=['GET'])
def monthly():
    try:
        # export_lowScoreApprentice_mosad
        export_lowScoreApprentice_mosad_csv = import_lowScoreApprentice_mosad("local")
        system_report1 = SystemReport(
            id=int(str(uuid.uuid4().int)[:5]),
            related_id=0,
            type="lowScoreApprentice_mosad",
            value=export_lowScoreApprentice_mosad_csv,
            creation_date=date.today(),
        )
        db.session.add(system_report1)

        # export_Eshcol_corrdintors_score
        export_Eshcol_corrdintors_score_csv = import_Eshcol_corrdintors_score("local")
        system_report1 = SystemReport(
            id=int(str(uuid.uuid4().int)[:5]),
            related_id=0,
            type="Eshcol_corrdintors_score",
            value=export_Eshcol_corrdintors_score_csv,
            creation_date=date.today(),
        )
        db.session.add(system_report1)

        # export_forgoten_mosad
        export_forgoten_mosad_csv = import_forgoten_mosad("local")
        system_report1 = SystemReport(
            id=int(str(uuid.uuid4().int)[:5]),
            related_id=0,
            type="forgoten_mosad",
            value=export_forgoten_mosad_csv,
            creation_date=date.today(),
        )
        db.session.add(system_report1)

        # export_forgoten_Tohnit
        export_forgoten_Tohnit_csv = import_forgoten_Tohnit("local")
        system_report1 = SystemReport(
            id=int(str(uuid.uuid4().int)[:5]),
            related_id=0,
            type="forgoten_Tohnit",
            value=export_forgoten_Tohnit_csv,
            creation_date=date.today(),
        )
        db.session.add(system_report1)
        # export_melave_corrdintors_score
        export_melave_corrdintors_score_csv = import_melave_corrdintors_score("local")
        system_report1 = SystemReport(
            id=int(str(uuid.uuid4().int)[:5]),
            related_id=0,
            type="melave_corrdintors_score",
            value=export_melave_corrdintors_score_csv,
            creation_date=date.today(),
        )
        db.session.add(system_report1)

        # export_lowScoreApprentice_tohnit_csv
        export_mosad_melavim_cnt_csv = import_mosad_melavim_cnt("local")
        system_report1 = SystemReport(
            id=int(str(uuid.uuid4().int)[:5]),
            related_id=0,
            type="mosad_melavim_cnt",
            value=export_mosad_melavim_cnt_csv,
            creation_date=date.today(),
        )
        db.session.add(system_report1)

        # export_mosad_corrdintors_score
        export_mosad_corrdintors_score_csv = import_mosad_corrdintors_score("local")
        system_report1 = SystemReport(
            id=int(str(uuid.uuid4().int)[:5]),
            related_id=0,
            type="mosad_corrdintors_score",
            value=export_mosad_corrdintors_score_csv,
            creation_date=date.today(),
        )
        db.session.add(system_report1)

        all_melave = db.session.query(User.id, User.name, User.institution_id).filter(
            User.role_ids.contains("0")).all()
        for melave in all_melave:
            melaveId = melave[0]
            all_melave_Apprentices = db.session.query(Apprentice.id).filter(
                Apprentice.accompany_id == melaveId).all()
            melave_score1, call_gap_avg, meet_gap_avg,group_meeting_gap_avg = md.melave_score(melaveId)
            system_report1 = SystemReport(
                id=int(str(uuid.uuid4().int)[:5]),
                related_id=melaveId,
                type=melave_Score,
                value=melave_score1,
                creation_date=date.today(),
            )
            db.session.add(system_report1)
            system_report1 = SystemReport(
                id=int(str(uuid.uuid4().int)[:5]),
                related_id=melaveId,
                type=visitcalls_melave_avg,
                value=call_gap_avg,
                creation_date=date.today(),
            )
            db.session.add(system_report1)
            system_report1 = SystemReport(
                id=int(str(uuid.uuid4().int)[:5]),
                related_id=melaveId,
                type=visitmeets_melave_avg,
                value=meet_gap_avg,
                creation_date=date.today(),
            )
            db.session.add(system_report1)
            # mosad Madadim:
            all_MosadCoordinator = db.session.query(User.id, User.institution_id).filter(
                User.role_ids.contains("1")).all()
            for mosadCoord in all_MosadCoordinator:
                mosadCoord_id = mosadCoord[0]
                res = md.mosadCoordinator(mosadCoord_id,False)[0].json

                system_report1 = SystemReport(
                    id=int(str(uuid.uuid4().int)[:5]),
                    related_id=mosadCoord_id,
                    type=visitcalls_melave_avg,
                    value=res['avg_apprenticeCall_gap'],
                    creation_date=date.today(),
                )
                db.session.add(system_report1)
                system_report1 = SystemReport(
                    id=int(str(uuid.uuid4().int)[:5]),
                    related_id=mosadCoord_id,
                    type=visitmeets_melave_avg,
                    value=res['avg_apprenticeMeeting_gap'],
                    creation_date=date.today(),
                )
                db.session.add(system_report1)

        try:
            db.session.commit()
            return jsonify({'result': 'success'}), HTTPStatus.OK

        except Exception as e:
            return jsonify({'result': 'error' + str(e)}), HTTPStatus.BAD_REQUEST
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@export_import_blueprint.route('/rivony', methods=['GET'])
def rivony():
    current_month = date.today().month
    month4index = current_month % 3
    start_Of_Rivon = datetime.today() - timedelta(days=30 * month4index)
    # melave Madadim:
    all_melave = db.session.query(User.id, User.name, User.institution_id).filter(User.role_ids.contains("0")).all()
    for melave in all_melave:
        melaveId = melave[0]
        all_melave_Apprentices = db.session.query(Apprentice.id).filter(
            Apprentice.accompany_id == melaveId).all()
        if len(all_melave_Apprentices) == 0:
            continue
        # מפגש מקצועי מלווה
        newvisit_professional = db.session.query(Report.user_id).filter(Report.ent_reported == melaveId,
                                                                       Report.title == professional_report,
                                                                       Report.visit_date > start_Of_Rivon).all()
        system_report1 = SystemReport(
            id=int(str(uuid.uuid4().int)[:5]),
            related_id=melaveId,
            type=proffesionalMeet_presence,
            value=len(newvisit_professional),
            creation_date=date.today(),
        )
        db.session.add(system_report1)
        Apprentice_ids_forgoten = [r[0] for r in all_melave_Apprentices]
        too_old = datetime.today() - timedelta(days=100)
        Oldvisitcalls = db.session.query(Report.ent_reported).filter(
            Apprentice.id == Report.ent_reported,
            or_(Report.title == call_report, Report.title == groupMeet_report, Report.title == personalMeet_report),
            Report.visit_date < too_old).all()
        for i in Oldvisitcalls:
            if i[0] in Apprentice_ids_forgoten:
                Apprentice_ids_forgoten.remove(i[0])

        system_report1 = SystemReport(
            id=int(str(uuid.uuid4().int)[:5]),
            related_id=melaveId,
            type=forgotenApprentice_cnt,
            value=len(Apprentice_ids_forgoten),
            creation_date=date.today(),
        )
        db.session.add(system_report1)

    # mosad Madadim:
    all_MosadCoordinator = db.session.query(User.id, User.institution_id).filter(User.role_ids.contains("1")).all()
    for mosadCoord in all_MosadCoordinator:
        mosadCoord_id = mosadCoord[0]
        inst = db.session.query(User.institution_id).filter(User.id == mosadCoord_id)
        all_Apprentices = db.session.query(Apprentice.id).filter(
            Apprentice.institution_id == inst).all()
        Apprentice_ids_forgoten = [r[0] for r in all_Apprentices]

        too_old = datetime.today() - timedelta(days=100)
        Oldvisitcalls = db.session.query(Report.ent_reported).filter(
            Apprentice.id == Report.ent_reported,
            or_(Report.title == call_report, Report.title == groupMeet_report, Report.title == personalMeet_report),
            Report.visit_date < too_old).all()
        for i in Oldvisitcalls:
            if i[0] in Apprentice_ids_forgoten:
                Apprentice_ids_forgoten.remove(i[0])

        system_report1 = SystemReport(
            id=int(str(uuid.uuid4().int)[:5]),
            related_id=mosadCoord_id,
            type=forgotenApprentice_cnt,
            value=len(Apprentice_ids_forgoten),
            creation_date=date.today(),
        )
        db.session.add(system_report1)

    # eshcol Madadim:
    all_eshcolCoordinator = db.session.query(User.id, User.eshcol).filter(User.role_ids.contains("2")).all()
    for eshcolCoord in all_eshcolCoordinator:
        eshcolCoord_id = eshcolCoord[0]
        eshco = eshcolCoord[1]
        all_Apprentices = db.session.query(Apprentice.id).filter(
            Apprentice.eshcol == eshco).all()
        Apprentice_ids_forgoten = [r[0] for r in all_Apprentices]
        too_old = datetime.today() - timedelta(days=100)
        Oldvisitcalls = db.session.query(Report.ent_reported).filter(
            Apprentice.id == Report.ent_reported,
            or_(Report.title == call_report, Report.title == groupMeet_report, Report.title == personalMeet_report),
            Report.visit_date < too_old).all()
        for i in Oldvisitcalls:
            if i[0] in Apprentice_ids_forgoten:
                Apprentice_ids_forgoten.remove(i[0])

        system_report1 = SystemReport(
            id=int(str(uuid.uuid4().int)[:5]),
            related_id=eshcolCoord_id,
            type=forgotenApprentice_cnt,
            value=len(Apprentice_ids_forgoten),
            creation_date=date.today(),
        )
        db.session.add(system_report1)

    try:
        db.session.commit()
        return jsonify({'result': 'success'}), HTTPStatus.OK

    except Exception as e:
        return jsonify({'result': 'error' + str(e)}), HTTPStatus.BAD_REQUEST


@export_import_blueprint.route('/yearly', methods=['GET'])
def yearly():
    current_month = date.today().month
    start_Of_year = datetime.today() - timedelta(days=30 * current_month)
    all_Apprentices = db.session.query(Apprentice.spirit_status, Apprentice.id).all()
    for apprentice1 in all_Apprentices:
        system_report1 = SystemReport(
            id=int(str(uuid.uuid4().int)[:5]),
            related_id=apprentice1.id,
            type="spirit_status",
            value=apprentice1.spirit_status,
            creation_date=date.today(),
        )
        db.session.add(system_report1)
    all_melave = db.session.query(User.id, User.name, User.institution_id).filter(User.role_ids.contains("0")).all()
    for melave in all_melave:
        melaveId = melave[0]
        all_melave_Apprentices = db.session.query(Apprentice.id).filter(
            Apprentice.accompany_id == melaveId).all()
        if len(all_melave_Apprentices) == 0:
            continue

        cenes_yearly = db.session.query(Report.user_id, func.max(Report.visit_date).label("visit_date")).group_by(
            Report.user_id).filter(Report.title == "כנס_שנתי", Report.user_id == melaveId,
                                  Report.visit_date > start_Of_year).first()
        if cenes_yearly:
            system_report1 = SystemReport(
                id=int(str(uuid.uuid4().int)[:5]),
                related_id=melaveId,
                type=cenes_presence,
                value=100,
                creation_date=date.today(),
            )
            db.session.add(system_report1)
        Horim_meeting = db.session.query(Report.ent_reported, func.max(Report.visit_date).label("visit_date")).group_by(
            Report.ent_reported).filter(Report.title == HorimCall_report, Report.user_id == melaveId,
                                       Report.visit_date > start_Of_year).all()
        if Horim_meeting:
            system_report1 = SystemReport(
                id=int(str(uuid.uuid4().int)[:5]),
                related_id=melaveId,
                type=horim_meeting,
                value=Horim_meeting,
                creation_date=date.today(),
            )
            db.session.add(system_report1)
        too_old = datetime.today() - timedelta(days=365)
        base_meeting = db.session.query(Report.visit_date).distinct(Report.visit_date).filter(
            or_(Report.title == personalMeet_report, Report.title == groupMeet_report),
            Report.visit_in_army == True,
            Report.visit_date > too_old,
            Report.user_id == melaveId).group_by(
            Report.visit_date).count()
        base_meeting_score = 0
        if base_meeting > 2:
            base_meeting_score += 10
            system_report1 = SystemReport(
                id=int(str(uuid.uuid4().int)[:5]),
                related_id=melaveId,
                type="base_meeting",
                value=2,
                creation_date=date.today(),
            )
            db.session.add(system_report1)
    try:
        db.session.commit()
        return jsonify({'result': 'success'}), HTTPStatus.OK

    except Exception as e:
        return jsonify({'result': 'error' + str(e)}), HTTPStatus.BAD_REQUEST


@export_import_blueprint.route('/uploadfile', methods=['post'])
def uploadfile():
    try:
        if correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        images_list = []
        for imagefile in request.files.getlist('file'):
            new_filename = uuid.uuid4().hex + '.' + imagefile.filename.rsplit('.', 1)[1].lower()
            bucket_name = "th01-s3"
            session = boto3.Session()
            s3_client = session.client('s3',
                                       aws_access_key_id=AWS_access_key_id,
                                       aws_secret_access_key=AWS_secret_access_key)
            s3 = boto3.resource('s3',
                                aws_access_key_id=AWS_access_key_id,
                                aws_secret_access_key=AWS_secret_access_key)
            try:
                s3_client.upload_fileobj(imagefile, bucket_name, new_filename)
            except:
                return jsonify({'result': 'faild', 'image path': new_filename}), HTTPStatus.OK
            images_list.append("https://th01-s3.s3.eu-north-1.amazonaws.com/" + new_filename)
        # if updatedEnt:
        #    updatedEnt.attachments=images_list
        #    db.session.commit()
        return jsonify({'result': 'success', 'image path': images_list}), HTTPStatus.OK
    except Exception:
        return jsonify({"result": str(Exception)}), HTTPStatus.BAD_REQUEST
