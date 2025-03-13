from openpyxl.reader.excel import load_workbook
from flask import Blueprint, request, jsonify
from http import HTTPStatus

from src.services import db
from src.models.gift_model import Gift
from src.models.task_model_v2 import Task

gift_blueprint = Blueprint("gift", __name__, url_prefix="/gift")


@gift_blueprint.route("/add_giftCode_excel", methods=["put"])
def add_gift_code_excel():
    try:

        file = request.files["file"]
        wb = load_workbook(file)
        sheet = wb.active
        not_commited = []
        index_ = 0
        for row in sheet.iter_rows(min_row=2):
            code = row[0].value
            if not code:
                not_commited.append(index_)
                index_ = index_ + 1
                continue
            was_used = False
            gift1 = Gift(code=code, was_used=was_used)
            db.session.add(gift1)
            try:
                db.session.commit()
            except Exception as e:
                print(index_, str(e))
            index_ = index_ + 1
        return (
            jsonify({"result": "success", "not_commited": not_commited}),
            HTTPStatus.OK,
        )
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST


@gift_blueprint.route("/getGift", methods=["GET"])
def get_gift():
    try:

        giftCode = db.session.query(Gift).filter(Gift.was_used == False).first()
        if not giftCode:
            # acount not found
            return jsonify({"result": "no code available"}), HTTPStatus.OK
        else:
            (
                db.session.query(Gift)
                .filter(
                    Gift.code == giftCode.code,
                )
                .delete()
            )
            db.session.commit()

            return jsonify({"result": str(giftCode.code)}), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": "error occured when reading code:"+str(e)}), HTTPStatus.BAD_REQUEST


@gift_blueprint.route("/delete", methods=["put"])
def delete():
    try:

        giftCode1 = request.args.get("giftCode")
        apprentice_id = request.args.get("")

        giftCode = db.session.query(Gift).filter(Gift.code == giftCode1).first()
        if giftCode is not None:
            giftCode.was_used = True
            (
                db.session.query(Task)
                .filter(
                    Task.subject == apprentice_id,
                    Task.event == "יומהולדת",
                )
                .delete()
            )

            db.session.commit()

        return jsonify({"result": "success"}), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST


@gift_blueprint.route("/deleteAll", methods=["put"])
def delete_all():
    try:

        db.session.query(Gift).delete()
        db.session.commit()

        return jsonify({"result": "success"}), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST


@gift_blueprint.route("/getGifts_cnt", methods=["get"])
def get_gifts_cnt():
    try:

        giftCodes_all = db.session.query(Gift).all()
        giftCodes_used = db.session.query(Gift).filter(Gift.was_used == True).all()
        giftCodes_all_cnt = len(giftCodes_all) or 0
        giftCodes_used_cnt = len(giftCodes_used) or 0

        return (
            jsonify(f"מומשו {giftCodes_used_cnt} מתוך {giftCodes_all_cnt} קודי מתנה"),
            HTTPStatus.OK,
        )
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST
