import uuid
from random import randrange

import boto3
from flask import Blueprint, request, jsonify
from http import HTTPStatus

from openpyxl.reader.excel import load_workbook
import src.routes.madadim as md

from src.routes.user_Profile import correct_auth
from src.services import db, red
from config import AWS_secret_access_key, AWS_access_key_id
from src.models.apprentice_model import Apprentice
from src.models.city_model import City
from src.models.cluster_model import Cluster
from src.models.institution_model import Institution, front_end_dict
from src.models.user_model import user1
from src.routes.setEntityDetails_form_routes import validate_email
#import base64
#from io import BytesIO

#from matplotlib.figure import Figure


institutionProfile_form_blueprint = Blueprint('institutionProfile_form', __name__,
                                              url_prefix='/institutionProfile_form')




@institutionProfile_form_blueprint.route('/mosad_generalInfo', methods=['GET'])
def mosad_generalInfo():
    if correct_auth() == False:
        return jsonify({'result': f"wrong access token "}), HTTPStatus.OK
    institution_id = request.args.get('institution_id')
    all_Apprentices = db.session.query(Apprentice.paying, Apprentice.militaryPositionNew, Apprentice.spirit_status,
                                       Apprentice.army_role, Apprentice.institution_mahzor).filter(
        Apprentice.institution_id == institution_id).all()
    all_melaves = db.session.query(user1.id).filter(user1.institution_id == institution_id).all()
    coordinator = db.session.query(user1.id).filter(user1.institution_id == institution_id,
                                                    user1.role_ids.contains("1")).first()
    if coordinator is None:
        return jsonify({'result': "error-no coordinator or such institution"}), HTTPStatus.BAD_REQUEST

    paying_dict = dict()
    Picud_dict = dict()
    matzbar_dict = dict()
    sugSherut_dict = dict()
    mahzor_dict = dict()

    for apprentice1 in all_Apprentices:
        paying_dict[apprentice1.paying] = paying_dict.get(apprentice1.paying, 0) + 1
        Picud_dict[apprentice1.militaryPositionNew] = Picud_dict.get(apprentice1.militaryPositionNew, 0) + 1
        matzbar_dict[apprentice1.spirit_status] = matzbar_dict.get(apprentice1.spirit_status, 0) + 1
        sugSherut_dict[apprentice1.army_role] = sugSherut_dict.get(apprentice1.army_role, 0) + 1
        mahzor_dict[apprentice1.institution_mahzor] = mahzor_dict.get(apprentice1.institution_mahzor, 0) + 1
    mitztainim = []
    for melaveId in all_melaves:
        melave_score, call_gap_avg, personal_meet_gap_avg, group_meeting_gap_avg = md.melave_score(melaveId.id)
        if melave_score > 95:
            mitztainim.append(melaveId.id)
    Mosad_coord_score, visitprofessionalMeet_melave_avg, visitMatzbar_melave_avg, call_gap_avg, personal_meet_gap_avg, group_meeting_gap_avg = md.mosad_Coordinators_score(
        coordinator.id)
    mosad_score, forgoten_Apprentice_count = md.mosad_score(institution_id)
    resJson = md.mosadCoordinator(coordinator.id,False)
    mosadCoordinatorJson = resJson[0].json
    # fig = Figure()
    # ax = fig.subplots()
    # ax.plot([1, 2])
    # # Save it to a temporary buffer.
    # buf = BytesIO()
    # fig.savefig(buf, format="png")
    # # Embed the result in the html output.
    # data = base64.b64encode(buf.getbuffer()).decode("ascii")
    data="iVBORw0KGgoAAAANSUhEUgAAAoAAAAHgCAYAAAA10dzkAAAAOXRFWHRTb2Z0d2FyZQBNYXRwbG90bGliIHZlcnNpb24zLjkuMCwgaHR0cHM6Ly9tYXRwbG90bGliLm9yZy80BEi2AAAACXBIWXMAAA9hAAAPYQGoP6dpAABAy0lEQVR4nO3dZ3xUdd7+8WvSJgkpECAJgdCrlBQRBGQVRREUxUKLq7iW1TUUwYoNsYGKhRJ0iyvrfROqgAUEEUUEcVlIQu8JEEpCT88kmTn/B/4397KCEkjmZOZ83q/XPJjDOeTKz5Bz+ftOJjbDMAwBAADAMnzMDgAAAAD3ogACAABYDAUQAADAYiiAAAAAFkMBBAAAsBgKIAAAgMVQAAEAACyGAggAAGAxFEAAAACLoQACAABYDAUQAADAYiiAAAAAFkMBBAAAsBgKIAAAgMVQAAEAACyGAggAAGAxFEAAAACLoQACAABYDAUQAADAYiiAAAAAFkMBBAAAsBgKIAAAgMVQAAEAACyGAggAAGAxFEAAAACLoQACAABYDAUQAADAYiiAAAAAFkMBBAAAsBgKIAAAgMVQAAEAACyGAggAAGAxFEAAAACLoQACAABYDAUQAADAYiiAAAAAFkMBBAAAsBgKIAAAgMVQAAEAACyGAggAAGAxFEAAAACLoQACAABYDAUQAADAYiiAAAAAFkMBBAAAsBgKIAAAgMVQAAEAACyGAggAAGAxFEAAAACLoQACAABYDAUQAADAYiiAAAAAFkMBBAAAsBgKIAAAgMVQAAEAACyGAggAAGAxfmYH8GQul0tHjx5VaGiobDab2XEAAMBFMAxDBQUFiomJkY+PNffCKICX4ejRo4qNjTU7BgAAuATZ2dlq0qSJ2TFMQQG8DKGhoZJ+/gIKCwszOQ0AALgY+fn5io2NrbyPWxEF8DL8e+wbFhZGAQQAwMNY+eVb1hx8AwAAWBgFEAAAwGIogAAAABZDAQQAALAYCiAAAIDFUAABAAAshgIIAABgMRRAAAAAi6EAAgAAWIxXFMBJkybpqquuUmhoqCIjIzVo0CDt3r37N69bsGCB2rdvr8DAQHXu3FnLli1zQ1oAAABzeUUB/P7775WcnKyffvpJK1euVHl5uW666SYVFRVd8Joff/xRw4cP14MPPqj09HQNGjRIgwYN0rZt29yYHAAAwP1shmEYZoeobidOnFBkZKS+//57/e53vzvvOUOHDlVRUZG+/PLLymNXX3214uPj9eGHH17Ux8nPz1d4eLjy8vL4XcAAAFQzwzBq5Pf1cv/2kh3A/5aXlydJioiIuOA569evV9++fc851q9fP61fv/6C1zgcDuXn55/zAAAA1S/90Bn1n/qDDpy88DQPl87rCqDL5dLjjz+uXr16qVOnThc8LycnR1FRUecci4qKUk5OzgWvmTRpksLDwysfsbGx1ZYbAAD8vOv31zWZGvzheu3KKdDkr3aZHckreV0BTE5O1rZt2zR37txq/7vHjx+vvLy8ykd2dna1fwwAAKzqTFGZHvrHRr2+bKcqXIYGdI7WW4O7mB3LK/mZHaA6jRw5Ul9++aXWrFmjJk2a/Oq50dHRys3NPedYbm6uoqOjL3iN3W6X3W6vlqwAAOD/bDxwWqPnpOtoXqkC/Hz04q1X6Pfdm9bIawDhJTuAhmFo5MiRWrx4sb799lu1aNHiN6/p0aOHVq1adc6xlStXqkePHjUVEwAA/BeXy9DM1fs09C8/6WheqVo0qKPFj/XUvVc3o/zVIK/YAUxOTlZqaqo+++wzhYaGVr6OLzw8XEFBQZKk++67T40bN9akSZMkSWPGjNG1116rd955R7fccovmzp2rjRs36i9/+YtpnwcAAFZyqtChcfM36/s9JyRJt8fH6PU7OivE7hX1pFbzihX+4IMPJEnXXXfdOcc//vhj3X///ZKkQ4cOycfn/zY8e/bsqdTUVL3wwgt67rnn1KZNGy1ZsuRXf3AEAABUj39mntLouenKzXfI7uejibd11NCrYtn1cxOvfB9Ad+F9hAAAqBqny9DM7/bpvW/2yGVIrRrWUco9iWof7b77KPdvL9kBBAAAtd/xglKNnZehdftOSZLuSmyiVwd1VHAAdcTdWHEAAFDj1u07qTFzM3Sy0KEgf1+9OqiT7r7y19+xAzWHAggAAGqM02Vo6qq9mv7tXhmG1DYqRClJiWoTFWp2NEujAAIAgBqRm1+q0XPS9c+s05KkYVfFasLAjgoK8DU5GSiAAACg2n2/54TGzcvQqaIy1Qnw1Rt3dtbt8Y3NjoX/jwIIAACqTYXTpXdW7tEHq/dLkjo0ClNKUoJaNgwxORn+EwUQAABUi6NnSzR6Tro2HjwjSfr91U31wi1XKNCfkW9tQwEEAACX7dtduRo3f7POFpcrxO6nyXd11q1dYsyOhQugAAIAgEtW7nTp7RW79Zc1mZKkzo3DNSMpQc3q1zE5GX4NBRAAAFyS7NPFGjUnXRnZZyVJ9/dsrvED2svux8i3tqMAAgCAKluxPUdPLdis/NIKhQX66a2743Rzp2izY+EiUQABAMBFc1Q4NfmrXfp43QFJUlxsXc0YnqDYiGBzg6FKKIAAAOCiHDpVrOTUNG09kidJerh3Cz3Vr70C/HxMToaqogACAIDftGzrMT2zcIsKHBWqG+yvKXfHqe8VUWbHwiWiAAIAgAsqLXfq9aU79T8/HZQkXdmsnqYNT1DjukEmJ8PloAACAIDzyjpZpOTZadpxLF+S9Oi1rfTETW3l78vI19NRAAEAwC98lnFEzy3aqqIypyLqBOjdIXG6rl2k2bFQTSiAAACgUmm5UxO/2K45G7IlSd1aRGjasARFhweanAzViQIIAAAkSfuOFyp5dpp25xbIZpNG9mmtMTe0kR8jX69DAQQAAPp002G9sGSbSsqdahBi1/tD43VNmwZmx0INoQACAGBhxWUVeumz7Vq46bAkqWer+np/WLwiQxn5ejMKIAAAFrU7p0DJqWnad7xQPjZpzA1tNfL61vL1sZkdDTWMAggAgMUYhqH5G7M14fPtKi13KTLUrqnDEtSjVX2zo8FNKIAAAFhIoaNCLyzeqiUZRyVJvds00HtD49UgxG5yMrgTBRAAAIvYcTRfI1PTlHmySL4+No27sa3+dG0r+TDytRwKIAAAXs4wDM3+5yG98uUOlVW4FB0WqOlJCbqqeYTZ0WASCiAAAF6soLRczy7aqqVbjkmSrm8fqSmD4xRRJ8DkZDATBRAAAC+19XCeRs5J08FTxfLzsenpm9vpoWtaMvIFBRAAAG9jGIb+8eMBvbFsl8qcLjWuG6TpSQlKbFrP7GioJSiAAAB4kbzicj396Wat2J4rSbrxiihNuTtO4cH+JidDbUIBBADAS2Rkn9XI1DQdPlMif1+bxvfvoD/0ai6bjZEvzkUBBADAwxmGoY/WZmnyV7tU4TIUGxGkGcMTFRdb1+xoqKUogAAAeLCzxWV6csFmfbPzuCRpQOdoTb6ri8ICGfniwiiAAAB4qE0HT2tUarqO5pUqwNdHL97aQb+/uhkjX/wmCiAAAB7G5TL05zWZmvL1bjldhprXD9aMpER1ahxudjR4CAogAAAe5FShQ08s2KzVu09Ikm6Li9Ebd3ZWiJ1bOi4eXy0AAHiIf2ae0ui56crNd8ju56OXb+uoYVfFMvJFlVEAAQCo5ZwuQzO/26f3vtkjlyG1alhHKfckqn10mNnR4KEogAAA1GInChwaOy9Da/edlCTdmdhYr97eSXUY+eIy8NUDAEAt9eO+kxozL0MnChwK8vfVK7d31OCusWbHghegAAIAUMs4XYamrtqr6d/ulWFIbaNClJKUqDZRoWZHg5egAAIAUIvk5pdqzNx0/ZR5WpI0tGusXr6to4ICfE1OBm9CAQQAoJZYs+eExs7L0KmiMgUH+OqNOzprUEJjs2PBC1EAAQAwWYXTpXdX7tHM1fslSe2jQ5VyT6JaNQwxORm8FQUQAAATHcsr0eg56frXgTOSpHu6N9WLt16hQH9Gvqg5FEAAAEzy3a7jGjc/Q2eKyxVi99OkOztrYFyM2bFgARRAAADcrNzp0pQVu/XnNZmSpE6NwzRjeKKaN6hjcjJYBQUQAAA3OnymWKPmpCv90FlJ0v09m2v8gPay+zHyhftQAAEAcJOvt+foqYVblFdSrtBAP719dxfd3KmR2bFgQRRAAABqWFmFS5O+2qmP1x2QJMU1CdeMpETFRgSbGwyWRQEEAKAGHTpVrJFz0rTlcJ4k6aFrWujpm9srwM/H5GSwMgogAAA1ZNnWY3pm4RYVOCoUHuSvdwbHqe8VUWbHAiiAAABUt9Jyp15fulP/89NBSVJi07qanpSoxnWDTE4G/IwCCABANco6WaSRqWnafjRfkvTItS315E3t5O/LyBe1BwUQAIBq8vnmoxr/6RYVlTkVUSdA7wyJU592kWbHAn6BAggAwGUqLXdq4hc7NGfDIUlSt+YRmjY8QdHhgSYnA87Pa/aj16xZo4EDByomJkY2m01Lliz5zWtmz56tuLg4BQcHq1GjRnrggQd06tSpmg8LAPAa+44XalDKOs3ZcEg2mzTq+tZKfbg75Q+1mtcUwKKiIsXFxSklJeWizl+3bp3uu+8+Pfjgg9q+fbsWLFigDRs26OGHH67hpAAAb7Eo7bBum7FWu3IK1CAkQJ880E1P3NROfrzeD7Wc14yA+/fvr/79+1/0+evXr1fz5s01evRoSVKLFi30yCOP6M0336ypiAAAL1FcVqGXPtuuhZsOS5J6tKyvqcPiFRnGrh88g2X/F6VHjx7Kzs7WsmXLZBiGcnNztXDhQg0YMMDsaACAWmxPboFun7FOCzcdlo9NGtu3rf73oe6UP3gUr9kBrKpevXpp9uzZGjp0qEpLS1VRUaGBAwf+6gjZ4XDI4XBUPs/Pz3dHVABALWAYhhZsPKyXPt+m0nKXGobaNW1Ygnq0qm92NKDKLLsDuGPHDo0ZM0YvvfSSNm3apOXLl+vAgQN69NFHL3jNpEmTFB4eXvmIjY11Y2IAgFmKHBUaOy9DT3+6RaXlLvVu00BfjelN+YPHshmGYZgdorrZbDYtXrxYgwYNuuA59957r0pLS7VgwYLKY2vXrlXv3r119OhRNWrU6BfXnG8HMDY2Vnl5eQoLC6vWzwEAUDvsOJqvkalpyjxZJB+b9MRN7fSna1vJx8dmdjRcovz8fIWHh1v6/m3ZEXBxcbH8/M799H19fSX9vM1/Pna7XXa7vcazAQDMZxiGUjcc0sQvdqiswqXosEBNG56gbi0izI4GXDavKYCFhYXat29f5fOsrCxlZGQoIiJCTZs21fjx43XkyBF98sknkqSBAwfq4Ycf1gcffKB+/frp2LFjevzxx9WtWzfFxMSY9WkAAGqBgtJyjV+0VV9uOSZJ6tOuod4ZEq+IOgEmJwOqh9cUwI0bN6pPnz6Vz8eNGydJGjFihGbNmqVjx47p0KFDlX9+//33q6CgQDNmzNATTzyhunXr6vrrr+dtYADA4rYdyVNyapoOniqWn49NT/Vrp4d7t2TkC6/ila8BdBdeQwAA3sMwDH2y/qBeX7pTZU6XGtcN0rThCbqyWT2zo6Gacf/2oh1AAAAuVV5JuZ5ZuEXLt+dIkvp2iNKUwV1UN5iRL7wTBRAAYGkZ2Wc1MjVNh8+UyN/XpvH9O+gPvZrLZmPkC+9FAQQAWJJhGPpobZbeXL5L5U5DsRFBmjE8UXGxdc2OBtQ4CiAAwHLOFpfpyQWb9c3O45Kk/p2iNfmuLgoP8jc5GeAeFEAAgKVsOnhao1LTdTSvVAG+Pnrh1g669+pmjHxhKRRAAIAluFyG/vJDpt5esVtOl6Hm9YM1IylRnRqHmx0NcDsKIADA650qdOiJBZu1evcJSdLAuBi9cUcnhQYy8oU1UQABAF5tQ9ZpjZqTptx8h+x+PpowsKOGd4tl5AtLowACALySy2Vo5up9enflHrkMqWXDOkpJSlSHRtZ841/gP1EAAQBe50SBQ+PmZ+iHvSclSXcmNNargzqpjp3bHiBRAAEAXubHfSc1Zl6GThQ4FOjvo1du76TBVzZh5Av8BwogAMArOF2Gpq3aq2nf7pVhSG0iQzTznkS1iQo1OxpQ61AAAQAe73h+qUbPTddPmaclSUO6NtHE2zopKMDX5GRA7UQBBAB4tDV7TmjsvAydKipTcICvXr+jk+5IaGJ2LKBWowACADxShdOl977Zo5mr98swpPbRoZqRlKjWkSFmRwNqPQogAMDjHMsr0Zg5Gdpw4OeRb1L3pnrp1isU6M/IF7gYFEAAgEf5btdxjZufoTPF5Qqx++mNOzvrtrgYs2MBHoUCCADwCOVOl6as2K0/r8mUJHWMCVNKUqKaN6hjcjLA81AAAQC13pGzJRqVmqa0Q2clSSN6NNP4AR0Y+QKXiAIIAKjVVu7I1ZMLNiuvpFyhgX56664u6t+5kdmxAI9GAQQA1EplFS5N/mqX/r4uS5IU1yRcM5ISFRsRbHIywPNRAAEAtU726WKNTE3T5sN5kqQHr2mhZ25urwA/H5OTAd6BAggAqFW+2npMT3+6RQWlFQoP8teUwXG68Yoos2MBXoUCCACoFUrLnXpj2U59sv6gJCmxaV1NG56gJvUY+QLVjQIIADDdgZNFSk5N0/aj+ZKkR65tqSdvaid/X0a+QE2gAAIATPX55qN6btFWFToqVC/YX+8OiVef9pFmxwK8GgUQAGCK0nKnJn6xQ3M2HJIkXdW8nqYNT1Cj8CCTkwHejwIIAHC7/ScKlTw7TbtyCmSzScnXtdbjfdvIj5Ev4BYUQACAWy1OP6znF29TcZlT9esE6P1h8erdpqHZsQBLoQACANyipMyplz7bpgWbDkuSerSsr6nD4hUZFmhyMsB6KIAAgBq3J7dAybPTtPd4oWw2acwNbTTq+jby9bGZHQ2wJAogAKDGGIahBZsO66XPtqm03KWGoXZNHRavnq0amB0NsDQKIACgRhQ5KvTCkm1anH5EktS7TQO9OyReDUPtJicDQAEEAFS7ncfylZyapswTRfKxSU/c1E5/uraVfBj5ArUCBRAAUG0Mw9CcDdl6+YvtKqtwKTosUNOGJ6hbiwizowH4DxRAAEC1KCgt13OLt+mLzUclSde1a6h3h8Qrok6AyckA/DcKIADgsm07kqeRqWk6cKpYvj42Pd2vnR7u3ZKRL1BLUQABAJfMMAz9z08H9dqXO1XmdCkmPFDTkxJ1ZbN6ZkcD8CsogACAS5JXUq5nP92ir7blSJL6dojSlMFdVDeYkS9Q21EAAQBVtjn7rEbOSVP26RL5+9r0bP8OeqBXc9lsjHwBT0ABBABcNMMw9Pd1BzT5q50qdxpqUi9IKUmJiouta3Y0AFVAAQQAXJSzxWV6csEWfbMzV5J0c8dovXl3F4UH+ZucDEBVUQABAL9p08EzGj0nXUfOlijA10cv3NpB917djJEv4KEogACAC3K5DP31h0y9vWK3KlyGmtUPVkpSojo1Djc7GoDLQAEEAJzX6aIyPTE/Q9/tPiFJurVLI026s7NCAxn5Ap6OAggA+IUNWac1ek66cvJLFeDno5cHdtTwbrGMfAEvQQEEAFRyuQx98P1+vbtyj5wuQy0b1FHKPYnq0CjM7GgAqhEFEAAgSTpZ6NDYeRn6Ye9JSdIdCY312qBOqmPnVgF4G/5VAwD04/6TGjM3QycKHAr099Ert3XS4K5NGPkCXooCCAAW5nQZmv7tXk1btVcuQ2oTGaKUexLVNirU7GgAahAFEAAs6nh+qR6fl6Ef95+SJA2+sokm3t5RwQHcGgBvx79yALCgH/ae0Nh5GTpZWKbgAF+9NqiT7kxsYnYsAG5CAQQAC6lwuvT+N3uVsnqfDENqHx2qGUmJah0ZYnY0AG5EAQQAiziWV6IxczK04cBpSdLwbk01YeAVCvT3NTkZAHejAAKABXy3+7jGzcvQmeJy1Qnw1aS7uui2uBizYwEwCQUQALxYudOlKV/v1p+/z5QkdYwJ04ykRLVoUMfkZADM5GN2gOqyZs0aDRw4UDExMbLZbFqyZMlvXuNwOPT888+rWbNmstvtat68uf7+97/XfFgAcIMjZ0s09M/rK8vffT2a6dM/9aT8AfCeHcCioiLFxcXpgQce0J133nlR1wwZMkS5ubn66KOP1Lp1ax07dkwul6uGkwJAzVu5I1dPLtisvJJyhdr99ObdXTSgcyOzYwGoJbymAPbv31/9+/e/6POXL1+u77//XpmZmYqIiJAkNW/evIbSAYB7lFW49ObyXfpobZYkqUuTcM0Ynqim9YNNTgagNvGaEXBVff755+rataveeustNW7cWG3bttWTTz6pkpISs6MBwCXJPl2swX9eX1n+HujVQgsf7Un5A/ALXrMDWFWZmZlau3atAgMDtXjxYp08eVKPPfaYTp06pY8//vi81zgcDjkcjsrn+fn57ooLAL9q+bZjemrhFhWUVigs0E9TBsfppo7RZscCUEtZtgC6XC7ZbDbNnj1b4eHhkqR3331Xd999t2bOnKmgoKBfXDNp0iRNnDjR3VEB4IIcFU69sXSn/rH+oCQpoWldTR+eoCb12PUDcGGWHQE3atRIjRs3rix/ktShQwcZhqHDhw+f95rx48crLy+v8pGdne2uuADwCwdOFumuD36sLH+P/K6l5j/Sg/IH4DdZdgewV69eWrBggQoLCxUS8vOvQNqzZ498fHzUpMn5fx+m3W6X3W53Z0wAOK8vNh/V+EVbVeioUL1gf70zJE7Xt48yOxYAD+E1O4CFhYXKyMhQRkaGJCkrK0sZGRk6dOiQpJ937+67777K85OSklS/fn394Q9/0I4dO7RmzRo99dRTeuCBB847/gWA2qC03KnnFm/VqDnpKnRU6Krm9bRsTG/KH4Aq8ZodwI0bN6pPnz6Vz8eNGydJGjFihGbNmqVjx45VlkFJCgkJ0cqVKzVq1Ch17dpV9evX15AhQ/Taa6+5PTsAXIz9JwqVPDtNu3IKZLNJj13XSmP7tpWfr9f8vzwAN7EZhmGYHcJT5efnKzw8XHl5eQoLCzM7DgAvtjj9sJ5fvE3FZU7VrxOg94bG63dtG5odC/BI3L+9aAcQALxRSZlTEz7fpvkbf/7htKtbRmjqsARFhQWanAyAJ6MAAkAttTe3QMmpadqTWyibTRp9fRuNvqGNfH1sZkcD4OEogABQCy3YmK0XP9um0nKXGobaNXVovHq2bmB2LABeggIIALVIkaNCL362TYvSjkiSrmndQO8NjVfDUN6CCkD1oQACQC2x81i+Rqamaf+JIvnYpHE3ttVj17WWDyNfANWMAggAJjMMQ3M2ZGviF9vlqHApKsyuacMS1L1lfbOjAfBSFEAAMFFBabmeW7xNX2w+Kkm6tm1DvTskTvVDGPkCqDkUQAAwybYjeRqZmqYDp4rl62PTU/3a6Y+9WzLyBVDjKIAA4GaGYeh/fzqoV7/cqTKnSzHhgZqelKArm0WYHQ2ARVAAAcCN8kvL9eynW7Rsa44kqW+HSL19d5zq1QkwORkAK6EAAoCbbM4+q5Fz0pR9ukR+PjY927+9HrymhWw2Rr4A3IsCCAA1zDAMfbzugCZ9tVPlTkNN6gVpRlKi4mPrmh0NgEVRAAGgBp0tLtNTC7do5Y5cSdLNHaP15t1dFB7kb3IyAFZGAQSAGpJ26IxGpabryNkSBfj66PlbOui+Hs0Y+QIwHQUQAKqZy2Xob2sz9dby3apwGWpWP1gzhieqc5Nws6MBgCQKIABUq9NFZXpywWZ9u+u4JOmWLo00+c7OCg1k5Aug9qAAAkA1+deB0xqVmq6c/FIF+PlowsArlNStKSNfALUOBRAALpPLZeiD7/fr3ZV75HQZatmgjmYkJeqKmDCzowHAeVEAAeAynCx0aOy8DP2w96QkaVB8jF67o7NC7Hx7BVB78R0KAC7R+v2nNGZuuo4XOBTo76NXbuukwV2bMPIFUOtRAAGgipwuQzO+3aepq/bIZUitI0OUkpSodtGhZkcDgItCAQSAKjheUKrH52box/2nJEmDr2yiibd3VHAA304BeA6+YwHARVq796Qen5euk4VlCvL31et3dNKdiU3MjgUAVUYBBIDfUOF0aeqqvZrx3T4ZhtQ+OlQzkhLVOjLE7GgAcEkogADwK3LySjV6bro2ZJ2WJA3vFqsJAzsq0N/X5GQAcOkogABwAat3H9e4+Zt1uqhMdQJ89cadnXV7fGOzYwHAZaMAAsB/KXe69M7Xe/Th9/slSVc0ClPKPYlq0aCOyckAoHpQAAHgPxw5W6LRc9K16eAZSdK9VzfT87d0YOQLwKtQAAHg//tmR66eXLhZZ4vLFWr305t3d9GAzo3MjgUA1Y4CCMDyyipcemv5Lv1tbZYkqUuTcM0Ynqim9YNNTgYANYMCCMDSsk8Xa+ScdG3OPitJ+kOv5nq2f3vZ/Rj5AvBeFEAAlrV8W46eWrhZBaUVCgv009uD49SvY7TZsQCgxlEAAViOo8KpSct2adaPByRJCU3ravrwBDWpx8gXgDVQAAFYysFTRRqZmq6tR/IkSX/8XUs91a+d/H19TE4GAO5DAQRgGV9uOapnP92qQkeF6gX7650hcbq+fZTZsQDA7SiAALxeablTr365Q7P/eUiS1LVZPU1PSlCj8CCTkwGAOSiAALza/hOFSp6dpl05BZKkx65rpXE3tpUfI18AFkYBBOC1lqQf0XOLt6q4zKn6dQL07tB4Xdu2odmxAMB0FEAAXqekzKmXP9+ueRuzJUlXt4zQ1GEJigoLNDkZANQOFEAAXmVvboGSU9O0J7dQNps06vo2GnNDG/n62MyOBgC1BgUQgNdYsDFbL322XSXlTjUIsWvqsHj1at3A7FgAUOtQAAF4vCJHhV78bJsWpR2RJF3TuoHeGxqvhqF2k5MBQO1EAQTg0Xbl5Ct5dpr2nyiSj00a27etHuvTmpEvAPwKCiAAj2QYhub9K1sTPt8uR4VLUWF2TR2WoKtb1jc7GgDUehRAAB6n0FGh5xZt1eebj0qSrm3bUO8OiVP9EEa+AHAxKIAAPMq2I3kamZqmA6eK5etj05M3tdMjv2spH0a+AHDRKIAAPIJhGPrfnw7q1aU7VVbhUqPwQE0fnqCuzSPMjgYAHocCCKDWyy8t17OfbtGyrTmSpBvaR2rK4DjVqxNgcjIA8EwUQAC12pbDZzUyNV2HThfLz8emZ/u314PXtJDNxsgXAC4VBRBArWQYhj5ed0CTvtqpcqehxnWDNCMpQQlN65kdDQA8HgUQQK2TV1yupxZu1tc7ciVJN10RpbfvjlN4sL/JyQDAO1AAAdQq6YfOaGRquo6cLVGAr4+eG9BeI3o2Z+QLANWIAgigVnC5DH20NktvLt+lCpehphHBSklKVOcm4WZHAwCvQwEEYLozRWV6YsFmfbvruCTpli6NNOnOzgoLZOQLADWBAgjAVBsPnNaoOek6lleqAD8fvXTrFbqne1NGvgBQgyiAAEzhchn64Pv9enflHjldhlo0qKMZSQnqGMPIFwBqmo/ZAarLmjVrNHDgQMXExMhms2nJkiUXfe26devk5+en+Pj4GssH4P+cLHTo/ln/0tsrdsvpMnR7fIy+GHUN5Q8A3MRrCmBRUZHi4uKUkpJSpevOnj2r++67TzfccEMNJQPwn37KPKUBU3/Qmj0nZPfz0Zt3ddb7Q+MVYmcgAQDu4jXfcfv376/+/ftX+bpHH31USUlJ8vX1rdKuIYCqcboMzfh2n6au2iOXIbWODFFKUqLaRYeaHQ0ALMdrdgAvxccff6zMzExNmDDB7CiAVzteUKr7/v5PvffNz+XvrsQm+nxkL8ofAJjEa3YAq2rv3r169tln9cMPP8jP7+KWweFwyOFwVD7Pz8+vqXiA11i376TGzM3QyUKHgvx99eqgTrr7yiZmxwIAS7NkAXQ6nUpKStLEiRPVtm3bi75u0qRJmjhxYg0mA7xHhdOlaav2avp3+2QYUruoUKXck6DWkez6AYDZbIZhGGaHqG42m02LFy/WoEGDzvvnZ8+eVb169eTr61t5zOVyyTAM+fr66uuvv9b111//i+vOtwMYGxurvLw8hYWFVfvnAXiq3PxSjZqTrg1ZpyVJw7vFasLAjgr09/2NKwGg5uXn5ys8PNzS929L7gCGhYVp69at5xybOXOmvv32Wy1cuFAtWrQ473V2u112u90dEQGPtXr3cY2bv1mni8pUJ8BXb9zZWbfHNzY7FgDgP3hNASwsLNS+ffsqn2dlZSkjI0MRERFq2rSpxo8fryNHjuiTTz6Rj4+POnXqdM71kZGRCgwM/MVxABen3OnSuyv36IPV+yVJHRqFKSUpQS0bhpicDADw37ymAG7cuFF9+vSpfD5u3DhJ0ogRIzRr1iwdO3ZMhw4dMise4NWOni3RqDnp2nTwjCTp3qub6flbOjDyBYBayitfA+guvIYAkFbtzNUTCzbrbHG5Qu1+mnxXF93SpZHZsQDggrh/e9EOIAD3Kqtw6a3lu/S3tVmSpM6NwzUjKUHN6tcxORkA4LdQAAFUWfbpYo2ak66M7LOSpPt7Ntf4Ae1l92PkCwCegAIIoEpWbM/RUws2K7+0QmGBfnp7cJz6dYw2OxYAoAoogAAuiqPCqUnLdmnWjwckSfGxdTV9eIJiI4LNDQYAqDIKIIDfdPBUkUampmvrkTxJ0sO9W+ipfu0V4GfpXycOAB6LAgjgVy3dckzPfrpFBY4K1Q321zuD43RDhyizYwEALgMFEMB5lZY79drSHfrfn35+/8yuzepp2vAExdQNMjkZAOByUQAB/ELmiUIlp6Zr57F8SdJj17XS2Bvbyt+XkS8AeAMKIIBzfJZxRM8t2qqiMqci6gTovaHxurZtQ7NjAQCqEQUQgCSppMypiV9s19x/ZUuSureI0LThCYoKCzQ5GQCgulEAAWjf8QIlz07X7twC2WzSqD6tNfqGNvJj5AsAXokCCFjcwk2H9eKSbSopd6pBiF3vD43XNW0amB0LAFCDKICARRWXVejFJdv1adphSVKv1vX13tB4RYYy8gUAb0cBBCxod06BHpu9SftPFMnHJj3et62S+7SWr4/N7GgAADegAAIWYhiG5v0rWxM+3y5HhUtRYXZNHZagq1vWNzsaAMCNKICARRQ6KvT84q36LOOoJOl3bRvqvSFxqh9iNzkZAMDdKICABWw/mqdRqenKPFkkXx+bnriprR79XSv5MPIFAEuiAAJezDAM/e8/D+nVL3eorMKlRuGBmj48QV2bR5gdDQBgIgog4KXyS8s1/tOtWrr1mCTphvaRmjI4TvXqBJicDABgNgog4IW2HD6rkanpOnS6WH4+Nj1zc3s91LuFbDZGvgAACiDgVQzD0KwfD+iNZTtV7jTUuG6QpiclKLFpPbOjAQBqEQog4CXyisv19KebtWJ7riTppiui9PbdcQoP9jc5GQCgtqEAAl4g/dAZjUxN15GzJfL3tem5AR10f8/mjHwBAOdFAQQ8mGEY+tsPWXpz+S5VuAw1jQjWjKQEdWlS1+xoAIBajAIIeKgzRWV6csFmrdp1XJI0oHO0Jt/VRWGBjHwBAL+OAgh4oI0HTmv0nHQdzStVgJ+PXrz1Cv2+e1NGvgCAi0IBBDyIy2XowzX79c7Xe+R0GWrRoI5mJCWoY0y42dEAAB6EAgh4iFOFDo2bv1nf7zkhSbo9Pkav39FZIXb+GQMAqoY7B+ABfso8pTFz05Wb75Ddz0cTb+uooVfFMvIFAFwSCiBQizldhlK+26f3v9kjlyG1alhHKfckqn10mNnRAAAejAII1FLHC0o1dl6G1u07JUm6K7GJXh3UUcEB/LMFAFwe7iRALbRu30mNmZuhk4UOBfn76tVBnXT3lU3MjgUA8BIUQKAWcboMTV21V9O/3SvDkNpGhSglKVFtokLNjgYA8CIUQKCWyM0v1eg56fpn1mlJ0rCrYjVhYEcFBfianAwA4G0ogEAt8P2eExo3L0OnispUJ8BXb9zZWbfHNzY7FgDAS1EAARNVOF16Z+UefbB6vySpQ6MwpSQlqGXDEJOTAQC8GQUQMMnRsyUaPSddGw+ekST9/uqmeuGWKxToz8gXAFCzKICACb7dlatx8zfrbHG5Qux+mnxXZ93aJcbsWAAAi6AAAm5U7nTpreW79NcfsiRJnRuHa0ZSgprVr2NyMgCAlVAAATfJPl2sUXPSlZF9VpJ0f8/mGj+gvex+jHwBAO5FAQTcYMX2HD21YLPySysUFuint+6O082dos2OBQCwKAogUIMcFU5N/mqXPl53QJIUF1tXM4YnKDYi2NxgAABLowACNeTQqWIlp6Zp65E8SdLDvVvoqX7tFeDnY3IyAIDVUQCBGrBs6zE9s3CLChwVqhvsryl3x6nvFVFmxwIAQBIFEKhWpeVOvb50p/7np4OSpCub1dO04QlqXDfI5GQAAPwfCiBQTbJOFil5dpp2HMuXJP3pulYad2Nb+fsy8gUA1C4UQKAafJZxRM8t2qqiMqci6gTo3SFxuq5dpNmxAAA4LwogcBlKy516+fPtmvuvbElStxYRmjYsQdHhgSYnAwDgwiiAwCXad7xAybPTtTu3QDabNLJPa425oY38GPkCAGo5CiBwCT7ddFgvLNmmknKnGoTY9f7QeF3TpoHZsQAAuCgUQKAKissq9NJn27Vw02FJUs9W9fX+sHhFhjLyBQB4DgogcJF25xQoOTVN+44XyscmjbmhrUZe31q+PjazowEAUCUUQOA3GIah+RuzNeHz7Sotdyky1K6pwxLUo1V9s6MBAHBJKIDAryh0VOiFxVu1JOOoJKl3mwZ6b2i8GoTYTU4GAMClowACF7DjaL5GpqYp82SRfH1seuKmtnr0d63kw8gXAODhKIDAfzEMQ7P/eUivfLlDZRUuNQoP1LThCbqqeYTZ0QAAqBYUQOA/FJSW69lFW7V0yzFJ0vXtIzVlcJwi6gSYnAwAgOrjNe9Yu2bNGg0cOFAxMTGy2WxasmTJr56/aNEi3XjjjWrYsKHCwsLUo0cPrVixwj1hUSttPZynW6ev1dItx+TnY9NzA9rrb/d1pfwBALyO1xTAoqIixcXFKSUl5aLOX7NmjW688UYtW7ZMmzZtUp8+fTRw4EClp6fXcFLUNoZhaNa6LN31wY86eKpYjesGaf6jPfRHXu8HAPBSNsMwDLNDVDebzabFixdr0KBBVbquY8eOGjp0qF566aWLOj8/P1/h4eHKy8tTWFjYJSSF2fKKy/X0p5u1YnuuJOnGK6I05e44hQf7m5wMAFBTuH/zGsBKLpdLBQUFioi48Av9HQ6HHA5H5fP8/Hx3REMNycg+q5GpaTp8pkT+vjaN799Bf+jVXDYbu34AAO9GAfz/pkyZosLCQg0ZMuSC50yaNEkTJ050YyrUBMMw9NHaLE3+apcqXIZiI4I0Y3ii4mLrmh0NAAC38JrXAF6O1NRUTZw4UfPnz1dkZOQFzxs/frzy8vIqH9nZ2W5MiepwtrhMD3+yUa8t3akKl6EBnaO1dHRvyh8AwFIsvwM4d+5cPfTQQ1qwYIH69u37q+fa7XbZ7fwGCE+16eBpjUpN19G8UgX4+ujFWzvo91c3Y+QLALAcSxfAOXPm6IEHHtDcuXN1yy23mB0HNcTlMvTnNZma8vVuOV2GmtcP1oykRHVqHG52NAAATOE1BbCwsFD79u2rfJ6VlaWMjAxFRESoadOmGj9+vI4cOaJPPvlE0s9j3xEjRmjq1Knq3r27cnJyJElBQUEKD6cYeItThQ49sWCzVu8+IUm6LS5Gb9zZWSF2r/nSBwCgyrzmbWBWr16tPn36/OL4iBEjNGvWLN1///06cOCAVq9eLUm67rrr9P3331/w/IvBj5HXbv/MPKXRc9OVm++Q3c9HL9/WUcOuimXkCwAWx/3biwqgGfgCqp2cLkMzv9un977ZI5chtWpYRyn3JKp9NP+NAADcvyUvGgEDknSiwKGx8zK0dt9JSdKdiY316u2dVIeRLwAAlbgrwmv8uO+kRs/N0MlCh4L8ffXK7R01uGus2bEAAKh1KIDweE6Xoamr9mr6t3tlGFLbqBClJCWqTVSo2dEAAKiVKIDwaLn5pRozN10/ZZ6WJA3tGquXb+uooABfk5MBAFB7UQDhsdbsOaGx8zJ0qqhMwQG+euOOzhqU0NjsWAAA1HoUQHicCqdL767co5mr90uS2keHKuWeRLVqGGJyMgAAPAMFEB7lWF6JRs9J178OnJEk3dO9qV689QoF+jPyBQDgYlEA4TG+23Vc4+Zn6ExxuULsfpp8V2fd2iXG7FgAAHgcCiBqvXKnS1NW7Naf12RKkjo1DtOM4Ylq3qCOyckAAPBMFEDUaofPFGvUnHSlHzorSbq/Z3ONH9Bedj9GvgAAXCoKIGqtr7fn6MkFm5VfWqHQQD+9fXcX3dypkdmxAADweBRA1DplFS5N+mqnPl53QJIU1yRcM5ISFRsRbG4wAAC8BAUQtcqhU8UaOSdNWw7nSZIeuqaFnr65vQL8fExOBgCA96AAotZYtvWYnlm4RQWOCoUH+eudwXHqe0WU2bEAAPA6FECYrrTcqdeX7tT//HRQkpTYtK6mJyWqcd0gk5MBAOCdKIAwVdbJIo1MTdP2o/mSpEeubaknb2onf19GvgAA1BQKIEzz+eajGv/pFhWVORVRJ0DvDIlTn3aRZscCAMDrUQDhdqXlTk38YofmbDgkSerWPELThicoOjzQ5GQAAFgDBRBute94oUampmlXToFsNmlkn9Yac0Mb+THyBQDAbSiAcJtPNx3WC0u2qaTcqQYhAXpvaLx6t2lodiwAACyHAogaV1xWoZc+266Fmw5Lknq0rK+pw+IVGcbIFwAAM1AAUaP25BYoeXaa9h4vlI9NGnNDW428vrV8fWxmRwMAwLIogKgRhmFowcbDeunzbSotd6lhqF3ThiWoR6v6ZkcDAMDyKICodkWOCj2/eKuWZByVJPVu00DvDY1XgxC7yckAAIBEAUQ123E0XyNT05R5skg+NumJm9rpT9e2kg8jXwAAag0KIKqFYRhK3XBIE7/YobIKl6LDAjVteIK6tYgwOxoAAPgvFEBctoLSco1ftFVfbjkmSerTrqHeGRKviDoBJicDAADnQwHEZdl2JE/JqWk6eKpYfj42PX1zOz10TUtGvgAA1GIUQFwSwzD0yfqDen3pTpU5XWpcN0jThifoymb1zI4GAAB+AwUQVZZXUq5nFm7R8u05kqS+HaI0ZXAX1Q1m5AsAgCegAKJKMrLPamRqmg6fKZG/r03j+3fQH3o1l83GyBcAAE9BAcRFMQxDH63N0pvLd6ncaSg2IkgzhicqLrau2dEAAEAVUQDxm84Wl+nJBZv1zc7jkqT+naI1+a4uCg/yNzkZAAC4FBRA/KpNB09rVGq6juaVKsDXRy/c2kH3Xt2MkS8AAB6MAojzcrkM/eWHTL29YrecLkPN6wdrRlKiOjUONzsaAAC4TBRA/MKpQoeeWLBZq3efkCQNjIvRG3d0UmggI18AALwBBRDn2JB1WqPmpCk33yG7n49evq2jhl0Vy8gXAAAvQgGEpJ9HvjNX79O7K/fIZUgtG9ZRSlKiOjQKMzsaAACoZhRA6ESBQ+PmZ+iHvSclSXcmNNargzqpjp0vDwAAvBF3eIv7cd9JjZmXoRMFDgX6++iV2ztp8JVNGPkCAODFKIAW5XQZmrZqr6Z9u1eGIbWJDNHMexLVJirU7GgAAKCGUQAt6Hh+qUbPTddPmaclSUO6NtHE2zopKMDX5GQAAMAdKIAWs2bPCY2dl6FTRWUKDvDV63d00h0JTcyOBQAA3IgCaBEVTpfe+2aPZq7eL8OQ2keHakZSolpHhpgdDQAAuBkF0AKO5ZVozJwMbTjw88g3qXtTvXTrFQr0Z+QLAIAVUQC93He7jmvc/AydKS5XiN1Pb9zZWbfFxZgdCwAAmIgC6KXKnS5NWbFbf16TKUnq1DhMM4YnqnmDOiYnAwAAZqMAeqEjZ0s0KjVNaYfOSpJG9Gim527pILsfI18AAEAB9Dord+TqyQWblVdSrtBAP711Vxf179zI7FgAAKAWoQB6ibIKlyZ/tUt/X5clSYprEq4ZSYmKjQg2ORkAAKhtKIBeIPt0sUampmnz4TxJ0oPXtNAzN7dXgJ+PyckAAEBtRAH0cF9tPaanP92igtIKhQf5a8rgON14RZTZsQAAQC1GAfRQpeVOvbFspz5Zf1CSlNi0rqYNT1CTeox8AQDAr6MAeqADJ4uUnJqm7UfzJUmPXNtST97UTv6+jHwBAMBvowB6mM83H9Vzi7aq0FGhesH+endIvPq0jzQ7FgAA8CAUQA9RWu7UxC92aM6GQ5Kkbs0jNHV4vBqFB5mcDAAAeBqvmRmuWbNGAwcOVExMjGw2m5YsWfKb16xevVqJiYmy2+1q3bq1Zs2aVeM5L8X+E4UalLJOczYcks0mjezTWqkPd6f8AQCAS+I1BbCoqEhxcXFKSUm5qPOzsrJ0yy23qE+fPsrIyNDjjz+uhx56SCtWrKjhpFWzOP2wBk5fq105BWoQEqBPHuimJ/u1kx+v9wMAAJfIa0bA/fv3V//+/S/6/A8//FAtWrTQO++8I0nq0KGD1q5dq/fee0/9+vWrqZgXrbisQhM+264Fmw5Lknq0rK+pw+IVGRZocjIAAODpvKYAVtX69evVt2/fc47169dPjz/++AWvcTgccjgclc/z8/NrJNue3AIlz07T3uOFstmkMTe00ajr28jXx1YjHw8AAFiLZeeIOTk5ioo69w2To6KilJ+fr5KSkvNeM2nSJIWHh1c+YmNjayTbjG/3ae/xQjUMtWv2Q931eN+2lD8AAFBtLFsAL8X48eOVl5dX+cjOzq6Rj/Pq7Z1095VNtGx0b/Vs1aBGPgYAALAuy46Ao6OjlZube86x3NxchYWFKSjo/D9da7fbZbfbazxbePDPv9INAACgJlh2B7BHjx5atWrVOcdWrlypHj16mJQIAADAPbymABYWFiojI0MZGRmSfn6bl4yMDB069PMbJ48fP1733Xdf5fmPPvqoMjMz9fTTT2vXrl2aOXOm5s+fr7Fjx5oRHwAAwG28pgBu3LhRCQkJSkhIkCSNGzdOCQkJeumllyRJx44dqyyDktSiRQstXbpUK1euVFxcnN555x397W9/qxVvAQMAAFCTbIZhGGaH8FT5+fkKDw9XXl6ewsLCzI4DAAAuAvdvL9oBBAAAwMWhAAIAAFgMBRAAAMBiKIAAAAAWQwEEAACwGAogAACAxVAAAQAALIYCCAAAYDEUQAAAAIvxMzuAJ/v3L1HJz883OQkAALhY/75vW/mXoVEAL0NBQYEkKTY21uQkAACgqgoKChQeHm52DFPwu4Avg8vl0tGjRxUaGiqbzVatf3d+fr5iY2OVnZ1t2d9T6A6ss3uwzu7BOrsH6+weNbnOhmGooKBAMTEx8vGx5qvh2AG8DD4+PmrSpEmNfoywsDC+wbgB6+werLN7sM7uwTq7R02ts1V3/v7NmrUXAADAwiiAAAAAFkMBrKXsdrsmTJggu91udhSvxjq7B+vsHqyze7DO7sE61yx+CAQAAMBi2AEEAACwGAogAACAxVAAAQAALIYCCAAAYDEUQJOkpKSoefPmCgwMVPfu3bVhw4ZfPX/BggVq3769AgMD1blzZy1btsxNST1fVdb6r3/9q3r37q169eqpXr166tu372/+t8HPqvo1/W9z586VzWbToEGDajagl6jqOp89e1bJyclq1KiR7Ha72rZty/ePi1DVdX7//ffVrl07BQUFKTY2VmPHjlVpaamb0nqmNWvWaODAgYqJiZHNZtOSJUt+85rVq1crMTFRdrtdrVu31qxZs2o8p9cy4HZz5841AgICjL///e/G9u3bjYcfftioW7eukZube97z161bZ/j6+hpvvfWWsWPHDuOFF14w/P39ja1bt7o5ueep6lonJSUZKSkpRnp6urFz507j/vvvN8LDw43Dhw+7Oblnqeo6/1tWVpbRuHFjo3fv3sbtt9/unrAerKrr7HA4jK5duxoDBgww1q5da2RlZRmrV682MjIy3Jzcs1R1nWfPnm3Y7XZj9uzZRlZWlrFixQqjUaNGxtixY92c3LMsW7bMeP75541FixYZkozFixf/6vmZmZlGcHCwMW7cOGPHjh3G9OnTDV9fX2P58uXuCexlKIAm6Natm5GcnFz53Ol0GjExMcakSZPOe/6QIUOMW2655Zxj3bt3Nx555JEazekNqrrW/62iosIIDQ01/vGPf9RURK9wKetcUVFh9OzZ0/jb3/5mjBgxggJ4Eaq6zh988IHRsmVLo6yszF0RvUJV1zk5Odm4/vrrzzk2btw4o1evXjWa05tcTAF8+umnjY4dO55zbOjQoUa/fv1qMJn3YgTsZmVlZdq0aZP69u1beczHx0d9+/bV+vXrz3vN+vXrzzlfkvr163fB8/GzS1nr/1ZcXKzy8nJFRETUVEyPd6nr/MorrygyMlIPPvigO2J6vEtZ588//1w9evRQcnKyoqKi1KlTJ73xxhtyOp3uiu1xLmWde/bsqU2bNlWOiTMzM7Vs2TINGDDALZmtgnth9fIzO4DVnDx5Uk6nU1FRUeccj4qK0q5du857TU5OznnPz8nJqbGc3uBS1vq/PfPMM4qJifnFNx38n0tZ57Vr1+qjjz5SRkaGGxJ6h0tZ58zMTH377be65557tGzZMu3bt0+PPfaYysvLNWHCBHfE9jiXss5JSUk6efKkrrnmGhmGoYqKCj366KN67rnn3BHZMi50L8zPz1dJSYmCgoJMSuaZ2AEELmDy5MmaO3euFi9erMDAQLPjeI2CggLde++9+utf/6oGDRqYHceruVwuRUZG6i9/+YuuvPJKDR06VM8//7w+/PBDs6N5ldWrV+uNN97QzJkzlZaWpkWLFmnp0qV69dVXzY4GXBA7gG7WoEED+fr6Kjc395zjubm5io6OPu810dHRVTofP7uUtf63KVOmaPLkyfrmm2/UpUuXmozp8aq6zvv379eBAwc0cODAymMul0uS5Ofnp927d6tVq1Y1G9oDXcrXc6NGjeTv7y9fX9/KYx06dFBOTo7KysoUEBBQo5k90aWs84svvqh7771XDz30kCSpc+fOKioq0h//+Ec9//zz8vFhr6U6XOheGBYWxu7fJeCr0s0CAgJ05ZVXatWqVZXHXC6XVq1apR49epz3mh49epxzviStXLnygufjZ5ey1pL01ltv6dVXX9Xy5cvVtWtXd0T1aFVd5/bt22vr1q3KyMiofNx2223q06ePMjIyFBsb6874HuNSvp579eqlffv2VRZsSdqzZ48aNWpE+buAS1nn4uLiX5S8f5duwzBqLqzFcC+sZmb/FIoVzZ0717Db7casWbOMHTt2GH/84x+NunXrGjk5OYZhGMa9995rPPvss5Xnr1u3zvDz8zOmTJli7Ny505gwYQJvA3ORqrrWkydPNgICAoyFCxcax44dq3wUFBSY9Sl4hKqu83/jp4AvTlXX+dChQ0ZoaKgxcuRIY/fu3caXX35pREZGGq+99ppZn4JHqOo6T5gwwQgNDTXmzJljZGZmGl9//bXRqlUrY8iQIWZ9Ch6hoKDASE9PN9LT0w1Jxrvvvmukp6cbBw8eNAzDMJ599lnj3nvvrTz/328D89RTTxk7d+40UlJSeBuYy0ABNMn06dONpk2bGgEBAUa3bt2Mn376qfLPrr32WmPEiBHnnD9//nyjbdu2RkBAgNGxY0dj6dKlbk7suaqy1s2aNTMk/eIxYcIE9wf3MFX9mv5PFMCLV9V1/vHHH43u3bsbdrvdaNmypfH6668bFRUVbk7teaqyzuXl5cbLL79stGrVyggMDDRiY2ONxx57zDhz5oz7g3uQ77777rzfb/+9tiNGjDCuvfbaX1wTHx9vBAQEGC1btjQ+/vhjt+f2FjbDYH8aAADASngNIAAAgMVQAAEAACyGAggAAGAxFEAAAACLoQACAABYDAUQAADAYiiAAAAAFkMBBAAAsBgKIAAAgMVQAAEAACyGAggAAGAxFEAAAACLoQACAABYDAUQAADAYiiAAAAAFkMBBAAAsBgKIAAAgMVQAAEAACyGAggAAGAxFEAAAACLoQACAABYDAUQAADAYiiAAAAAFkMBBAAAsBgKIAAAgMVQAAEAACyGAggAAGAxFEAAAACLoQACAABYDAUQAADAYiiAAAAAFvP/ACy//9yZVp61AAAAAElFTkSuQmCC"
    return data
    return jsonify({
        'good_Melave_ids_matzbar': mosadCoordinatorJson["good_Melave_ids_matzbar"],
        'visitDoForBogrim_list': mosadCoordinatorJson["visitDoForBogrim_list"],

        'mitztainim': mitztainim,
        'forgoten_Apprentice_count': len(forgoten_Apprentice_count),
        'call_gap_avg': call_gap_avg,
        'personal_meet_gap_avg': personal_meet_gap_avg,
        'group_meeting_gap_avg': group_meeting_gap_avg,

        'paying_dict': [{"key":k,"value": v} for k, v in paying_dict.items()],
        'mahzor_dict': [{"key":k,"value": v} for k, v in mahzor_dict.items()],
        'sugSherut_dict': [{"key":k,"value": v} for k, v in sugSherut_dict.items()],
        'matzbar_dict': [{"key":k,"value": v} for k, v in matzbar_dict.items()],
        'Picud_dict': [{"key":k,"value": v} for k, v in Picud_dict.items()],

    })


@institutionProfile_form_blueprint.route('/uploadPhoto', methods=['post'])
def uploadPhoto_form():
    if correct_auth() == False:
        return jsonify({'result': f"wrong access token "}), HTTPStatus.OK
    if request.method == "POST":
        institution_id = request.args.get('institution_id')
        print(institution_id)
        print(request.files)
        imagefile = request.files['image']
        # filename = werkzeug.utils.secure_filename(imagefile.filename)
        # print("\nReceived image File name : " + imagefile.filename)
        # imagefile.save( filename)
        new_filename = uuid.uuid4().hex + '.' + imagefile.filename.rsplit('.', 1)[1].lower()
        bucket_name = "th01-s3"
        session = boto3.Session()
        s3_client = session.client('s3',
                                   aws_access_key_id=AWS_access_key_id,
                                   aws_secret_access_key=AWS_secret_access_key)
        s3 = boto3.resource('s3',
                            aws_access_key_id=AWS_access_key_id,
                            aws_secret_access_key=AWS_secret_access_key)
        print(new_filename)
        try:
            s3_client.upload_fileobj(imagefile, bucket_name, new_filename)
        except:
            return jsonify({'result': 'faild', 'image path': new_filename}), HTTPStatus.OK
        updatedEnt = Institution.query.get(institution_id)
        updatedEnt.logo_path = "https://th01-s3.s3.eu-north-1.amazonaws.com/" + new_filename
        db.session.commit()
        # head = s3_client.head_object(Bucket=bucket_name, Key=new_filename)
        return jsonify({'result': 'success', 'image path': new_filename}), HTTPStatus.OK


@institutionProfile_form_blueprint.route('/apprentice_and_melave', methods=['GET'])
def getmyApprentices_form():
    if correct_auth() == False:
        return jsonify({'result': f"wrong access token "}), HTTPStatus.OK
    institution_id = int(request.args.get('institution_id'))
    print(institution_id)
    melave_List = db.session.query(user1).filter(user1.institution_id == institution_id,
                                                 user1.role_ids.contains("0")).all()
    apprenticeList = db.session.query(Apprentice).filter(Apprentice.institution_id == institution_id).all()
    print(melave_List)
    my_dict = []
    for noti in melave_List:
        city = db.session.query(City).filter(City.id == noti.city_id).first()
        my_dict.append(
            {
                "id": str(noti.id),
                "role": "מלווה",
                "first_name": noti.name,
                "last_name": noti.last_name,
                "address": {
                    "country": "IL",
                    "city": city.name,
                    "cityId": str(noti.city_id),
                    "street": noti.address,
                    "houseNumber": "1",
                    "apartment": "1",
                    "region": str(city.cluster_id),
                    "entrance": "a",
                    "floor": "1",
                    "postalCode": "12131",
                    "lat": 32.04282620026557,
                    "lng": 34.75186193813887
                },
            })
        print(my_dict)
    for noti in apprenticeList:
        city = db.session.query(City).filter(City.id == noti.city_id).first()
        my_dict.append(
            {
                "id": str(noti.id),
                "address": {
                    "country": "IL",
                    "city": city.name if city is not None else "",
                    "cityId": str(noti.city_id),
                    "street": noti.address,
                    "houseNumber": "1",
                    "apartment": "1",
                    "region": str(city.cluster_id) if city is not None else "",
                    "entrance": "a",
                    "floor": "1",
                    "postalCode": "12131",
                    "lat": 32.04282620026557,
                    "lng": 34.75186193813887
                },
                "id": str(noti.id), "thMentor": str(noti.accompany_id),
                "militaryPositionNew": str(noti.militaryPositionNew)
                , "avatar": noti.photo_path if noti.photo_path is not None else 'https://www.gravatar.com/avatar',
                "name": str(noti.name), "last_name": str(noti.last_name),
                "institution_id": str(noti.institution_id), "thPeriod": str(noti.hadar_plan_session),
                "serve_type": noti.serve_type,
                "marriage_status": str(noti.marriage_status), "militaryCompoundId": str(noti.base_address),
                "phone": noti.phone, "email": noti.email, "teudatZehut": noti.teudatZehut,
                "highSchoolInstitution": noti.highSchoolInstitution, "army_role": noti.army_role,
                "unit_name": noti.unit_name,
                "onlineStatus": noti.accompany_connect_status, "matsber": str(noti.spirit_status),

                "militaryPositionOld": noti.militarypositionold, "educationalInstitution": noti.educationalinstitution
                , "educationFaculty": noti.educationfaculty, "workOccupation": noti.workoccupation,
                "workType": noti.worktype, "workPlace": noti.workplace, "workStatus": noti.workstatus

            })

    if apprenticeList is None:
        # acount not found
        return jsonify({"result": "Wrong id"})
    if apprenticeList == []:
        # acount not found
        return jsonify({"result": "empty"})
    else:
        # print(f' notifications: {my_dict}]')
        # TODO: get Noti form to DB
        return jsonify(my_dict), HTTPStatus.OK
        # return jsonify([{'id':str(noti.id),'result': 'success',"apprenticeId":str(noti.apprenticeid),"date":str(noti.date),"timeFromNow":str(noti.timefromnow),"event":str(noti.event),"allreadyread":str(noti.allreadyread)}]), HTTPStatus.OK


@institutionProfile_form_blueprint.route('/getProfileAtributes', methods=['GET'])
def getProfileAtributes_form():
    if correct_auth() == False:
        return jsonify({'result': f"wrong access token "}), HTTPStatus.OK
    institution_id = request.args.get('institution_id')
    institution_Ent=db.session.query(Institution).filter(Institution.id == institution_id).first()
    print(institution_Ent)
    if institution_Ent:
        city = db.session.query(City).filter(str(City.id) == institution_Ent.city_id).first()
        list = {"id": str(institution_Ent.id), "name": institution_Ent.name, "owner_id": institution_Ent.owner_id,
                "contact_phone": institution_Ent.contact_phone,
                "city": city.name if city is not None else "", "contact_name": str(institution_Ent.contact_name),
                "phone": str(institution_Ent.phone), "address": institution_Ent.address,
                "avatar": institution_Ent.logo_path if institution_Ent.logo_path is not None else 'https://www.gravatar.com/avatar',
                "eshcol": str(institution_Ent.eshcol_id), "roshYeshiva_phone": institution_Ent.roshYeshiva_phone,
                "roshYeshiva_name": institution_Ent.roshYeshiva_name,
                "admin_phone": str(institution_Ent.admin_phone), "admin_name": institution_Ent.admin_name}
        return jsonify(list), HTTPStatus.OK
    else:
        return jsonify(results="no such id"), HTTPStatus.OK


@institutionProfile_form_blueprint.route("/add_mosad", methods=['post'])
def add_mosad():
    if correct_auth() == False:
        return jsonify({'result': f"wrong access token "}), HTTPStatus.OK
    data = request.json
    print(data)
    name = data['name']
    eshcol = data['eshcol']
    roshYeshiva_phone = data['roshYeshiva_phone']
    roshYeshiva_name = data['roshYeshiva_name']
    admin_phone = data['admin_phone']
    admin_name = data['admin_name']
    contact_name = data['contact_name']
    contact_phone = data['contact_phone']
    owner_id = data['owner_id']

    city = data['city'] + " " if data['city'] is not None else None
    phone = data['phone']
    print(city)
    try:
        cityid = db.session.query(City.id).filter(City.name == city).first()
        print(cityid)
        Institution1 = Institution(
            id=int(str(uuid.uuid4().int)[:5]),
            name=name,
            phone=phone,
            city_id=cityid[0] if cityid is not None else "",
            eshcol_id=eshcol,
            roshYeshiva_phone=roshYeshiva_phone,
            roshYeshiva_name=roshYeshiva_name,
            admin_phone=admin_phone,
            admin_name=admin_name,
            contact_name=contact_name,
            owner_id=owner_id,
            contact_phone=contact_phone
        )
        db.session.add(Institution1)
        db.session.commit()
    except Exception as e:
        return jsonify({'result': 'error while inserting' + str(e)}), HTTPStatus.BAD_REQUEST

    if Institution1:
        # TODO: add contact form to DB
        return jsonify({'result': 'success'}), HTTPStatus.OK


@institutionProfile_form_blueprint.route('/getAll', methods=['GET'])
def getAll():
    if correct_auth() == False:
        return jsonify({'result': f"wrong access token "}), HTTPStatus.OK
    inst_List = db.session.query(Institution).all()
    if inst_List == []:
        return jsonify([]), HTTPStatus.OK
    # print(inst_List)
    my_list = []
    for r in inst_List:
        city = None
        region = None
        if r.city_id != "":
            city = db.session.query(City).filter(City.id == r.city_id).first()
            region = db.session.query(Cluster).filter(Cluster.id == city.cluster_id).first()
        melave_List = db.session.query(user1).filter(user1.institution_id == r.id, user1.role_ids.contains("0")).all()
        apprenticeList = db.session.query(Apprentice.id).filter(Apprentice.institution_id == r.id).all()
        owner_details = db.session.query(user1.name,user1.last_name).filter(user1.id == r.owner_id).first()

        my_list.append(
            {"id": str(r.id), "roshYeshiva_phone": r.roshYeshiva_phone, "roshYeshiva_name": r.roshYeshiva_name,
             "admin_name": r.admin_name, "admin_phone": r.admin_phone,
             "name": r.name, "racaz_firstName": owner_details.name if owner_details else "no owner","racaz_lastName": owner_details.last_name if owner_details else "no owner", "logo_path": r.logo_path or "",
             "contact_phone": r.contact_phone, "address": {
                "country": "IL",
                "city": city.name if city else "",
                "cityId": r.city_id,
                "street": r.address,
                "houseNumber": "1",
                "apartment": "1",
                "shluha": "12131",

                "region": region.name if region else "",
                "entrance": "a",
                "floor": "1",
                "postalCode": "12131",
                "lat": 32.04282620026557,
                "lng": 34.75186193813887
            },
             "score": randrange(100),
             "apprenticeList": [str(row.id) for row in apprenticeList],
             "melave_List": [str(row.id) for row in melave_List],

             "phone": r.phone, "city_id": r.city_id})
    return jsonify(my_list), HTTPStatus.OK


@institutionProfile_form_blueprint.route('/getMahzors', methods=['get'])
def getMahzors():
    try:
        if correct_auth()==False:
            return jsonify({'result': f"wrong access token "}), HTTPStatus.OK
        eshcols_appren = db.session.query(Apprentice.institution_mahzor).distinct(Apprentice.institution_mahzor).all()
        eshcols_appren_ids = [str(row[0]) for row in eshcols_appren]
        return eshcols_appren_ids
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.OK


@institutionProfile_form_blueprint.route("/update", methods=['put'])
def update():
    try:
        if correct_auth()==False:
            return jsonify({'result': f"wrong access token "}), HTTPStatus.OK
        # get tasksAndEvents
        mosad_Id = request.args.get("mosad_Id")
        data = request.json
        updatedEnt = Institution.query.get(mosad_Id)
        for key in data:
            if key == "city":
                CityId = db.session.query(City).filter(
                    City.name == str(data[key])).first()
                setattr(updatedEnt, "city_id", CityId.id)
            if key == "region":
                ClusterId = db.session.query(Cluster.id).filter(
                    Cluster.name == str(data[key])).first()
                setattr(updatedEnt, "cluster_id", ClusterId.id)
            elif key == "email" or key == "birthday":
                if validate_email(data[key]):
                    setattr(updatedEnt, key, data[key])
                else:
                    return jsonify({'result': "email or date -wrong format"}), 401
            else:
                setattr(updatedEnt, front_end_dict[key], data[key])
        db.session.commit()
        if updatedEnt:
            return jsonify({'result': 'success'}), HTTPStatus.OK
        return jsonify({'result': 'error'}), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.OK


@institutionProfile_form_blueprint.route("/add_mosad_excel", methods=['put'])
def add_mosad_excel():
    if correct_auth() == False:
        return jsonify({'result': f"wrong access token "}), HTTPStatus.OK
    file = request.files['file']
    print(file)
    wb = load_workbook(file)
    sheet = wb.active
    not_commited = []
    for row in sheet.iter_rows(min_row=2):
        name = row[0].value.strip()
        phone = str(row[1].value)
        email = row[2].value.strip()
        eshcol = row[3].value.strip()
        roshYeshiva_phone = row[4].value
        roshYeshiva_name = row[5].value.strip()
        admin_phone = row[6].value.strip()
        admin_name = row[7].value.strip()
        owner_id = row[8].value
        logo_path = row[9].value.strip() if row[9].value else ""
        address = row[10].value.strip()
        city = row[11].value.strip()
        contact_name = row[12].value.strip()
        contact_phone = row[13].value
        try:
            CityId = db.session.query(City.id).filter(City.name == city).first()
            Institution1 = db.session.query(Institution.id).filter(Institution.name == name).first()
            if Institution1:
                not_commited.append(name)
                continue
            Institution1 = Institution(
                # email=email,
                id=int(str(uuid.uuid4().int)[:5]),
                eshcol_id=eshcol,
                roshYeshiva_phone=roshYeshiva_phone,
                roshYeshiva_name=roshYeshiva_name,
                admin_phone=admin_phone,
                admin_name=admin_name,
                name=name,
                owner_id=owner_id,
                logo_path=logo_path,
                contact_phone=str(contact_phone),
                contact_name=str(contact_name),
                phone=phone,
                city_id=CityId.id,
                address=address
            )
            db.session.add(Institution1)
            db.session.commit()
        except Exception as e:
            not_commited.append(name)
    return jsonify({'result': 'success', "not_commited": not_commited}), HTTPStatus.OK


@institutionProfile_form_blueprint.route('/delete', methods=['DELETE', 'post'])
def deleteEnt():
    if correct_auth() == False:
        return jsonify({'result': f"wrong access token "}), HTTPStatus.OK
    data = request.json
    try:
        entityId = str(data['entityId'])
        res = db.session.query(Institution).filter(Institution.id == entityId).delete()
        db.session.commit()
        return jsonify({'result': 'sucess'}), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': 'error' + str(e)}), HTTPStatus.BAD_REQUEST
