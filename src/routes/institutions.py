import json
import uuid
from flask import Blueprint, request, jsonify
from http import HTTPStatus
from openpyxl.reader.excel import load_workbook
from config import TH_PNG
import src.routes.performance_metrics as md
from src.logic.compute_score import mosad_score
from src.logic.performence_metric import compute_apprentice_profile, interaction_count
from src.models import  system_report_model
from src.models.cluster_model import Cluster
from src.models.madadim_setting_model import MadadimSetting
from src.models.system_report_model import SystemReport
from src.routes.utils.validations import validate_email
from src.services import db
from src.models.apprentice_model import Apprentice
from src.models.city_model import City
from src.models.region_model import Region
from src.models.institution_model import Institution, front_end_dict
from src.models.user_model import User


institutionProfile_form_blueprint = Blueprint(
    "institutions", __name__, url_prefix="/institutions"
)


def get_institution_list(user_id):
    user_ent = (
        db.session.query(User.role_ids, User.institution_id, User.cluster_id)
        .filter(User.id == user_id)
        .first()
    )
    inst_List = []
    if user_ent is None:
        inst_List = db.session.query(Institution).all()
        return inst_List
    if "0" in user_ent.role_ids:
        inst_List = [
            db.session.query(Institution)
            .filter(Institution.id == user_ent.institution_id)
            .first()
        ]
    if "1" in user_ent.role_ids:
        inst_List = [
            db.session.query(Institution)
            .filter(Institution.id == user_ent.institution_id)
            .first()
        ]
    if "2" in user_ent.role_ids:
        inst_List = (
            db.session.query(Institution)
            .filter(Institution.cluster_id == user_ent.cluster_id)
            .all()
        )
    if "3" in user_ent.role_ids:
        inst_List = db.session.query(Institution).all()

    return inst_List


@institutionProfile_form_blueprint.route("", methods=["GET"])
def get_all():
    user_id = request.args.get("userId")
    inst_List = get_institution_list(user_id)
    if inst_List == []:
        return jsonify([]), HTTPStatus.OK
    # print(inst_List)
    my_list = []
    for r in inst_List:
        city = None
        region = None
        if r.city_id:
            city = db.session.query(City).filter(City.id == r.city_id).first()
            region = (
                db.session.query(Region).filter(Region.id == city.region_id).first()
            )
        melave_List = (
            db.session.query(User)
            .filter(User.institution_id == r.id, User.role_ids.contains("0"))
            .all()
        )
        apprenticeList = (
            db.session.query(Apprentice.id)
            .filter(Apprentice.institution_id == r.id)
            .all()
        )
        if r.owner_id and r.owner_id!="" :
            owner_details = (
                db.session.query(User.name, User.last_name)
                .filter(User.id == r.owner_id)
                .first()
            )
        else:
            owner_details=None
        mosad_score_, forgoten_Apprentice_count = mosad_score(r.id)
        my_list.append(
            r.to_attributes(
                city, region, melave_List, apprenticeList, owner_details, mosad_score_
            )
        )

    return jsonify(my_list), HTTPStatus.OK


@institutionProfile_form_blueprint.route("/add_mosad", methods=["post"])
def add_mosad_manual():
    try:
        data = request.json
        name = data["name"]
        cluster = data["cluster"]
        roshYeshiva_phone = data["roshYeshiva_phone"]
        roshYeshiva_name = data["roshYeshiva_name"]
        admin_phone = data["admin_phone"]
        admin_name = data["admin_name"]
        contact_name = data["contact_name"]
        contact_phone = data["contact_phone"]
        owner_id = data["racaz_id"]
        city = data["city"]
        phone = data["phone"]
        shluha = data["shluha"]
        logo_path = data["logo_path"]

        owner_role = (
            db.session.query(User.id)
            .filter(User.id == owner_id, User.role_ids.contains("1"))
            .first()
        )
        if not owner_role:
            return (
                jsonify({"result": "error- no such racaz id in system"}),
                HTTPStatus.OK,
            )
        CityId = db.session.query(City.id).filter(City.name == city).first()
        if CityId is None:
            return (
                jsonify({"result": "error- no such city"}),
                HTTPStatus.OK,
            )
        cluster_id = (
            db.session.query(Cluster).filter(Cluster.name == cluster).first()
        )
        if cluster_id is None:
            cluster_id=int(str(uuid.uuid4().int)[:5])
            cluster = Cluster(name=cluster, id=int(str(uuid.uuid4().int)[:5]))
            db.session.add(cluster)
            db.session.commit()
        else:
            cluster_id=cluster_id.id
        Institution1 = Institution(
            id=int(str(uuid.uuid4().int)[:5]),
            name=name,
            phone=phone,
            city_id=CityId.id,
            cluster_id=cluster_id,
            roshYeshiva_phone=roshYeshiva_phone,
            roshYeshiva_name=roshYeshiva_name,
            admin_phone=admin_phone,
            admin_name=admin_name,
            contact_name=contact_name,
            owner_id=owner_id,
            contact_phone=contact_phone,
            logo_path=logo_path if logo_path else "",
            shluha=shluha
        )
        db.session.add(Institution1)
        db.session.commit()
        return jsonify({"result": "success"}), HTTPStatus.OK
    except Exception as e:
        return (
            jsonify({"result": "error while inserting" + str(e)}),
            HTTPStatus.BAD_REQUEST,
        )


@institutionProfile_form_blueprint.route("/mosad_general_info", methods=["GET"])
def mosad_general_info():
    try:
        institution_id = request.args.get("institution_id")
        all_Apprentices = (
            db.session.query(
                Apprentice.paying,
                Apprentice.militaryPositionNew,
                Apprentice.spirit_status,
                Apprentice.army_role,
                Apprentice.institution_mahzor,
            )
            .filter(Apprentice.institution_id == institution_id)
            .all()
        )
        Institution_ = (
            db.session.query(Institution.logo_path)
            .filter(Institution.id == institution_id)
            .first()
        )
        all_melaves = (
            db.session.query(User.id)
            .filter(User.institution_id == institution_id)
            .all()
        )
        coordinator = (
            db.session.query(User.id, User.name)
            .filter(User.institution_id == institution_id, User.role_ids.contains("1"))
            .first()
        )
        if coordinator is None:
            return (
                jsonify({"result": "error-no coordinator or such institution"}),
                HTTPStatus.BAD_REQUEST,
            )

        paying_dict, Picud_dict, matzbar_dict, sugSherut_dict, mahzor_dict = (
            compute_apprentice_profile(all_Apprentices)
        )
        mitztainim = []
        for melaveId in all_melaves:
            melave_score, call_gap_avg, personal_meet_gap_avg, group_meeting_gap_avg = (
                md.melave_score(melaveId.id)
            )
            if melave_score > 95:
                mitztainim.append(melaveId.id)
        (
            Mosad_coord_score,
            visitprofessionalMeet_melave_avg,
            visitMatzbar_melave_avg,
            call_gap_avg,
            personal_meet_gap_avg,
            group_meeting_gap_avg,
        ) = md.mosad_Coordinators_score(coordinator.id)
        madadim_setting1 = db.session.query(MadadimSetting).first()
        all_melaves_ids = [melave.id for melave in all_melaves]
        visitcallsCount, visitMeetCount = interaction_count(
            all_melaves_ids, madadim_setting1
        )
        mosad_score_, forgoten_Apprentice_count = md.mosad_score(institution_id)
        resJson = md.mosad_coordinator(coordinator.id)
        mosadCoordinatorJson = resJson[0].json

        matzbar_apprentice_status_mosad = (
            db.session.query(SystemReport.value)
            .filter(
                SystemReport.type
                == system_report_model.matzbar_apprentice_status_mosad,
                SystemReport.related_id == institution_id,
            )
            .first()
        )
        visitmeets_mosad_avg = (
            db.session.query(SystemReport.value)
            .filter(
                SystemReport.type == system_report_model.visitmeets_mosad_avg,
                SystemReport.related_id == institution_id,
            )
            .first()
        )
        visitcalls_mosad_avg = (
            db.session.query(SystemReport.value)
            .filter(
                SystemReport.type == system_report_model.visitcalls_mosad_avg,
                SystemReport.related_id == institution_id,
            )
            .first()
        )
        avg_groupMeeting_gap = (
            db.session.query(SystemReport.value)
            .filter(
                SystemReport.type == system_report_model.avg_groupMeeting_gap,
                SystemReport.related_id == institution_id,
            )
            .first()
        )
        Apprentice_forgoten_count = (
            db.session.query(SystemReport.value)
            .filter(
                SystemReport.type
                == system_report_model.apprentice_forgoten_cnt_mosad_monthly,
                SystemReport.related_id == institution_id,
            )
            .first()
        )
        matzbar_apprentice_status_mosad_dict = (
            json.loads(matzbar_apprentice_status_mosad.value.replace("'", '"'))
            if matzbar_apprentice_status_mosad
            else None
        )
        return jsonify(
            {
                "coordinator_name": coordinator.name,
                "interaction_dict": [
                    {"key": "visitMeetCount", "value": visitMeetCount},
                    {"key": "visitcallsCount", "value": visitcallsCount},
                ],
                "mahzor_dict": [{"key": k, "value": v} for k, v in mahzor_dict.items()],
                "sugSherut_dict": [
                    {"key": k, "value": v} for k, v in sugSherut_dict.items()
                ],
                "matzbar_dict": [
                    {"key": k, "value": v} for k, v in matzbar_dict.items()
                ],
                "call_gap_avg": call_gap_avg,
                "personal_meet_gap_avg": personal_meet_gap_avg,
                "group_meeting_gap_avg": group_meeting_gap_avg,
                "paying_count": paying_dict["TRUE"],
                "mitztainim": mitztainim,
                "apprentice_count": len(all_Apprentices),
                "melave_count": len(all_melaves_ids),
                "done_matzbar_meeting": mosadCoordinatorJson["good_Melave_ids_matzbar"],
                "done_melavim_meeting": mosadCoordinatorJson["new_MelavimMeeting"],
                "visitDoForBogrim_list": mosadCoordinatorJson["visitDoForBogrim_list"],
                "forgoten_Apprentice_count": len(forgoten_Apprentice_count),
                "logo_path": Institution_.logo_path,
                "tochnit_hadar_logo_path": TH_PNG,
                "Apprentice_forgoten_count_is_increasing": (
                    Apprentice_forgoten_count.value < len(forgoten_Apprentice_count)
                    if Apprentice_forgoten_count
                    else True
                ),
                "group_meeting_gap_avg_is_increasing": (
                    avg_groupMeeting_gap.value < group_meeting_gap_avg
                    if avg_groupMeeting_gap
                    else True
                ),
                "call_gap_avg_is_increasing": (
                    visitcalls_mosad_avg.value < call_gap_avg
                    if visitcalls_mosad_avg
                    else True
                ),
                "personal_meet_gap_avg_is_increasing": (
                    visitmeets_mosad_avg.value < personal_meet_gap_avg
                    if visitmeets_mosad_avg
                    else True
                ),
                "matzbar_apprentice_status_is_increasing": (
                    matzbar_apprentice_status_mosad_dict["אדוק"] < matzbar_dict["אדוק"]
                    if matzbar_apprentice_status_mosad_dict
                    else True
                ),
                "MelavimMeeting_todo": mosadCoordinatorJson["MelavimMeeting_todo"],
            }
        )
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST


@institutionProfile_form_blueprint.route("/getMahzors", methods=["get"])
def get_mahzors():
    try:

        eshcols_appren = (
            db.session.query(Apprentice.institution_mahzor)
            .distinct(Apprentice.institution_mahzor)
            .all()
        )
        eshcols_appren_ids = [str(row[0]) for row in eshcols_appren]
        return eshcols_appren_ids
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.OK


@institutionProfile_form_blueprint.route("/update", methods=["put"])
def update():
    try:

        # get tasksAndEvents
        mosad_Id = request.args.get("mosad_Id")
        data = request.json
        updatedEnt = Institution.query.get(mosad_Id)
        for key in data:
            if key == "city":
                CityId = (
                    db.session.query(City).filter(City.name == str(data[key])).first()
                )
                setattr(updatedEnt, "city_id", CityId.id)
            elif key == "eshcol":
                ClusterID = (
                    db.session.query(Cluster)
                    .filter(Cluster.name == str(data[key]))
                    .first()
                )

                setattr(updatedEnt, "cluster_id", ClusterID.id)
            elif key == "region":
                region_id = (
                    db.session.query(Region.id)
                    .filter(Region.name == str(data[key]))
                    .first()
                )
                setattr(updatedEnt, "region_id", region_id.id)
            elif key == "email" or key == "birthday":
                if validate_email(data[key]):
                    setattr(updatedEnt, key, data[key])
                else:
                    return jsonify({"result": "email or date -wrong format"}), 401
            else:
                setattr(updatedEnt, front_end_dict[key], data[key])
        db.session.commit()
        if updatedEnt:
            return jsonify({"result": "success"}), HTTPStatus.OK
        return jsonify({"result": "error"}), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.OK


@institutionProfile_form_blueprint.route("/add_mosad_excel", methods=["put"])
def add_mosad_excel():

    file = request.files["file"]
    wb = load_workbook(file)
    sheet = wb.active
    not_commited = []
    index_ = 2

    columns = [cell.value for cell in sheet[1]]

    def column_value(column_name):
        # Use the Hebrew labels, to handle files with mixing column orders
        try:
            return row[columns.index(column_name)].value
        except ValueError:
            return None

    def strip_or_none(label, default=""):
        value = column_value(label)
        return str(value).strip() if not value is None else default

    for row in sheet.iter_rows(min_row=2):
        name = strip_or_none("שם מוסד")
        owner_id = str(strip_or_none("פלאפון רכז מוסד"))
        cluster_name = strip_or_none("שיוך לאשכול")
        shluha = strip_or_none("שלוחה")
        phone = strip_or_none("טלפון מוסד")
        roshYeshiva_phone = strip_or_none("טלפון ראש מכינה/ישיבה")
        roshYeshiva_name = strip_or_none("שם ראש מכינה/ישיבה")
        admin_phone = strip_or_none("טלפון מנהל אדמינסטרטיבי")
        admin_name = strip_or_none("שם מנהל אדמינסטרטיבי")
        logo_path = strip_or_none("קישור לתמונת לוגו")
        address = strip_or_none("שם")
        city = strip_or_none("מיקום מוסד - עיר")
        contact_name = strip_or_none("שם")
        contact_phone = strip_or_none("שם")
        if not name or not cluster_name or not owner_id:
            not_commited.append(str(index_) + ": חסר שם מוסד או אשכול או רכז מוסד")
            index_ = index_ + 1
            continue
        try:
            cluster_id = (
                db.session.query(Cluster).filter(Cluster.name == cluster_name).first()
            )
            if cluster_id is None:
                cluster = Cluster(name=cluster_name, id=int(str(uuid.uuid4().int)[:5]))
                db.session.add(cluster)
                db.session.commit()
            # owner_role = db.session.query(User.id).filter(User.id == owner_id, User.role_ids.contains("1")).first()
            CityId = db.session.query(City.id).filter(City.name == city).first()
            if CityId is None:
                not_commited.append(str(index_) + ": עיר לא חוקית " + city)
                index_ = index_ + 1
                continue
            Institution1 = (
                db.session.query(Institution).filter(Institution.name == name).first()
            )
            if Institution1:
                Institution1.name = name
                Institution1.address = address
                Institution1.city_id = CityId.id
                Institution1.phone = phone
                Institution1.cluster_id = cluster_id.id
                Institution1.roshYeshiva_phone = roshYeshiva_phone
                Institution1.roshYeshiva_name = roshYeshiva_name
                Institution1.admin_phone = admin_phone
                Institution1.admin_name = admin_name
                Institution1.owner_id = owner_id
                Institution1.logo_path = logo_path
                Institution1.contact_phone = contact_phone
                Institution1.contact_name = contact_name
                Institution1.shluha = shluha

                db.session.add(Institution1)
                db.session.commit()
                index_ = index_ + 1
                continue
            Institution1 = Institution(
                # email=email,
                id=int(str(uuid.uuid4().int)[:5]),
                cluster_id=cluster_id.id,
                roshYeshiva_phone=roshYeshiva_phone,
                roshYeshiva_name=roshYeshiva_name,
                admin_phone=admin_phone,
                admin_name=admin_name,
                name=name,
                owner_id=owner_id,
                logo_path=logo_path,
                contact_phone=str(contact_phone),
                contact_name=str(contact_name),
                phone=phone,
                city_id=CityId.id,
                address=address,
                shluha=shluha,
            )
            db.session.add(Institution1)
            db.session.commit()
            index_ = index_ + 1
        except Exception as e:
            not_commited.append(str(index_) + ":" + str(e))
            index_ = index_ + 1
    return jsonify({"result": "success", "not_commited": not_commited}), HTTPStatus.OK


@institutionProfile_form_blueprint.route("/delete", methods=["DELETE", "post"])
def delete_ent():

    data = request.json
    try:
        entityId = str(data["entityId"])
        db.session.query(Institution).filter(Institution.id == entityId).delete()
        db.session.commit()
        return jsonify({"result": "sucess"}), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": "error" + str(e)}), HTTPStatus.BAD_REQUEST