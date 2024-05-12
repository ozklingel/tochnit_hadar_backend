from http import HTTPStatus

import requests
from flask import Blueprint, request, jsonify
from openpyxl.reader.excel import load_workbook

import config
from src.services import db
from src.models.apprentice_model import Apprentice
from src.models.base_model import Base
from src.models.city_model import City
from src.models.cluster_model import Cluster
from src.models.contact_form_model import ContactForm
from src.models.gift import gift
from src.models.institution_model import Institution
from src.models.notification_model import notifications
from src.models.user_model import user1
from src.models.visit_model import Visit

master_user_form_blueprint = Blueprint('master_user', __name__, url_prefix='/master_user')

@master_user_form_blueprint.route('/setSetting_madadim', methods=['post'])
def setSetting_madadim():
    data = request.json
    try:
        config.call_madad_date = data['call_madad_date']
        config.meet_madad_date = data['meet_madad_date']
        config.groupMeet_madad_date = data['groupMeet_madad_date']
        config.callHorim_madad_date = data['callHorim_madad_date']
        config.madadA_date = data['madadA_date']
        config.madadB_date = data['madadB_date']
        config.madadC_date = data['madadC_date']
        config.madadD_date = data['madadD_date']
    except Exception as e:
        return jsonify({'result': "fail-"+str(e)}), HTTPStatus.OK
    return jsonify({'result': 'success'}), HTTPStatus.OK


@master_user_form_blueprint.route('/getAllSetting_madadim', methods=['GET'])
def getNotificationSetting_form():
    try:
        return jsonify({"eshcolMosadMeet_madad_date":config.eshcolMosadMeet_madad_date,
                        "tochnitMeet_madad_date":config.tochnitMeet_madad_date
                        ,"doForBogrim_madad_date":config.doForBogrim_madad_date,
                        "matzbarmeet_madad_date": config.matzbarmeet_madad_date,
                        "professionalMeet_madad_date": config.professionalMeet_madad_date
                           , "callHorim_madad_date": config.callHorim_madad_date,
                        "groupMeet_madad_date": config.groupMeet_madad_date,
                        "meet_madad_date": config.meet_madad_date
                           , "call_madad_date": config.call_madad_date
                           , "cenes_report": config.cenes_report

                        }), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


def addApperntice(wb):
    sheet = wb.active
    uncommited_ids=[]
    for row in sheet.iter_rows(min_row=2):
        if row[0].value is None or row[1].value is None:
            uncommited_ids.append(row[2].value)
            continue
        first_name = row[0].value.strip()
        last_name = row[1].value.strip()
        phone = row[2].value
        city = row[3].value.strip() if not row[3].value is None else "לא ידוע"
        address = row[4].value.strip() if not row[4].value is None else ""
        teudatZehut = row[5].value  if not row[5].value is None else ""
        birthday_Ivry_month = row[6].value.strip()  if not row[6].value is None else ""
        birthday_Ivry_day = row[7].value.strip()  if not row[7].value is None else ""
        birthday_Ivry = birthday_Ivry_day + " " + birthday_Ivry_month
        birthday_loazi = row[8].value if not row[8].value is None else None
        mail = row[9].value.strip()  if not row[9].value is None else ""
        marriage_status = row[10].value.strip() if not row[10].value is None else ""
        serve_type = row[11].value.strip() if not row[11].value is None else ""
        hadar_plan_session = row[12].value if not row[12].value is None else ""
        contact1_first_relation = row[13].value.strip() if not row[13].value is None else ""
        contact1_first_name = row[14].value.strip() if not row[14].value is None else ""
        contact1_phone = row[15].value if not row[15].value is None else ""
        contact1_email = row[16].value.strip() if not row[16].value is None else ""
        contact2_first_relation = row[17].value.strip() if not row[17].value is None else ""
        contact2_first_name = row[18].value.strip() if not row[18].value is None else ""
        contact2_phone = row[19].value if not row[19].value is None else ""
        contact2_email = row[20].value.strip() if not row[20].value is None else ""
        contact3_first_relation = row[21].value.strip() if not row[21].value is None else ""
        contact3_first_name = row[22].value.strip() if not row[22].value is None else ""
        contact3_phone = row[23].value if not row[23].value is None else ""
        contact3_email = row[24].value.strip() if not row[24].value is None else ""
        marriage_date_month = row[25].value if not row[25].value is None else ""
        marriage_date_day = row[26].value if not row[26].value is None else ""
        marriage_date = marriage_date_day + " " + marriage_date_month
        marriage_date_loazi = row[27].value if not row[27].value is None else None
        institution_mahzor = row[28].value if not row[28].value is None else ""
        teacher_grade_a = row[29].value.strip() if not row[29].value is None else ""
        teacher_grade_a_phone = row[30].value if not row[30].value is None else ""
        teacher_grade_b = row[31].value.strip() if not row[31].value is None else ""
        teacher_grade_b_phone = row[32].value if not row[32].value is None else ""
        paying = row[33].value.strip() if not row[33].value is None else ""
        matzbar = row[34].value.strip() if not row[34].value is None else ""
        high_school_name = row[35].value.strip() if not row[35].value is None else ""
        high_school_teacher = row[36].value if not row[36].value is None else ""
        high_school_teacher_phone = row[37].value if not row[37].value is None else ""
        workstatus = row[38].value.strip()   if not row[38].value is None else ""
        workplace = row[39].value.strip()  if not row[39].value is None else ""
        educationfaculty = row[40].value.strip()  if not row[40].value is None else ""
        workoccupation = row[41].value.strip()  if not row[41].value is None else ""
        #worktype = row[42].value.strip()  #
        unit_name = row[42].value.strip()  if not row[42].value is None else ""# חיל שריון
        army_role = row[43].value.strip()  if not row[43].value is None else ""# צנחנים
        militaryPositionNew = row[45].value.strip() if not row[45].value is None else "" # מפקצ
        militaryPositionOld = row[46].value.strip()  if not row[46].value is None else ""# צנחנים
        recruitment_date = row[47].value if not row[47].value is None else None # צנחנים
        release_date = row[48].value if not row[48].value is None else None # צנחנים
        base_name = row[49].value.strip() if not row[49].value is None else "לא ידוע"
        institution_name = row[50].value.strip() if not row[50].value is None else ""
        accompany_id = row[51].value if not row[51].value is None else "" # מפקד?

        CityId = db.session.query(City.id).filter(City.name == city).first()
        militaryCompoundId = db.session.query(Base.id).filter(Base.name == base_name).first()
        institution_id = db.session.query(Institution.id).filter(Institution.name == institution_name).first()
        if institution_id is None or CityId is None or militaryCompoundId is None:
            uncommited_ids.append(row[2].value)
            continue
        eshcol = db.session.query(Institution.eshcol_id).filter(Institution.id == institution_id.id).first()
        if institution_id is None or eshcol is None or CityId is None or militaryCompoundId is None:
            uncommited_ids.append(row[2].value)
            continue

        try:
            institution_id = db.session.query(Institution.id).filter(Institution.name == str(institution_name)).first()
            Apprentice1 = Apprentice(
                email=mail,
                high_school_teacher=high_school_teacher,
                 release_date=release_date,
                 recruitment_date=recruitment_date,
                militaryPositionOld=militaryPositionOld,
                militaryPositionNew=militaryPositionNew,
                institution_mahzor=institution_mahzor,
                teacher_grade_a_phone=teacher_grade_a_phone,
                teacher_grade_b_phone=teacher_grade_b_phone,
                city_id=CityId.id,
                id=phone,
                eshcol=eshcol.eshcol_id,
                base_address=militaryCompoundId.id,
                institution_id=institution_id.id if institution_id is not None else 0,
                address=address,
                serve_type=serve_type,
                name=first_name,
                last_name=last_name,
                phone=str(phone),
                army_role=army_role,
                marriage_status=marriage_status,
                contact1_email=contact1_email,
                teacher_grade_b=teacher_grade_b,
                teacher_grade_a=teacher_grade_a,
                hadar_plan_session=hadar_plan_session,
                contact2_phone=contact2_phone,
                contact2_first_name=contact2_first_name,
                contact1_phone=contact1_phone,
                contact1_first_name=contact1_first_name,
                teudatZehut=teudatZehut,
                birthday_ivry=birthday_Ivry,
                 birthday=birthday_loazi,
                unit_name=unit_name,
                accompany_id=accompany_id,
                contact3_email=contact3_email,
                contact3_first_name=contact3_first_name,
                contact3_phone=contact3_phone,
                marriage_date_ivry=marriage_date,
                 marriage_date=marriage_date_loazi,
                spirit_status=matzbar,
                contact2_email=contact2_email,
                #worktype=worktype,
                workoccupation=workoccupation,
                educationfaculty=educationfaculty,
                workplace=workplace,
                workstatus=workstatus,
                high_school_teacher_phone=high_school_teacher_phone,
                high_school_name=high_school_name,
                paying=paying,
                contact1_relation=contact1_first_relation,
                contact2_relation=contact2_first_relation,
                contact3_relation=contact3_first_relation

            )
            db.session.add(Apprentice1)
        except Exception as e:
            return jsonify({'result': 'error while inserting' + str(e)}), HTTPStatus.BAD_REQUEST
    db.session.commit()
    print(uncommited_ids)
    return [x for x in uncommited_ids if x is not None]


def addUsers(wb):

    sheet = wb.active
    uncommited_ids=[]
    for row in sheet.iter_rows(min_row=2):
        if row[0].value is None or row[1].value is None or row[2].value is None or row[5].value is None:
            uncommited_ids.append(row[5].value)
            continue
        role_ids = ""
        if "מלווה" in row[2].value.strip():
            role_ids += "0,"
        if "רכז מוסד" in row[2].value.strip():
            role_ids += "1,"
        if "רכז אשכול" in row[2].value.strip():
            role_ids += "2,"
        if "אחראי תוכנית" in row[2].value.strip():
            role_ids += "3,"
        role_ids = role_ids[:-1]
        first_name = row[0].value.strip()
        last_name = row[1].value.strip()
        institution_name = row[3].value.strip() if not row[3].value is None else "לא ידוע"
        eshcol = row[4].value.strip() if not row[4].value is None else "" if not row[4].value is None else "לא ידוע"
        phone = str(row[5].value).replace("-", "").strip()
        # email = row[3].value.strip()
        try:
            institution_id = db.session.query(Institution.id).filter(
                Institution.name == str(institution_name)).first()

            user = user1(
                id=int(str(phone).replace("-", "")),
                name=first_name,
                last_name=last_name,
                role_ids=role_ids,
                # email=str(email),
                eshcol=eshcol,
                institution_id=institution_id[0],
            )

            db.session.add(user)

            db.session.commit()
        except Exception as e:
            return jsonify({'result': 'error while inserting' + str(e)}), HTTPStatus.BAD_REQUEST
    return [x for x in uncommited_ids if x is not None]

@master_user_form_blueprint.route("/initDB", methods=['put'])
def initDB():
    try:
        type = request.args.get('type')

        giftCode = db.session.query(gift).delete()
        giftCode = db.session.query(Visit).delete()
        giftCode = db.session.query(ContactForm).delete()
        giftCode = db.session.query(notifications).delete()
        giftCode = db.session.query(user1).delete()
        giftCode = db.session.query(Apprentice).delete()
        db.session.commit()
        uncommited_ids=[]
        if type=="lab":
            path = 'data/apprentice_enter_lab.xlsx'
            wb = load_workbook(filename=path)
            for i in addApperntice(wb):
                uncommited_ids.append(i)
            print("appretice lab loaded")
            path = 'data/user_enter_lab.xlsx'
            wb = load_workbook(filename=path)
            for i in addUsers(wb):
                uncommited_ids.append(i)
            print("user  lab loaded")

        else :
            path = 'data/טבלת הכנסת חניכים - סופית ינון.xlsx'
            wb = load_workbook(filename=path)
            for i in addApperntice(wb):
                uncommited_ids.append(i)
            path = 'data/טבלת הכנסת משתמשים  - סופית ינון.xlsx'
            wb = load_workbook(filename=path)
            for i in addUsers(wb):
                uncommited_ids.append(i)
        return jsonify({'result': 'success', "uncommited_ids": uncommited_ids}), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST