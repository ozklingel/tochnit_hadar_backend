import uuid
from http import HTTPStatus
from flask import Blueprint, request, jsonify
from openpyxl.reader.excel import load_workbook
from src.models.cluster_model import Cluster
from src.services import db
from src.models.apprentice_model import Apprentice
from src.models.base_model import Base
from src.models.city_model import City
from src.models.message_model import Message
from src.models.gift_model import Gift
from src.models.institution_model import Institution
from src.models.task_model_v2 import Task
from src.models.user_model import User
from src.models.report_model import Report

master_user_form_blueprint = Blueprint(
    "master_user", __name__, url_prefix="/master_user"
)
base_dir = "/home/ubuntu/flaskapp/"


@master_user_form_blueprint.route("/get_db_tree", methods=["get"])
def get_db_tree():
    tochint_dict = {}
    tochint_coords = (
        db.session.query(
            User.role_ids,
            User.institution_id,
            User.cluster_id,
            User.id,
            User.name,
            User.id,
        )
        .filter(User.role_ids.contains("3"))
        .all()
    )
    for tochint_coord in tochint_coords:
        tochint_dict[str(tochint_coord.name) + ":" + str(tochint_coord.id)] = (
            tochint_coord.id
        )

    cluster_coords = (
        db.session.query(
            User.role_ids, User.institution_id, User.cluster_id, User.id, User.name
        )
        .filter(User.role_ids.contains("2"))
        .all()
    )
    cluster_coords_dict = {}
    for cluster_coord in cluster_coords:
        institution_coords = (
            db.session.query(
                User.role_ids, User.institution_id, User.cluster_id, User.id, User.name
            )
            .filter(
                User.role_ids.contains("1"), User.cluster_id == cluster_coord.cluster_id
            )
            .all()
        )
        institution_coord_dict = {}
        for institution_coord in institution_coords:
            institution_Ent = (
                db.session.query(Institution)
                .filter(Institution.id == institution_coord.institution_id)
                .first()
            )
            accompanys = (
                db.session.query(
                    User.role_ids,
                    User.institution_id,
                    User.cluster_id,
                    User.id,
                    User.name,
                    User.role_ids,
                )
                .filter(
                    User.role_ids.contains("0"),
                    User.institution_id == institution_coord.institution_id,
                )
                .all()
            )
            apprentice_dict = {}
            for accompany in accompanys:
                apprenticeList = (
                    db.session.query(Apprentice)
                    .filter(Apprentice.accompany_id == accompany.id)
                    .all()
                )
                apprenticeList = [r.id for r in apprenticeList]
                apprentice_dict[str(accompany.name) + ":" + str(accompany.id)] = (
                    apprenticeList
                )
            institution_coord_dict[
                str(institution_coord.name)
                + ":"
                + str(institution_coord.id)
                + "-"
                + str(institution_Ent.name)
            ] = apprentice_dict
        cluster_coords_dict[
            str(cluster_coord.name)
            + ":"
            + str(cluster_coord.id)
            + "-"
            + str(cluster_coord.cluster_id)
        ] = institution_coord_dict
    result_dict = {}
    result_dict["clusters"] = cluster_coords_dict
    result_dict["tochint_coordinators"] = tochint_dict

    return result_dict


def addApperntice(wb):
    sheet = wb.active
    columns = [cell.value for cell in sheet[1]]

    commited_ids = []
    not_commited = []

    unknown_value = "לא ידוע"
    index_ = 2

    for row in sheet.iter_rows(min_row=2):

        def column_value(column_name):
            # Use the Hebrew labels, to handle files with mixing column orders
            try:
                return row[columns.index(column_name)].value
            except ValueError:
                return None

        def strip_or_none(label, default=""):
            value = column_value(label)
            return str(value).strip() if not value is None else default

        try:
            name = strip_or_none("שם")
            last_name = strip_or_none("שם משפחה")
            phone = column_value("פלאפון חניך")
            phone = int(str(phone)[:9].replace("-", ""))
            # mandatory values
            if not name or not last_name or not phone or phone < 500000000:
                not_commited.append(
                    str(index_)
                    + ":"
                    + "שדה חובה לא חוקי-"
                    + name
                    + " "
                    + last_name
                    + " "
                    + str(phone)
                )
                index_ = index_ + 1
                continue

            city = strip_or_none("עיר", unknown_value)
            base_name = strip_or_none("בסיס", unknown_value)
            institution_name = strip_or_none("מוסד")

            city_id = db.session.query(City.id).filter(City.name == city).first()
            if city_id is None:
                not_commited.append(str(index_) + ":ערך עיר לא חוקי- " + str(city))
                index_ = index_ + 1
                continue
            military_compound_id = (
                db.session.query(Base.id).filter(Base.name == base_name).first()
            )

            # check valide ids
            if (
                military_compound_id is None
                and base_name != "משוחרר"
                and base_name != "טרום גיוס"
            ):
                not_commited.append(
                    str(index_) + ": ערך בסיס לא חוקי- " + str(base_name)
                )
                index_ = index_ + 1
                continue

            institution_id = (
                db.session.query(Institution.id)
                .filter(Institution.name == institution_name)
                .first()
            )
            if institution_id is None:
                not_commited.append(str(index_) + ": מוסד לא חוקי- " + institution_name)
                index_ = index_ + 1
                continue

            cluster_id = (
                db.session.query(Institution.cluster_id)
                .filter(Institution.id == institution_id.id)
                .first()
            )
            if cluster_id is None:
                not_commited.append(
                    str(index_) + " אשכול לא חוקי עבור מוסד- " + institution_name
                )
                index_ = index_ + 1
                continue

            Apprentice_ = (
                db.session.query(Apprentice).filter(Apprentice.id == str(phone)).first()
            )
            birthday_ivry = None
            if (
                strip_or_none("יום הולדת עברי - יום")
                and strip_or_none("יום הולדת עברי חודש")
                and strip_or_none("יום הולדת עברי - שנה")
            ):
                birthday_ivry = (
                    strip_or_none("יום הולדת עברי - יום")
                    + "' "
                    + strip_or_none("יום הולדת עברי חודש")
                    + strip_or_none("יום הולדת עברי - שנה")
                )

            if Apprentice_:
                Apprentice_.phone = (str(phone),)
                Apprentice_.city_id = city_id.id
                Apprentice_.cluster_id = (cluster_id.cluster_id,)
                Apprentice_.base_address = (
                    (military_compound_id.id,) if military_compound_id else None
                )
                Apprentice_.institution_id = (
                    (institution_id.id if institution_id is not None else None),
                )
                Apprentice_.name = (name,)
                Apprentice_.last_name = (last_name,)
                Apprentice_.address = (strip_or_none("רחוב"),)
                Apprentice_.teudatZehut = (strip_or_none("תעודת זהות"),)
                Apprentice_.birthday_ivry = (birthday_ivry,)
                Apprentice_.birthday = (column_value("יום הולדת לועזי"),)
                Apprentice_.email = (strip_or_none("מייל חניך"),)
                Apprentice_.marriage_status = (strip_or_none("סטאטוס משפחתי"),)
                Apprentice_.serve_type = (strip_or_none("סטאטוס"),)
                Apprentice_.hadar_plan_session = (strip_or_none("מחזור תוכנית הדר"),)
                Apprentice_.contact1_relation = (strip_or_none("איש קשר - 1 - קירבה"),)
                Apprentice_.contact1_first_name = (strip_or_none("איש קשר - 1 - שם "),)
                Apprentice_.contact1_phone = (strip_or_none("איש קשר - 1 -פלאפון"),)
                Apprentice_.contact1_email = (strip_or_none("איש קשר - 1 - מייל "),)
                Apprentice_.contact2_relation = (strip_or_none("איש קשר - 2 - קירבה"),)
                Apprentice_.contact2_first_name = (strip_or_none("איש קשר - 2 - שם "),)
                Apprentice_.contact2_phone = (strip_or_none("איש קשר - 2 -פלאפון"),)
                Apprentice_.contact2_email = (strip_or_none("איש קשר - 2 - מייל "),)
                Apprentice_.contact3_relation = (strip_or_none("איש קשר - 3 - קירבה"),)
                Apprentice_.contact3_first_name = (strip_or_none("איש קשר - 3 - שם "),)
                Apprentice_.contact3_phone = (strip_or_none("איש קשר - 3 -פלאפון"),)
                Apprentice_.contact3_email = (strip_or_none("איש קשר - 3 - מייל "),)
                Apprentice_.marriage_date_ivry = (
                    strip_or_none("יום נישואין תאריך עברי - יום")
                    + "' "
                    + strip_or_none("יום נישואין תאריך עברי - חודש"),
                )
                # marriage_date=column_value("יום נישואין תאריך לועזי"),
                # Temporary, until we fix front-end to handle a null value for this
                Apprentice_.marriage_date = ("2000-01-01",)
                Apprentice_.institution_mahzor = (strip_or_none("מחזור במכינה/ישיבה"),)
                Apprentice_.teacher_grade_a = (strip_or_none('ר"מ שנה א'),)
                Apprentice_.teacher_grade_a_phone = (strip_or_none("פלאפון ר״מ שנה א"),)
                Apprentice_.teacher_grade_b = (strip_or_none('ר"מ שנה ב'),)
                Apprentice_.teacher_grade_b_phone = (strip_or_none("פלאפון ר״מ שנה ב"),)
                Apprentice_.paying = column_value("משלם או לא") == "משלם"
                Apprentice_.spirit_status = (strip_or_none("מצב״ר - מצב רוחני"),)
                Apprentice_.high_school_name = (strip_or_none("שם תיכון"),)
                Apprentice_.high_school_teacher = (strip_or_none("שם מחנך תיכון"),)
                Apprentice_.high_school_teacher_phone = (
                    strip_or_none("פלאפון מחנך תיכון"),
                )
                Apprentice_.workstatus = (strip_or_none("עיסוק"),)
                Apprentice_.workplace = (strip_or_none("מקום עבודה"),)
                Apprentice_.educationfaculty = (strip_or_none("פקולטה"),)
                Apprentice_.workoccupation = (strip_or_none("סוג עבודה"),)
                Apprentice_.army_role = (strip_or_none("סוג שירות"),)  # סיירות
                Apprentice_.unit_name = (strip_or_none("חטיבה - שיוך יחידתי"),)
                Apprentice_.militaryPositionNew = (strip_or_none("תפקיד נוכחי"),)
                Apprentice_.militaryPositionOld = (strip_or_none("תפקיד קודם"),)
                Apprentice_.recruitment_date = (column_value("תאריך גיוס"),)
                Apprentice_.release_date = (column_value("תאריך שחרור"),)
                Apprentice_.accompany_id = (column_value("מלווה אחראי - טלפון"),)
                commited_ids.append(phone)
                index_ = index_ + 1
                db.session.add(Apprentice_)
                continue
            apprentice = Apprentice(
                id=phone,
                phone=str(phone),
                city_id=city_id.id,
                cluster_id=cluster_id.cluster_id,
                base_address=military_compound_id.id if military_compound_id else None,
                institution_id=(institution_id.id if institution_id is not None else 0),
                name=name,
                last_name=last_name,
                address=strip_or_none("רחוב"),
                teudatZehut=strip_or_none("תעודת זהות"),
                birthday_ivry=birthday_ivry,
                birthday=column_value("יום הולדת לועזי"),
                email=strip_or_none("מייל חניך"),
                marriage_status=strip_or_none("סטאטוס משפחתי"),
                serve_type=strip_or_none("סטאטוס"),
                hadar_plan_session=strip_or_none("מחזור תוכנית הדר"),
                contact1_relation=strip_or_none("איש קשר - 1 - קירבה"),
                contact1_first_name=strip_or_none("איש קשר - 1 - שם "),
                contact1_phone=strip_or_none("איש קשר - 1 -פלאפון"),
                contact1_email=strip_or_none("איש קשר - 1 - מייל "),
                contact2_relation=strip_or_none("איש קשר - 2 - קירבה"),
                contact2_first_name=strip_or_none("איש קשר - 2 - שם "),
                contact2_phone=strip_or_none("איש קשר - 2 -פלאפון"),
                contact2_email=strip_or_none("איש קשר - 2 - מייל "),
                contact3_relation=strip_or_none("איש קשר - 3 - קירבה"),
                contact3_first_name=strip_or_none("איש קשר - 3 - שם "),
                contact3_phone=strip_or_none("איש קשר - 3 -פלאפון"),
                contact3_email=strip_or_none("איש קשר - 3 - מייל "),
                marriage_date_ivry=strip_or_none("יום נישואין תאריך עברי - יום")
                + "' "
                + strip_or_none("יום נישואין תאריך עברי - חודש"),
                # marriage_date=column_value("יום נישואין תאריך לועזי"),
                # Temporary, until we fix front-end to handle a null value for this
                marriage_date="2000-01-01",
                institution_mahzor=strip_or_none("מחזור במכינה/ישיבה"),
                teacher_grade_a=strip_or_none('ר"מ שנה א'),
                teacher_grade_a_phone=strip_or_none("פלאפון ר״מ שנה א"),
                teacher_grade_b=strip_or_none('ר"מ שנה ב'),
                teacher_grade_b_phone=strip_or_none("פלאפון ר״מ שנה ב"),
                paying=column_value("משלם או לא") == "משלם",
                spirit_status=strip_or_none("מצב״ר - מצב רוחני"),
                high_school_name=strip_or_none("שם תיכון"),
                high_school_teacher=strip_or_none("שם מחנך תיכון"),
                high_school_teacher_phone=strip_or_none("פלאפון מחנך תיכון"),
                workstatus=strip_or_none("עיסוק"),
                workplace=strip_or_none("מקום עבודה"),
                educationfaculty=strip_or_none("פקולטה"),
                workoccupation=strip_or_none("סוג עבודה"),
                army_role=strip_or_none("סוג שירות"),  # סיירות
                unit_name=strip_or_none("חטיבה - שיוך יחידתי"),
                militaryPositionNew=strip_or_none("תפקיד נוכחי"),
                militaryPositionOld=strip_or_none("תפקיד קודם"),
                recruitment_date=column_value("תאריך גיוס"),
                release_date=column_value("תאריך שחרור"),
                accompany_id=column_value("מלווה אחראי - טלפון"),
                # worktype=worktype,
            )
            db.session.add(apprentice)
            commited_ids.append(phone)
            index_ = index_ + 1
        except Exception as e:
            not_commited.append(str(index_) + ":" + str(e))
            index_ = index_ + 1
    db.session.commit()
    return [x for x in not_commited if x is not None]


def addUsers(wb):
    sheet = wb.active
    uncommited_ids = []
    index_ = 2
    for row in sheet.iter_rows(min_row=2):
        if (
            row[0].value is None
            or row[1].value is None
            or row[2].value is None
            or row[5].value < 500000000
            or row[5].value is None
        ):
            uncommited_ids.append(index_)
            index_ = index_ + 1
            continue
        role_ids = ""
        if "מלווה" in row[2].value.strip():
            role_ids += "0,"
        if "רכז מוסד" in row[2].value.strip():
            role_ids += "1,"
        if "רכז אשכול" in row[2].value.strip():
            role_ids += "2,"
        if "אחראי תוכנית" in row[2].value.strip():
            role_ids += "3,"
        role_ids = role_ids[:-1]
        first_name = row[0].value.strip()
        last_name = row[1].value.strip()
        institution_name = row[3].value.strip() if not row[3].value is None else None
        cluster_name = row[4].value.strip() if not row[4].value is None else None
        phone = str(row[5].value).replace("-", "").strip()
        # email = row[3].value.strip()
        try:
            institution_id = (
                db.session.query(Institution.id)
                .filter(Institution.name == str(institution_name))
                .first()
            )
            cluster_id = (
                db.session.query(Cluster).filter(Cluster.name == cluster_name).first()
            )
            id_ = int(str(phone)[:9].replace("-", ""))
            User_ = db.session.query(User).filter(User.id == id_).first()
            if User_:
                User_.name = first_name
                User_.last_name = last_name
                User_.role_ids = role_ids
                User_.cluster_id = cluster_id.id if cluster_id else None
                User_.institution_id = institution_id.id if cluster_id else None
                db.session.add(User_)
                db.session.commit()
                index_ = index_ + 1
                continue
            user = User(
                id=id_,
                name=first_name,
                last_name=last_name,
                role_ids=role_ids,
                # email=str(email),
                cluster_id=cluster_id.id if cluster_id else None,
                institution_id=institution_id[0] if institution_id else None,
            )
            db.session.add(user)
            db.session.commit()
            index_ = index_ + 1
        except Exception as e:
            print(str(e))
            return (
                jsonify({"result": "error while inserting" + str(e)}),
                HTTPStatus.BAD_REQUEST,
            )
    return [x for x in uncommited_ids if x is not None]


@master_user_form_blueprint.route("/delete_db", methods=["put"])
def delete_db():
    giftCode = db.session.query(Gift).delete()
    giftCode = db.session.query(Report).delete()
    giftCode = db.session.query(Message).delete()
    giftCode = db.session.query(Task).delete()
    giftCode = db.session.query(User).delete()
    giftCode = db.session.query(Apprentice).delete()
    res = db.session.query(Institution).delete()
    db.session.commit()
    return jsonify({"result": "success"}), HTTPStatus.OK


@master_user_form_blueprint.route("/initDB", methods=["put"])
def initDB():
    try:
        type = request.args.get("type")
        total = request.args.get("total")

        giftCode = db.session.query(Gift).delete()
        giftCode = db.session.query(Report).delete()
        giftCode = db.session.query(Message).delete()
        giftCode = db.session.query(Task).delete()
        giftCode = db.session.query(User).delete()
        giftCode = db.session.query(Apprentice).delete()

        db.session.commit()
        uncommited_ids = []
        if total:
            path = base_dir + "data/citiesToAdd.xlsx"
            wb = load_workbook(filename=path)
            res = db.session.query(City).delete()
            upload_CitiesDB(wb)
            print("upload_CitiesDB loaded")
            path = base_dir + "data/טבלת מוסדות לפרודקשן - ינון סגר.xlsx"
            wb = load_workbook(filename=path)
            res = db.session.query(Institution).delete()
            add_mosad_excel(wb)
            print("mosad loaded")
            res = db.session.query(Base).delete()
            upload_baseDB()
            print("upload_baseDB loaded")

        if type == "lab":
            path = "data/apprentice_enter_lab.xlsx"
            wb = load_workbook(filename=path)
            for i in addApperntice(wb):
                uncommited_ids.append(i)
            for row in db.session.query(Apprentice).all():
                setattr(row, "association_date", "2023-05-01")
            db.session.commit()
            print("appretice lab loaded")
            path = "data/user_enter_lab.xlsx"
            wb = load_workbook(filename=path)
            for i in addUsers(wb):
                uncommited_ids.append(i)
            print("user  lab loaded")
        else:
            path = "data/טבלת הכנסת משתמשים לפרודקשן - ינון סגר.xlsx"
            wb = load_workbook(filename=path)
            for i in addUsers(wb):
                uncommited_ids.append(":משתמשים:" + str(i))
            print("user  loaded")
            path = "data/טבלת הכנסת חניכים לפרודקשן - ינון סגר .xlsx"
            wb = load_workbook(filename=path)
            for i in addApperntice(wb):
                uncommited_ids.append(":חניכים:" + str(i))
            # print("appretice  loaded")
            # path = "data/messages.xlsx"
            # wb = load_workbook(filename=path)
            # add_message(wb)
            # print("message  loaded")
            # path = "data/report.xlsx"
            # wb = load_workbook(filename=path)
            # add_report(wb)
            # print("report  loaded")
        return (
            jsonify({"result": "success", "uncommited_ids": uncommited_ids}),
            HTTPStatus.OK,
        )
    except Exception as e:
        print(str(e))
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST


def add_message(wb):
    # /home/ubuntu/flaskapp/

    sheet = wb.active
    for row in sheet.iter_rows(min_row=2):
        created_by_id = row[0].value
        created_for_id = row[1].value
        created_at = row[2].value
        subject = row[3].value
        content = row[4].value
        ent_group = row[6].value.strip() if row[6].value else ""
        attachments = str(row[5].value).split(",")
        icon = row[7].value.strip()
        type = row[8].value.strip()

        if attachments == ["None"]:
            attachments = []
        rep = Message(
            icon=icon,
            id=int(str(uuid.uuid4().int)[:5]),
            type=type,
            created_by_id=created_by_id or "",
            created_at=created_at,
            ent_group=ent_group or "",
            content=content or "",
            subject=subject or "",
            attachments=attachments,
            allreadyread=False,
            created_for_id=created_for_id or "",
        )
        db.session.add(rep)
    try:
        db.session.commit()
    except Exception as e:
        return (
            jsonify({"result": "error while inserting" + str(e)}),
            HTTPStatus.BAD_REQUEST,
        )

    return jsonify({"result": "success"}), HTTPStatus.OK


def add_report(wb):
    # /home/ubuntu/flaskapp/
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2):
        user_id = row[0].value
        ent_reported = row[1].value
        visit_date = row[2].value
        title = row[3].value
        visit_in_army = row[4].value
        description = row[5].value
        attachments = str(row[6].value).split(",")
        ent_group = row[7].value
        if attachments == ["None"]:
            attachments = []
        rep = Report(
            id=int(str(uuid.uuid4().int)[:5]),
            user_id=user_id,
            ent_reported=ent_reported,
            visit_date=visit_date,
            title=title,
            visit_in_army=visit_in_army,
            description=description,
            attachments=attachments,
            allreadyread=False,
            ent_group=ent_group,
        )
        db.session.add(rep)
    try:
        db.session.commit()
    except Exception as e:
        return (
            jsonify({"result": "error while inserting" + str(e)}),
            HTTPStatus.BAD_REQUEST,
        )

    return jsonify({"result": "success"}), HTTPStatus.OK


def upload_CitiesDB(wb):
    try:
        import csv

        my_list = []
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2):
            my_list.append(City(row[2].value, row[1].value.strip(), row[0].value))
        for ent in my_list:
            db.session.add(ent)
            db.session.commit()

        return jsonify({"result": "success"}), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.OK


def add_mosad_excel(wb):
    sheet = wb.active
    not_commited = []
    index_ = 2

    columns = [cell.value for cell in sheet[1]]

    def column_value(column_name):
        # Use the Hebrew labels, to handle files with mixing column orders
        try:
            return row[columns.index(column_name)].value
        except ValueError:
            return None

    def strip_or_none(label, default=""):
        value = column_value(label)
        return str(value).strip() if not value is None else default

    for row in sheet.iter_rows(min_row=2):
        name = strip_or_none("שם מוסד")
        owner_id = strip_or_none("פלאפון רכז מוסד")
        cluster_name = strip_or_none("שיוך לאשכול")
        shluha = strip_or_none("שלוחה")
        phone = strip_or_none("טלפון מוסד")
        roshYeshiva_phone = strip_or_none("טלפון ראש מכינה/ישיבה")
        roshYeshiva_name = strip_or_none("שם ראש מכינה/ישיבה")
        admin_phone = strip_or_none("טלפון מנהל אדמינסטרטיבי")
        admin_name = strip_or_none("שם מנהל אדמינסטרטיבי")
        logo_path = strip_or_none("logo_path")
        address = strip_or_none("שם")
        city = strip_or_none("מיקום מוסד - עיר")
        contact_name = strip_or_none("שם")
        contact_phone = strip_or_none("שם")
        if not name or not cluster_name or not owner_id:
            not_commited.append(index_)
            index_ = index_ + 1
            continue
        try:
            cluster_id = (
                db.session.query(Cluster).filter(Cluster.name == cluster_name).first()
            )
            if cluster_id is None:
                cluster = Cluster(name=cluster_name, id=int(str(uuid.uuid4().int)[:5]))
                db.session.add(cluster)
                db.session.commit()
            # owner_role = db.session.query(User.id).filter(User.id == owner_id, User.role_ids.contains("1")).first()
            CityId = db.session.query(City.id).filter(City.name == city).first()
            if CityId is None:  # or owner_role is None:
                not_commited.append(index_)
                index_ = index_ + 1
                continue

            Institution1 = (
                db.session.query(Institution.id)
                .filter(Institution.name == name)
                .first()
            )
            if Institution1:
                Institution1.name = name
                Institution1.address = address
                Institution1.role_ids = CityId.id
                Institution1.phone = phone
                Institution1.cluster_id = cluster_id.id
                Institution1.roshYeshiva_phone = roshYeshiva_phone
                Institution1.roshYeshiva_name = roshYeshiva_name
                Institution1.admin_phone = admin_phone
                Institution1.admin_name = admin_name
                Institution1.owner_id = owner_id
                Institution1.logo_path = logo_path
                Institution1.contact_phone = contact_phone
                Institution1.contact_name = contact_name
                Institution1.shluha = shluha

                db.session.add(Institution1)
                db.session.commit()
                continue
            Institution1 = Institution(
                # email=email,
                id=int(str(uuid.uuid4().int)[:5]),
                cluster_id=cluster_id.id,
                roshYeshiva_phone=roshYeshiva_phone,
                roshYeshiva_name=roshYeshiva_name,
                admin_phone=admin_phone,
                admin_name=admin_name,
                name=name,
                owner_id=owner_id,
                logo_path=logo_path,
                contact_phone=str(contact_phone),
                contact_name=str(contact_name),
                phone=phone,
                city_id=CityId.id,
                address=address,
                shluha=shluha,
            )
            db.session.add(Institution1)
            db.session.commit()
            index_ = index_ + 1
        except Exception as e:
            not_commited.append(index_)
            index_ = index_ + 1
    return jsonify({"result": "success", "not_commited": not_commited}), HTTPStatus.OK


def upload_baseDB():
    try:
        import csv

        my_list = []
        # /home/ubuntu/flaskapp/
        with open(base_dir + "data/base_add.csv", "r", encoding="utf8") as f:
            reader = csv.reader(f)
            for row in reader:
                ent = Base(str(uuid.uuid4().int)[:5], row[0].strip(), row[1].strip())
                db.session.add(ent)
        db.session.commit()
        return jsonify({"result": "success"}), HTTPStatus.OK
    except Exception as e:
        print(str(e))
        return jsonify({"result": str(e)}), HTTPStatus.OK
