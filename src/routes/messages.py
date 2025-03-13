import json
import uuid
from http import HTTPStatus
from typing import Union, List

import arrow
import requests
from flask import Blueprint, request, jsonify
from openpyxl.reader.excel import load_workbook
from sqlalchemy import or_
from datetime import datetime as dt

import config
from src.services import db
from src.routes.utils.sms import send_sms_019
from .utils.validations import parse_payload
from ..models.apprentice_model import Apprentice
from ..models.institution_model import Institution
from ..models.message_model import Message, IconTypeEnum, MessageTypeEnum
from ..models.user_model import User

messages_form_blueprint = Blueprint("messages", __name__, url_prefix="/messages")


@messages_form_blueprint.route("/add", methods=["POST"])
def add():
    try:
        data = request.json
        parse_payload_ = parse_payload(data)
        if not isinstance(parse_payload_, dict):
            return (
                jsonify({"result": "error" + str(parse_payload_)}),
                HTTPStatus.BAD_REQUEST,
            )
        subject = parse_payload_["subject"]
        content = parse_payload_["content"]
        type_ = MessageTypeEnum[parse_payload_["type"]]
        icon = IconTypeEnum[parse_payload_["icon"]].name
        attachments = parse_payload_["attachments"]
        ent_group_name = parse_payload_["ent_group"]
        created_by_id = parse_payload_["created_by_id"]
        created_for_ids = parse_payload_["created_for_ids"]
        if created_for_ids == []:
            achrah_tohnit = User.query.filter(User.role_ids.contains("3")).all()
            created_for_ids = [str(a.id) for a in achrah_tohnit]
        mess_id = str(uuid.uuid1().int)[:8]

        for key in created_for_ids:
            contact_form1 = Message(
                id=mess_id,  # if ent_group_name!="" else str(uuid.uuid1().int)[:5],
                created_for_id=key,
                created_by_id=created_by_id,
                content=content,
                subject=subject,
                created_at=arrow.now().format("YYYY-MM-DDTHH:mm:ss"),
                allreadyread=False,
                attachments=attachments,
                type=type_,
                ent_group=ent_group_name,
                icon=icon,
            )
            db.session.add(contact_form1)
        db.session.commit()
        if icon == "WHATSAPP":
            message = subject+ "\n\n"+ content
            message += "\n\n*תוכנית הדר*"
            recipients = data["created_for_ids"]
            attachments_1=attachments[0] if attachments and attachments!=[] else None
            returned: List[int] = send_green_whatsapp(attachments_1,message, recipients)
            count_200 = returned.count(200)
            if count_200 == len(returned):
                return jsonify({"result": "success"}), HTTPStatus.OK
        if icon == "SMS":
            message: str = data["content"]
            recipients: List[str] = data["created_for_ids"]
            sources: Union[List[str], str] = [str(559482844)]
            return send_one_sms_019(message, recipients, sources)

        return jsonify({"result": "success"}), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST


@messages_form_blueprint.route("", methods=["GET"])
def get_all_messages_form():
    try:

        user = request.args.get("userId")
        user_role = db.session.query(User.role_ids).filter(user == User.id).first()[0]
        messages_list = (
            db.session.query(
                Message.created_for_id,
                Message.created_at,
                Message.id,
                Message.attachments,
                Message.type,
                Message.icon,
                Message.allreadyread,
                Message.subject,
                Message.content,
                Message.ent_group,
                Message.created_by_id,
            )
            .filter(or_(Message.created_for_id == user, Message.created_by_id == user))
            .all()
        )
        my_dict = []
        groped_mess = []
        group_report_dict = dict()
        for mess in messages_list:
            if (mess.icon == "WHATSAPP" or mess.icon == "SMS") and mess.created_by_id!=user:
                continue
            # filter melave -only  insernal messages
            if (
                user_role == "0"
                and str(mess.created_for_id) != str(user)
                and mess.icon != "INTERNAL"
            ):
                continue
            # filter melave only past messages
            if mess.created_at > dt.today() and str(mess.created_for_id) == str(user):
                continue
            # change tupe to incoming if sender!=creator
            if str(mess.created_by_id) != str(user):
                mess_type = MessageTypeEnum.INCOMING.name
            else:
                mess_type = mess.type.name
            if mess.ent_group != "":
                if mess.ent_group + str(mess.id) in group_report_dict:
                    group_report_dict[mess.ent_group + str(mess.id)].append(
                        str(mess.created_for_id)
                    )
                else:
                    group_report_dict[mess.ent_group + str(mess.id)] = [
                        str(mess.created_for_id)
                    ]
                groped_mess.append(mess)
            else:
                created_for_id = (
                    db.session.query(User.name, User.last_name)
                    .filter(User.id == mess.created_for_id)
                    .first()
                )
                date_ = str(mess.created_at).replace(" ", "T")
                mess = Message(
                    id=mess.id,  # if ent_group_name!="" else str(uuid.uuid1().int)[:5],
                    created_for_id=mess.created_for_id,
                    created_by_id=mess.created_by_id,
                    content=mess.content,
                    subject=mess.subject,
                    created_at=mess.created_at,
                    allreadyread=mess.allreadyread,
                    attachments=mess.attachments,
                    type=mess_type,
                    ent_group=mess.ent_group,
                    icon=mess.icon,
                )
                to_ = [created_for_id.name + " " + created_for_id.last_name]
                my_dict.append(mess.to_attributes(mess_type, to_, date_, ""))
        for mess in groped_mess:
            if group_report_dict[mess.ent_group + str(mess.id)] != None:
                created_for_id_str = ""
                for id_ in group_report_dict[mess.ent_group + str(mess.id)]:
                    created_for_id = (
                        db.session.query(User.name, User.last_name)
                        .filter(User.id == id_)
                        .first()
                    )
                    created_for_id_str += (
                        created_for_id.name + " " + created_for_id.last_name + ","
                    )
                created_for_id_str = created_for_id_str[:-1]

                if str(mess.created_by_id) != str(user):
                    mess_type = MessageTypeEnum.INCOMING.name
                else:
                    mess_type = mess.type
                mess = Message(
                    id=mess.id,  # if ent_group_name!="" else str(uuid.uuid1().int)[:5],
                    created_for_id=created_for_id_str,
                    created_by_id=mess.created_by_id,
                    content=mess.content,
                    subject=mess.subject,
                    created_at=mess.created_at,
                    allreadyread=mess.allreadyread,
                    attachments=mess.attachments,
                    type=mess.type,
                    ent_group=mess.ent_group,
                    icon=mess.icon,
                )
                date_ = str(mess.created_at).replace(" ", "T")
                my_dict.append(
                    mess.to_attributes(
                        mess_type, created_for_id_str, date_, mess.ent_group
                    )
                )
                group_report_dict[mess.ent_group + str(mess.id)] = None
        return jsonify(my_dict), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST


@messages_form_blueprint.route("/send_per_persona", methods=["POST"])
def send_per_persona():
    try:
        data = request.json
        parse_payload_ = parse_payload(data)
        if not isinstance(parse_payload_, dict):
            return (
                jsonify({"result": "error" + str(parse_payload_)}),
                HTTPStatus.BAD_REQUEST,
            )
        roles = parse_payload_["roles"]
        subject = parse_payload_["subject"]
        content = parse_payload_["content"]
        type_ = MessageTypeEnum.OUTGOING
        icon = parse_payload_["icon"]
        attachments = parse_payload_["attachments"]
        created_by_id = str(parse_payload_["created_by_id"])
        created_for_ids = parse_payload_["created_for_apprentices"]
        institution_ent = (
            db.session.query(Institution.id, Institution.cluster_id)
            .filter(
                Institution.id == Apprentice.institution_id,
                Apprentice.id == created_for_ids[0],
            )
            .first()
        )
        user_list = []
        if "0" in roles:
            melave_list = (
                db.session.query(Apprentice.accompany_id)
                .filter(Apprentice.id.in_(created_for_ids))
                .all()
            )
            melave_list_ids = [str(r.accompany_id) for r in melave_list]
            user_list = user_list + melave_list_ids
        if "1" in roles:
            racaz_mosad_list = (
                db.session.query(User.id)
                .filter(
                    User.institution_id == institution_ent.id,
                    User.role_ids.contains("1"),
                )
                .all()
            )
            racaz_mosad_list_ids = [str(r.id) for r in racaz_mosad_list]
            user_list = user_list + racaz_mosad_list_ids
        if "2" in roles:
            racaz_eshcol_list = (
                db.session.query(User.id)
                .filter(
                    User.cluster_id == institution_ent.cluster_id,
                    User.role_ids.contains("2"),
                )
                .all()
            )
            racaz_eshcol_list_ids = [str(r.id) for r in racaz_eshcol_list]
            user_list = user_list + racaz_eshcol_list_ids

        mess_id = str(uuid.uuid1().int)[:5]
        if icon == "INTERNAL":
            for key in set(user_list):
                contact_form_1 = Message(
                    id=mess_id,  # if ent_group_name!="" else str(uuid.uuid1().int)[:5],
                    created_for_id=key,
                    created_by_id=created_by_id,
                    content=content,
                    subject=subject,
                    created_at=arrow.now().format("YYYY-MM-DDTHH:mm:ss"),
                    allreadyread=False,
                    attachments=attachments,
                    type=type_,
                    ent_group="",
                    icon=icon,
                )
                db.session.add(contact_form_1)
            db.session.commit()
        if icon == "WHATSAPP":
            created_for_ids_whatapp = ["0" + a for a in created_for_ids]
            returned: List[int] = send_green_whatsapp(content, created_for_ids_whatapp)
            count_200 = returned.count(200)
            if count_200 == len(returned):
                return jsonify({"result": "success"}), HTTPStatus.OK
        if icon == "SMS":
            created_for_ids_sms = ["0" + a for a in created_for_ids]
            sources: str = "559482844"
            return send_one_sms_019(
                sources=sources, recipients=created_for_ids_sms, message=content
            )
        return jsonify({"result": "success"}), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST


@messages_form_blueprint.route("/send_sms", methods=["POST"])
def send_sms():
    try:
        data = request.json
        message: str = data["message"]
        recipients: List[str] = data["recipients"]
        sources: Union[List[str], str] = data["sources"]
        return send_one_sms_019(message, recipients, sources)
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST


def send_one_sms_019(message, recipients, sources):
    responses = send_sms_019(sources=sources, recipients=recipients, message=message)
    if responses is not None:
        numbers_to_add_to_019 = []
        for number in responses:
            if isinstance(responses[number], dict):
                if (
                    config.SendMessages.Sms.error_message_019
                    in responses[number].values()
                ):
                    numbers_to_add_to_019.append(number)
            else:
                if config.SendMessages.Sms.error_message_019 in responses[number]:
                    numbers_to_add_to_019.append(number)
        if len(numbers_to_add_to_019) > 0:
            return (
                jsonify(
                    {
                        "result": {
                            "response": str(responses),
                            config.SendMessages.Sms.at_least_one_error: config.SendMessages.Sms.message_add_to_019
                            + str(numbers_to_add_to_019),
                        }
                    }
                ),
                HTTPStatus.INTERNAL_SERVER_ERROR,
            )
        return jsonify({"result": str(responses)}), HTTPStatus.INTERNAL_SERVER_ERROR
    return jsonify({"result": "success"}), HTTPStatus.OK


def send_green_whatsapp(
    urlFile :str,message: str, numbers: List[str], delay_send_messages_milliseconds: int = 0
):
    api_url: str = config.GREENAPI_FILE # TODO - put in config
    id_instance: str = config.GREENAPI_INSTANCE  # TODO - put in config
    api_token_instance: str = config.GREENAPI_TOKEN
    url = f"{api_url}/waInstance{id_instance}/sendFileByUrl/{api_token_instance}"
    responses = []
    for number in numbers:
        israel_country_code = "972"
        number_to_send = israel_country_code + str(number)
        if urlFile:
            payload = {
                "chatId": f"{number_to_send}@c.us",
                "caption": message,
                "urlFile": urlFile,
                "fileName": "img_th", #add the filename

            }
            headers = {"Content-Type": "application/json"}
            response = requests.request("POST", url, headers=headers, json=payload)
        else:
            payload = {
            "chatId": f"{number_to_send}@c.us",
            "message": message,
            }
            api_url: str = config.GREENAPI_NO_FILE  # TODO - put in config
        
            url = f"{api_url}/waInstance{id_instance}/sendMessage/{api_token_instance}"
            payload = json.dumps(payload)
            headers = {"Content-Type": "application/json"}
            response = requests.request("POST", url, headers=headers, data=payload)
        responses.append(response.status_code)

    return responses



@messages_form_blueprint.route("/send_whatsapp", methods=["POST"])
def send_whatsapp():
    try:
        data = request.json
        if "message" not in data or "recipients" not in data:
            return (
                jsonify({"result": "missing message or recipients"}),
                HTTPStatus.BAD_REQUEST,
            )
        message = data["message"]
        message += "\n\n*תוכנית הדר*"

        recipients = data["recipients"]
        attachments = data["attachments"]

        returned: List[int] = send_green_whatsapp(attachments[0],message, recipients)
        count_200 = returned.count(200)
        if count_200 == len(returned):
            return jsonify({"result": "success"}), HTTPStatus.OK

        return (
            jsonify(
                {
                    "result": str(
                        f"success with: {count_200}, failed with: {(len(returned) - count_200)}"
                    )
                }
            ),
            HTTPStatus.INTERNAL_SERVER_ERROR,
        )
    except Exception as e:
        return (
            jsonify({"result": str(e), "input, request:": str(request)}),
            HTTPStatus.BAD_REQUEST,
        )


@messages_form_blueprint.route("/setWasRead", methods=["post"])
def set_was_read_message_form():
    data = request.json
    message_id = data["message_id"]
    try:
        Message.query.filter_by(id=message_id).update(
            dict(allreadyread=True)
        )
        db.session.commit()

        if message_id:
            # TODO: add contact form to DB
            return jsonify({"result": "success"}), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST


@messages_form_blueprint.route("/delete", methods=["DELETE", "post"])
def delete_ent():
    data = request.json
    try:
        entity_id = str(data["entityId"])
        db.session.query(Message).filter(Message.id == entity_id).delete()
        db.session.commit()
        return jsonify({"result": "sucess"}), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": "error" + str(e)}), HTTPStatus.BAD_REQUEST


@messages_form_blueprint.route("/add_message_excel", methods=["put"])
def add_message_excel():
    # /home/ubuntu/flaskapp/
    file = request.files["file"]

    wb = load_workbook(file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2):
        created_by_id = row[0].value
        created_for_id = row[1].value
        created_at = row[2].value
        subject = row[3].value
        content = row[4].value
        ent_group = row[6].value.strip() if row[6].value else ""
        attachments = str(row[5].value).split(",")
        icon = row[7].value.strip()
        type_ = row[8].value.strip()

        if attachments == ["None"]:
            attachments = []
        rep = Message(
            icon=icon,
            id=int(str(uuid.uuid4().int)[:5]),
            type=type_,
            created_by_id=created_by_id or "",
            created_at=created_at,
            ent_group=ent_group or "",
            content=content or "",
            subject=subject or "",
            attachments=attachments,
            allreadyread=False,
            created_for_id=created_for_id or "",
        )
        db.session.add(rep)
    try:
        db.session.commit()
    except Exception as e:
        return (
            jsonify({"result": "error while inserting" + str(e)}),
            HTTPStatus.BAD_REQUEST,
        )

    return jsonify({"result": "success"}), HTTPStatus.OK


@messages_form_blueprint.route("/get_recipients", methods=["GET"])
def get_recipients():
    try:
        users = db.session.query(User.id, User.name, User.last_name).all()
        return (
            jsonify(
                [
                    {"id": str(row.id), "name": row.name, "last_name": row.last_name}
                    for row in users
                ],
            ),
            HTTPStatus.OK,
        )
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST