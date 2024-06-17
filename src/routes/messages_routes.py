import json
import uuid
from http import HTTPStatus
from typing import Union, List

import arrow
import requests
from flask import Blueprint, request, jsonify
from openpyxl.reader.excel import load_workbook
from sqlalchemy import or_

import config
from src.services import db
from .utils.sms import send_sms_019
from .search_ent import filter_by_request
from .user_profile import correct_auth
from ..models.message_model import Message
from ..models.user_model import User

messages_form_blueprint = Blueprint('messegaes_form', __name__, url_prefix='/messegaes_form')


@messages_form_blueprint.route('/send_per_persona', methods=['POST'])
def send_per_persona():
    try:
        if correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        data = request.json
        roles = data['roles']
        subject = data['subject']
        content = data['content']
        type = "יוצאות"
        icon = ""
        attachments = []
        try:
            icon = data['icon']
            attachments = data['attachments']
        except Exception as e:
            print(str(e))
        created_by_id = str(data['created_by_id'])
        personas = User.query.filter(User.role_id.in_(roles)).all()
        created_for_ids = [str(a.id) for a in personas]
        if type == "draft":
            created_for_ids = [str(created_by_id)]  # do not send any one
        mess_id = str(uuid.uuid1().int)[:5]
        print("date ", arrow.now().format('YYYY-MM-DDThh:mm:ss'))
        for key in created_for_ids:
            contact_form_1 = Message(
                id=mess_id,  # if ent_group_name!="" else str(uuid.uuid1().int)[:5],
                created_for_id=key,
                created_by_id=created_by_id,
                content=content,
                subject=subject,
                created_at=arrow.now().format('YYYY-MM-DDThh:mm:ss'),
                allreadyread=False,
                attachments=attachments,
                type=type,
                ent_group="",
                icon=icon
            )
            db.session.add(contact_form_1)
        db.session.commit()
        if icon == "WHATSAPP":
            created_for_ids_whatapp = ["0" + a for a in created_for_ids]
            returned: List[int] = send_green_whatsapp(content, created_for_ids_whatapp)
            count_200 = returned.count(200)
            if count_200 == len(returned):
                return jsonify({'result': 'success'}), HTTPStatus.OK
        if icon == "SMS":
            created_for_ids_sms = ["0" + a for a in created_for_ids]
            sources: str = "559482844"
            return send_one_sms_019(sources=sources, recipients=created_for_ids_sms, message=content)
        return jsonify({'result': 'success'}), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@messages_form_blueprint.route('/send_sms', methods=['POST'])
def send_sms():
    try:
        if correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        data = request.json
        message: str = data['message']
        recipients: List[str] = data['recipients']
        sources: Union[List[str], str] = data['sources']
        return send_one_sms_019(message, recipients, sources)
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


def send_one_sms_019(message, recipients, sources):
    responses = send_sms_019(sources=sources, recipients=recipients, message=message)
    if responses is not None:
        numbers_to_add_to_019 = []
        for number in responses:
            if isinstance(responses[number], dict):
                if (config.SendMessages.Sms.error_message_019
                        in responses[number].values()):
                    numbers_to_add_to_019.append(number)
            else:
                if (config.SendMessages.Sms.error_message_019
                        in responses[number]):
                    numbers_to_add_to_019.append(number)
        if len(numbers_to_add_to_019) > 0:
            return jsonify({'result': {
                "response": str(responses),
                config.SendMessages.Sms.at_least_one_error: config.SendMessages.Sms.message_add_to_019 + str(
                    numbers_to_add_to_019)
            }
            }), HTTPStatus.INTERNAL_SERVER_ERROR
        return jsonify({'result': str(responses)}), HTTPStatus.INTERNAL_SERVER_ERROR
    return jsonify({'result': 'success'}), HTTPStatus.OK


def send_green_whatsapp(message: str, numbers: List[str], delay_send_messages_milliseconds: int = 0):
    api_url: str = "https://7103.api.greenapi.com"  # TODO - put in config
    id_instance: str = "7103922187"  # TODO - put in config
    api_token_instance: str = "2a53bacca96949e5a92d03125886fd12cefb7415bf974f4a87"  # TODO - put in config
    responses = []
    for number in numbers:
        # number validation, assuming only israeli numbers, without the 0 in the beginning
        if number.startswith("0"):
            number = number[1:]
        if len(number) != 9:
            print(f"Invalid phone number: {number}, should be only 9 digits long.")
            return

        israel_country_code = "972"
        number_to_send = israel_country_code + number

        payload = {
            "chatId": f"{number_to_send}@c.us",
            "message": message,
        }

        if delay_send_messages_milliseconds > 0:
            payload["delaySendMessagesMilliseconds"] = delay_send_messages_milliseconds

        payload = json.dumps(payload)

        url = f"{api_url}/waInstance{id_instance}/sendMessage/{api_token_instance}"

        headers = {
            'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, data=payload)

        responses.append(response.status_code)

    return responses
    # print(response.text.encode('utf8')) # get message id, for future reference


@messages_form_blueprint.route('/send_whatsapp', methods=['POST'])
def send_whatsapp():
    try:
        if correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        data = request.json
        if 'message' not in data or 'recipients' not in data:
            return jsonify({'result': 'missing message or recipients'}), HTTPStatus.BAD_REQUEST
        message = data['message']
        message += "\n\n*תוכנית הדר*"

        recipients = data['recipients']
        returned: List[int] = send_green_whatsapp(message, recipients)
        count_200 = returned.count(200)
        if count_200 == len(returned):
            return jsonify({'result': 'success'}), HTTPStatus.OK

        return (jsonify({'result': str(f"success with: {count_200}, failed with: {(len(returned) - count_200)}")}),
                HTTPStatus.INTERNAL_SERVER_ERROR)
    except Exception as e:
        return jsonify({'result': str(e), "input, request:": str(request)}), HTTPStatus.BAD_REQUEST


# from chat box
@messages_form_blueprint.route('/add', methods=['POST'])
def add_contact_form():
    try:
        if correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        data = request.json
        subject = data['subject']
        content = data['content']
        type = "פניות_שירות"
        icon = ""
        attachments = []
        ent_group_name = ""
        try:
            type = data['type']
            icon = data['icon']
            attachments = data['attachments']
            ent_group_name = str(data['ent_group'])

        except Exception as e:
            print(str(e))
        created_by_id = str(data['created_by_id'])
        created_for_ids = data['created_for_ids']
        if created_for_ids == [""]:
            achrah_tohnit = User.query.filter(User.role_id == "3").all()
            created_for_ids = [str(a.id) for a in achrah_tohnit]
        if type == "draft":
            created_for_ids = [str(created_by_id)]
        mess_id = str(uuid.uuid1().int)[:5]
        print("date ", arrow.now().format('YYYY-MM-DDThh:mm:ss'))
        for key in created_for_ids:
            contact_form1 = Message(
                id=mess_id,  # if ent_group_name!="" else str(uuid.uuid1().int)[:5],
                created_for_id=key,
                created_by_id=created_by_id,
                content=content,
                subject=subject,
                created_at=arrow.now().format('YYYY-MM-DDThh:mm:ss'),
                allreadyread=False,
                attachments=attachments,
                type=type,
                ent_group=ent_group_name,
                icon=icon
            )
            db.session.add(contact_form1)
            print(contact_form1.created_at)
        db.session.commit()
        return jsonify({'result': 'success'}), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@messages_form_blueprint.route('/getAll', methods=['GET'])
def get_all_messages_form():
    try:
        if correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        user = request.args.get('userId')
        user_role = db.session.query(User.role_id).filter(
            user == User.id).first()[0]
        messages_list = db.session.query(Message.created_for_id, Message.created_at, Message.id,
                                         Message.attachments, Message.type, Message.icon,
                                         Message.allreadyread, Message.subject, Message.content,
                                         Message.ent_group, Message.created_by_id) \
            .filter(or_(Message.created_for_id == user, Message.created_by_id == user)).all()
        my_dict = []
        groped_mess = []
        group_report_dict = dict()
        for mess in messages_list:
            if user_role == "0" and  str(mess.created_for_id) != str(user) and  mess.icon!="INTERNAL" :
                continue
            if str(mess.created_by_id) != str(user):
                mess_type="נכנסות"
            else:
                mess_type=mess.type
            if mess.ent_group != "":
                if mess.ent_group + str(mess.id) in group_report_dict:
                    group_report_dict[mess.ent_group + str(mess.id)].append(str(mess.created_for_id))
                else:
                    group_report_dict[mess.ent_group + str(mess.id)] = [str(mess.created_for_id)]
                groped_mess.append(mess)
            else:
                created_for_id = db.session.query(User.name, User.last_name).filter(
                    User.id == mess.created_for_id).first()
                created_by_id = db.session.query(User.name, User.last_name).filter(
                    User.id == mess.created_by_id).first()
                my_dict.append(
                    {"type": mess_type, "attachments": mess.attachments, "id": str(mess.id),
                     "to": [created_for_id.name + " " + created_for_id.last_name], "ent_group": "",
                     "from": created_by_id.name + " " + created_by_id.last_name,
                     "date": str(mess.created_at).replace(" ", "T"),
                     "content": mess.content, "title": str(mess.subject), "allreadyread": str(mess.allreadyread),
                     "icon": mess.icon})
        for mess in groped_mess:
            if group_report_dict[mess.ent_group + str(mess.id)] != None:
                created_for_id_str = ""
                for id in group_report_dict[mess.ent_group + str(mess.id)]:
                    created_for_id = db.session.query(User.name, User.last_name).filter(User.id == id).first()
                    created_for_id_str += created_for_id.name + " " + created_for_id.last_name + ","
                created_for_id_str = created_for_id_str[:-1]
                created_by_id = db.session.query(User.name, User.last_name).filter(
                    User.id == mess.created_by_id).first()
                if str(mess.created_by_id) != str(user):
                    mess_type = "נכנסות"
                else:
                    mess_type = mess.type
                my_dict.append(
                    {"type": mess_type, "attachments": mess.attachments, "id": str(mess.id),
                     "from": created_by_id.name + " " + created_by_id.last_name,
                     "date": str(mess.created_at).replace(" ", "T"),
                     "to": [created_for_id_str],
                     "content": mess.content, "title": str(mess.subject), "allreadyread": str(mess.allreadyread),
                     "ent_group": mess.ent_group,
                     "icon": mess.icon})
                group_report_dict[mess.ent_group + str(mess.id)] = None
        # print(f' notifications: {my_dict}]')
        # TODO: get Noti form to DB
        return jsonify(my_dict), HTTPStatus.OK
        # return jsonify([{'id':str(noti.id),'result': 'success',"apprenticeId":str(noti.apprenticeid),"date":str(noti.date),"timeFromNow":str(noti.timefromnow),"event":str(noti.event),"allreadyread":str(noti.allreadyread)}]), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@messages_form_blueprint.route('/setWasRead', methods=['post'])
def set_was_read_message_form():
    if correct_auth() == False:
        return jsonify({'result': "wrong access token"}), HTTPStatus.OK
    data = request.json
    message_id = data['message_id']
    print(message_id)
    try:
        # notis =Message.query.filter_by(id=message_id)#db.session.query(Message.id,Message.allreadyread).filter(Message.id==message_id).all()
        num_rows_updated = Message.query.filter_by(id=message_id).update(dict(allreadyread=True))
        db.session.commit()

        if message_id:
            # print(f'setWasRead form: subject: [{subject}, notiId: {notiId}]')
            # TODO: add contact form to DB
            return jsonify({'result': 'success'}), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@messages_form_blueprint.route('/delete', methods=['DELETE', 'post'])
def delete_ent():
    data = request.json
    try:
        if correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        entity_id = str(data['entityId'])
        res = db.session.query(Message).filter(Message.id == entity_id).delete()
        db.session.commit()
        return jsonify({'result': 'sucess'}), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': 'error' + str(e)}), HTTPStatus.BAD_REQUEST


@messages_form_blueprint.route("/add_message_excel", methods=['put'])
def add_message_excel():
    if correct_auth() == False:
        return jsonify({'result': "wrong access token"}), HTTPStatus.OK
    # /home/ubuntu/flaskapp/
    file = request.files['file']

    wb = load_workbook(file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2):
        created_by_id = row[0].value
        created_for_id = row[1].value
        created_at = row[2].value
        subject = row[3].value
        content = row[4].value
        ent_group = row[6].value.strip() if row[6].value else ""
        attachments = str(row[5].value).split(",")
        icon = row[7].value.strip()
        type = row[8].value.strip()

        if attachments == ["None"]:
            attachments = []
        print(row)
        rep = Message(
            icon=icon,
            id=int(str(uuid.uuid4().int)[:5]),
            type=type,
            created_by_id=created_by_id or "",
            created_at=created_at,
            ent_group=ent_group or "",
            content=content or "",
            subject=subject or "",
            attachments=attachments,
            allreadyread=False,
            created_for_id=created_for_id or ""
        )
        db.session.add(rep)
    try:
        db.session.commit()
    except Exception as e:
        return jsonify({'result': 'error while inserting' + str(e)}), HTTPStatus.BAD_REQUEST

    return jsonify({'result': 'success'}), HTTPStatus.OK


@messages_form_blueprint.route("/get_recipients", methods=['GET'])
def get_recipients():
    try:
        if correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        users = db.session.query(User.id, User.name, User.last_name).all()
        return jsonify([{"id": str(row.id), "name": row.name, "last_name": row.last_name} for row in users],
                       ), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST
