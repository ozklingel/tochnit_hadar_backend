import uuid
import time
from datetime import timedelta
from typing import List

from http import HTTPStatus
from flask import Blueprint, request, jsonify
import pyotp

from openpyxl.reader.excel import load_workbook
from timer_dict import TimerDict
from twilio.rest import Client

from src.services import red, db
from src.models.city_model import City
from src.models.user_model import User
from src.routes.Utils.sms import send_sms_019
from src.routes.messages_routes import send_green_whatsapp

secret_key = pyotp.random_base32()
hotp = pyotp.HOTP(secret_key)

user_otp_dict = TimerDict(default_duration=timedelta(minutes=1))
onboarding_form_blueprint = Blueprint('onboarding_form', __name__, url_prefix='/onboarding_form')


@onboarding_form_blueprint.route('/getOTP', methods=['GET'])
def getOTP_form():
    try:
        created_by_phone = request.args.get('created_by_phone')
        userEnt = User.query.get(created_by_phone)
        if userEnt is None:
            return jsonify({"result": "not in system"}), 401
        if str(created_by_phone) in user_otp_dict:
            return jsonify({"result": "already got otp"}), 401
        otp_code = hotp.at(int(created_by_phone))  # You can use different counter values for different OTPs
        print("Your HMAC-based OTP:", otp_code)
        # Generate an OTP using TOTP after every 30 seconds
        send_sms_019(["559482844"], [created_by_phone],
                     "your verify service verification code from **tochnit hadar** is : " + str(otp_code))
        user_otp_dict[str(created_by_phone)] = otp_code
        return jsonify({"result": "success"}), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST

@onboarding_form_blueprint.route('/getOTP_whatsapp', methods=['get'])
def getOTP_whatsapp():
    try:
        created_by_phone = request.args.get('created_by_phone')
        created_by_phone_asList = ["0" + str(created_by_phone)]
        userEnt = User.query.get(created_by_phone)
        if userEnt is None:
            return jsonify({"result": "not in system"}), 401
        print(created_by_phone)
        otp_code = hotp.at(int(created_by_phone))
        print("your verify service verification code from tochnit hadar is:" + otp_code)
        returned: List[int] = send_green_whatsapp(
            "your verify service verification code from *tochnit hadar* is : " + otp_code, created_by_phone_asList)
        user_otp_dict[str(created_by_phone)] = otp_code

        count_200 = returned.count(200)
        if count_200 == len(returned):
            return jsonify({'result': 'success'}), HTTPStatus.OK
        return (jsonify({'result': str(f"success with: {count_200}, failed with: {(len(returned) - count_200)}")}),
                HTTPStatus.INTERNAL_SERVER_ERROR)
    except Exception as e:
        return jsonify({'result': str(e), "input:": str(created_by_phone)}), HTTPStatus.BAD_REQUEST


@onboarding_form_blueprint.route('/verifyOTP', methods=['GET'])
def verifyOTP_form():
    try:
        user_otp = request.args.get('otp')
        created_by_phone = request.args.get('created_by_phone')
        print("user_otp", user_otp)
        print("current TOTP is: ", user_otp_dict[str(created_by_phone)])
        print("created_by_phone", created_by_phone)
        if user_otp_dict[str(created_by_phone)] != user_otp:
            return jsonify({"result": "wrong otp", "firsOnboarding": True}), 401
        userEnt = User.query.get(created_by_phone)
        if userEnt is None:
            return jsonify({"result": "not in system", "firsOnboarding": True}), 401
        if userEnt.name and userEnt.last_name and userEnt.email and userEnt.birthday and userEnt.region_id:
            accessToken = int(str(uuid.uuid4().int)[:5])
            red.hset(created_by_phone, "accessToken", accessToken)
            redisaccessToken = red.hget(created_by_phone, "accessToken").decode("utf-8")
            return jsonify({"result": accessToken, "firsOnboarding": False}), HTTPStatus.OK
        red.hdel(created_by_phone, "accessToken")
        accessToken = int(str(uuid.uuid4().int)[:5])
        red.hset(created_by_phone, "accessToken", accessToken)
        # red.hset(int(str(created_by_phone)), "accessToken", accessToken)
        return jsonify({"result": accessToken, "firsOnboarding": True}), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e), "firsOnboarding": True}), HTTPStatus.BAD_REQUEST


@onboarding_form_blueprint.route('/get_CitiesDB', methods=['GET'])
def get_CitiesDB():
    try:
        CityList = db.session.query(City.name).all()
        return jsonify([i[0] for i in [tuple(row) for row in CityList]])
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@onboarding_form_blueprint.route('/upload_CitiesDB', methods=['GET'])
def upload_CitiesDB():
    try:
        my_list = []
        # /home/ubuntu/flaskapp/src/routes/
        path = 'data/citiesToAdd.xlsx'
        wb = load_workbook(filename=path)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2):
            print(row[0].value)
            my_list.append(City(int(row[2].value), row[1].value.strip(), int(row[0].value)))
        for ent in my_list:
            db.session.add(ent)
        db.session.commit()
        return jsonify({"result": "success"}), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@onboarding_form_blueprint.route('/verifyOTP_whatsapp', methods=['GET'])
def verifyOTP_twilo():
    account_sid = "AC7e7b44337bff9de0cb3702ad5e23e1e8"
    auth_token = "61caa0e3ab00c4d8928e97fdad1a8d52"
    client = Client(account_sid, auth_token)
    otp = request.args.get('otp')
    created_by_phone = request.args.get('created_by_phone')
    print(otp)
    print(created_by_phone)

    result = "error"
    try:
        verification_check = client.verify \
            .v2 \
            .services('VA280c3b665cf155bb76e5bc77bb5c750a') \
            .verification_checks \
            .create(to="+972" + created_by_phone, code=otp)
    except:
        return jsonify({"result": "not in system"}), HTTPStatus.OK

    time.sleep(2.4)
    if verification_check.status != "approved":
        return jsonify({"result": "wrong otp"}), HTTPStatus.OK
    userEnt = User.query.get(created_by_phone)
    if userEnt is None:
        return jsonify({"result": "not in system"}), HTTPStatus.OK

    if userEnt.name and userEnt.last_name and userEnt.email and userEnt.birthday and userEnt.region_id:
        print("not null")
        red.hdel(int(str(created_by_phone)), "accessToken")
        accessToken = int(str(uuid.uuid4().int)[:5])
        red.hset(int(str(created_by_phone)), "accessToken", accessToken)
        return jsonify({"result": accessToken, "firsOnboarding": False}), HTTPStatus.OK
    accessToken = int(str(uuid.uuid4().int)[:5])
    print(accessToken)
    # red.hset(int(str(created_by_phone)), "accessToken", accessToken)
    return jsonify({"result": accessToken, "firsOnboarding": True}), HTTPStatus.OK
