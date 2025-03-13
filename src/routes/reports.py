import datetime
from datetime import datetime
from flask import Blueprint, request, jsonify
from http import HTTPStatus

from openpyxl.reader.excel import load_workbook
import arrow as arrow
from sqlalchemy import or_
from .search_ent import filter_by_request
from .utils.validations import parse_payload
from ..logic.reports import task_to_done, get_reports_as_dict
from ..models import report_model
from ..models.apprentice_model import Apprentice
from ..models.user_model import User
from src.services import db
import uuid
from ..models.report_model import Report, report_as_meet, call_report

reports_form_blueprint = Blueprint("reports", __name__, url_prefix="/reports")


@reports_form_blueprint.route("", methods=["GET"])
def get_all_reports_form():
    try:
        user = request.args.get("userId")
        page_num = (
            int(request.args.get("page_num"))
            if request.args.get("page_num") is not None
            else None
        )
        query = request.args.get("query") or ""
        sortBy = request.args.get("sortBy") or ""
        user_ent = (
            db.session.query(
                User.role_ids, User.institution_id, User.cluster_id, User.id
            )
            .filter(User.id == user)
            .first()
        )
        new_list = []
        if not user_ent:
            apprentice = (
                db.session.query(Apprentice.id).filter(Apprentice.id == user).first()
            )
            if apprentice:
                reportList = apprentice_reports(apprentice.id)
                my_reports = get_reports_as_dict(reportList)
                new_list = sorted(my_reports, key=lambda x: x["date"], reverse=True)
            else:
                return jsonify({"result": "no user"}), HTTPStatus.BAD_REQUEST

        else:
            user_list = []
            if "0" in user_ent.role_ids:
                user_list = (
                    db.session.query(User.id).filter(User.id == user_ent.id).all()
                )
            if "1" in user_ent.role_ids:
                user_list = (
                    db.session.query(User.id)
                    .filter(
                        User.institution_id == user_ent.institution_id,
                        User.role_ids.contains("0") | User.role_ids.contains("1"),
                    )
                    .all()
                )
            if "2" in user_ent.role_ids:
                user_list = (
                    db.session.query(User.id)
                    .filter(
                        User.cluster_id == user_ent.cluster_id,
                        or_(
                            User.role_ids.contains("0"),
                            User.role_ids.contains("1"),
                            User.role_ids.contains("2"),
                        ),
                    )
                    .all()
                )
            if "3" in user_ent.role_ids:
                user_list = db.session.query(User.id).all()
            user_list_ids = [str(row_.id) for row_ in user_list]
            if page_num != None and query != "":
                paginated_query = (
                    db.session.query(
                        Report.ent_reported,
                        Report.ent_group,
                        Report.note,
                        Report.visit_date,
                        Report.id,
                        Report.title,
                        Report.description,
                        Report.attachments,
                        Report.allreadyread,
                        Report.created_at,
                        Report.user_id,
                    )
                    .filter(
                        Report.user_id.in_(user_list_ids), Report.title.contains(query)
                    )
                    .filter(Report.user_id.in_(user_list_ids))
                    .filter(Report.user_id!=Report.ent_reported)
                    .distinct(Report.id)
                ).all()
                results = sorted(paginated_query, key=lambda d: d[9], reverse=True)
                new_list = get_reports_as_dict(results[page_num * 30 - 30:page_num * 30])
                new_list = sorted(new_list, key=lambda x: x["date"], reverse=True)

            elif page_num != None and sortBy != "" and sortBy == "asc":
                all_distict_reports = (
                    db.session.query(
                        Report.ent_reported,
                        Report.ent_group,
                        Report.note,
                        Report.visit_date,
                        Report.id,
                        Report.title,
                        Report.description,
                        Report.attachments,
                        Report.allreadyread,
                        Report.created_at,
                        Report.user_id,
                    )
                    .filter(Report.user_id.in_(user_list_ids))
                    .filter(Report.user_id != Report.ent_reported)
                    .distinct(Report.id)
                ).all()
                results = sorted(all_distict_reports, key=lambda d: d[9], reverse=False)
                new_list = get_reports_as_dict(results[page_num * 30 - 30:page_num * 30])
                new_list = sorted(new_list, key=lambda x: x["date"], reverse=False)
            elif page_num != None and sortBy != "" and sortBy == "desc":
                all_distict_reports = (
                    db.session.query(
                        Report.ent_reported,
                        Report.ent_group,
                        Report.note,
                        Report.visit_date,
                        Report.id,
                        Report.title,
                        Report.description,
                        Report.attachments,
                        Report.allreadyread,
                        Report.created_at,
                        Report.user_id,
                    )
                    .filter(Report.user_id.in_(user_list_ids))
                    .filter(Report.user_id != Report.ent_reported)
                    .distinct(Report.id)
                ).all()
                results = sorted(all_distict_reports, key=lambda d: d[9], reverse=True)
                new_list = get_reports_as_dict(results[page_num * 30 - 30:page_num * 30])
                new_list = sorted(new_list, key=lambda x: x["date"], reverse=True)
            elif page_num != None and sortBy != "" and sortBy == "alphabeit":
                all_distict_reports = (
                    db.session.query(
                        Report.ent_reported,
                        Report.ent_group,
                        Report.note,
                        Report.visit_date,
                        Report.id,
                        Report.title,
                        Report.description,
                        Report.attachments,
                        Report.allreadyread,
                        Report.created_at,
                        Report.user_id,
                    )
                    .filter(Report.user_id.in_(user_list_ids))
                    .filter(Report.user_id != Report.ent_reported)
                    .distinct(Report.id)
                ).all()
                results = sorted(all_distict_reports, key=lambda d: d[9], reverse=True)
                new_list = get_reports_as_dict(results[page_num * 30 - 30:page_num * 30])
                new_list = sorted(new_list, key=lambda x: x["search"], reverse=False)
            elif page_num != None:
                #get all my  distinct reports
                all_distict_reports = (
                    db.session.query(
                        Report.ent_reported,
                        Report.ent_group,
                        Report.note,
                        Report.visit_date,
                        Report.id,
                        Report.title,
                        Report.description,
                        Report.attachments,
                        Report.allreadyread,
                        Report.created_at,
                        Report.user_id,
                    )
                    .filter(Report.user_id.in_(user_list_ids))
                    .distinct(Report.id)
                ).all()
                #sort by creation time
                results = sorted(all_distict_reports, key=lambda d: d[9], reverse=True)
                #paginate +create the list of reports with combine same id reports
                new_list = get_reports_as_dict(results[page_num*30-30:page_num*30])
                #sort again by creation date because order was lost in get_reports_as_dict
                new_list = sorted(new_list, key=lambda d: d['creation_date'], reverse=True)
        return jsonify(new_list), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST


def add_inherited_report(Visit1):
    if Visit1.title in report_as_meet:
        Visit1.title=call_report
        Visit1.id = uuid.uuid4()
        db.session.add(Visit1)
        task_to_done(Visit1.title, Visit1.user_id, Visit1.ent_reported)


@reports_form_blueprint.route("/add", methods=["post"])
def add_reports_form():
    try:
        data = request.json
        parse_payload_ = parse_payload(data)
        if not isinstance(parse_payload_, dict):
            return (
                jsonify({"result": "error" + str(parse_payload_)}),
                HTTPStatus.BAD_REQUEST,
            )

        user = parse_payload_["userId"]
        ent_group_name = parse_payload_["ent_group"]
        attachments = parse_payload_["attachments"]
        description = parse_payload_["description"]
        List_of_repored = parse_payload_["List_of_repored"]
        event_type = parse_payload_["event_type"]
        if event_type not in report_model.all_reports:
            return (
                jsonify({"result": "error" + ",no such report type"}),
                HTTPStatus.BAD_REQUEST,
            )

        if user:
            vis_id = uuid.uuid4()
            for ent_reported_ in List_of_repored:
                Visit1 = Report(
                    created_at=arrow.now().format("YYYY-MM-DDThh:mm:ss"),
                    user_id=user,
                    ent_reported=str(ent_reported_),
                    visit_in_army=(
                        True
                        if data["event_type"] == report_model.basis_report
                        else False
                    ),
                    visit_date=datetime.fromisoformat(data["date"]),
                    allreadyread=False,
                    id=vis_id,
                    title=event_type,
                    attachments=attachments,
                    ent_group=ent_group_name,
                    description=description,
                )
                db.session.add(Visit1)
                db.session.commit()
                add_inherited_report(Visit1)
                task_to_done(event_type, user, ent_reported_)

                db.session.commit()
        return jsonify({"result": "success"}), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": "error" + str(e)}), HTTPStatus.BAD_REQUEST


@reports_form_blueprint.route("/delete", methods=["POST"])
def delete():
    try:

        data = request.json
        reportIds = data["reportId"]
        for id_ in reportIds:
            db.session.query(Report).filter(Report.id == id_).delete()
        db.session.commit()
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST
    return jsonify({"result": "success"}), HTTPStatus.OK


@reports_form_blueprint.route("/update", methods=["put"])
def update():
    try:

        # get tasksAndEvents
        reportId = request.args.get("reportId")
        data = request.json
        updatedEnt = Report.query.get(reportId)
        for key in data:
            setattr(updatedEnt, key, data[key])
        db.session.commit()
        if updatedEnt:
            # TODO: add contact form to DB
            return jsonify({"result": "success"}), HTTPStatus.OK
        return jsonify({"result": "error"}), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST


@reports_form_blueprint.route("/filter_report", methods=["GET"])
def filter_report():
    try:
        users, apprentice, ent_group_dict = filter_by_request(request)
        types = (
            request.args.get("type").split(",")
            if request.args.get("type") is not None
            else None
        )
        page_num = int(
            request.args.get("page_num")
            if request.args.get("page_num") is not None
            else 0
        )
        
        # Get date range parameters
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        
        # Base query
        query = db.session.query(
            Report.ent_reported,
            Report.ent_group,
            Report.note,
            Report.visit_date,
            Report.id,
            Report.title,
            Report.description,
            Report.attachments,
            Report.allreadyread,
            Report.created_at,
            Report.user_id,
        ).filter(
            or_(Report.user_id.in_(users), Report.ent_reported.in_(apprentice))
        )

        # Add type filter if specified
        if types:
            query = query.filter(Report.title.in_(types))

        # Add date range filter if specified
        if start_date:
            query = query.filter(Report.visit_date >= datetime.fromisoformat(start_date))
        if end_date:
            query = query.filter(Report.visit_date <= datetime.fromisoformat(end_date))

        # Execute query with pagination
        reports_user = query.order_by(Report.created_at.desc()).distinct(Report.id, Report.created_at).paginate(
            page=page_num, per_page=30, error_out=False
        )

        return jsonify(get_reports_as_dict(reports_user)), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST


@reports_form_blueprint.route("/filter_recipients", methods=["GET"])
def filter_to():
    try:

        users, apprentice, ent_group_dict = filter_by_request(request)
        ent_group_concat = ""
        if apprentice != [] or users != []:
            ent_group_concat = ", ".join(ent_group_dict.values())
        result = set(users + apprentice)
        return (
            jsonify(
                {
                    "filtered": [str(row) for row in result],
                    "ent_group": ent_group_concat,
                }
            ),
            HTTPStatus.OK,
        )
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.OK


@reports_form_blueprint.route("/add_report_excel", methods=["put"])
def add_report_excel():
    # /home/ubuntu/flaskapp/

    file = request.files["file"]

    wb = load_workbook(file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2):
        user_id = row[0].value
        ent_reported = row[1].value
        visit_date = row[2].value
        title = row[3].value
        visit_in_army = row[4].value
        description = row[5].value
        attachments = str(row[6].value).split(",")
        ent_group = row[7].value
        if attachments == ["None"]:
            attachments = []
        rep = Report(
            user_id=user_id,
            ent_reported=ent_reported,
            visit_date=visit_date,
            title=title,
            visit_in_army=visit_in_army,
            description=description,
            attachments=attachments,
            allreadyread=False,
            ent_group=ent_group,
        )
        db.session.add(rep)
        task_to_done(title, user_id, str(ent_reported))

    try:
        db.session.commit()
    except Exception as e:
        return (
            jsonify({"result": "error while inserting" + str(e)}),
            HTTPStatus.BAD_REQUEST,
        )

    return jsonify({"result": "success"}), HTTPStatus.OK


def apprentice_reports(apprentice_id):

    reportList = (
        db.session.query(
            Report.ent_reported,
            Report.ent_group,
            Report.note,
            Report.visit_date,
            Report.id,
            Report.title,
            Report.description,
            Report.attachments,
            Report.allreadyread,
            Report.created_at,
            Report.user_id,
        )
        .filter(Report.ent_reported == apprentice_id)
        .all()
    )
    return reportList
