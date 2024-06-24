import datetime
from datetime import datetime
from flask import Blueprint, request, jsonify
from http import HTTPStatus

from openpyxl.reader.excel import load_workbook
import arrow as arrow
from sqlalchemy import or_

import config
from src.models.models_utils import to_iso
from .search_ent import filter_by_request
from .user_profile import correct_auth
from ..models.apprentice_model import Apprentice
from ..models.user_model import User
from src.services import db
import uuid
from ..models.report_model import Report

reports_form_blueprint = Blueprint('reports_form', __name__, url_prefix='/reports_form')


@reports_form_blueprint.route('/add', methods=['post'])
def add_reports_form():
    if correct_auth() == False:
        return jsonify({'result': "wrong access token"}), HTTPStatus.OK
    data = request.json
    user = str(data['userId'])
    ent_group_name = ""
    attachments = []
    description = ""

    ent_group_name = str(data.get('ent_group'))
    attachments = data.get('attachments')
    description = data.get('description')

    if user:
        List_of_repored = data['List_of_repored']
        vis_id = int(str(uuid.uuid4().int)[:5])
        event_type = data['event_type']
        if event_type == config.zoom_report or event_type == config.fiveMess_report:
            event_type = config.call_report
        for key in List_of_repored:
            Visit1 = Report(
                created_at=arrow.now().format('YYYY-MM-DDThh:mm:ss'),

                user_id=user,
                ent_reported=str(key),
                visit_in_army=True if data['event_type'] == config.basis_report else False,
                visit_date=datetime.fromisoformat(data['date']),
                allreadyread=False,
                id=vis_id,
                title=event_type,
                attachments=attachments,
                ent_group=ent_group_name,
                description=description
            )

            db.session.add(Visit1)

    try:
        db.session.commit()
    except Exception as e:
        return jsonify({'result': 'error' + str(e)}), HTTPStatus.BAD_REQUEST
    return jsonify({'result': "success"}), HTTPStatus.OK


@reports_form_blueprint.route('/getAll', methods=['GET'])
def getAll_reports_form():
    try:

        user = request.args.get('userId')
        reportList = db.session.query(Report.ent_reported, Report.ent_group, Report.note, Report.visit_date, Report.id,
                                      Report.title, Report.description, Report.attachments, Report.allreadyread,
                                      Report.created_at).filter(Report.user_id == user).all()
        user_name = db.session.query(User.name, User.last_name).filter(
            user == User.id).first()
        user_name = user_name.name + " " + user_name.last_name
        my_dict = []
        used_report = dict()
        for noti in reportList:
            used_report[str(noti.id)] = used_report.get(str(noti.id), []) + [noti]
        for k, notiList in used_report.items():
            reportedList = []
            reported_name_str = ""
            for noti in notiList:
                reportedList.append(str(noti.ent_reported))
                reported_name = db.session.query(Apprentice.name, Apprentice.last_name).filter(
                    Apprentice.id == noti.ent_reported).first()
                if reported_name is None:
                    reported_name = db.session.query(User.name, User.last_name).filter(
                        User.id == user).first()
                reported_name_str += reported_name.name + " " + reported_name.last_name + ","
            my_dict.append(
                {"search": reported_name_str + "," + user_name, "id": str(k), "reported_on": reportedList,
                 "date": to_iso(noti.visit_date), "creation_date": str(noti.created_at), "ent_group": noti.ent_group,
                 "title": str(noti.title), "allreadyread": str(noti.allreadyread), "description": str(noti.description),
                 "attachments": noti.attachments})
        #add my personas reports:
        user1ent = db.session.query(User.role_ids, User.institution_id, User.cluster_id,User.id).filter(
            User.id == user).first()
        if "0" in user1ent.role_ids:
            userList = []
        if "1" in user1ent.role_ids:
            userList = db.session.query(User.id).filter(User.institution_id == user1ent.institution_id,User.role_ids.contains("0")).all()
        if "2" in user1ent.role_ids:
            userList = db.session.query(User.id).filter(User.cluster_id == user1ent.cluster_id,or_(User.role_ids.contains("0"),User.role_ids.contains("1"))).all()
        if "3" in user1ent.role_ids:
            userList = db.session.query(User.id).all()
        for ent in userList:
            reportList = db.session.query(Report.ent_reported, Report.ent_group, Report.note, Report.visit_date, Report.id,
                                          Report.title, Report.description, Report.attachments, Report.allreadyread,
                                          Report.created_at).filter(Report.user_id == ent.id).all()
            user_name = db.session.query(User.name, User.last_name).filter(
                ent.id == User.id).first()
            user_name = user_name.name + " " + user_name.last_name
            used_report = dict()
            for noti in reportList:
                used_report[str(noti.id)] = used_report.get(str(noti.id), []) + [noti]
            for k, notiList in used_report.items():
                reportedList = []
                reported_name_str = ""
                for noti in notiList:
                    reportedList.append(str(noti.ent_reported))
                    reported_name = db.session.query(Apprentice.name, Apprentice.last_name).filter(
                        Apprentice.id == noti.ent_reported).first()
                    if reported_name is None:
                        reported_name = db.session.query(User.name, User.last_name).filter(
                            User.id == user).first()
                    reported_name_str += reported_name.name + " " + reported_name.last_name + ","
                my_dict.append(
                    {"search": reported_name_str + "," + user_name, "id": str(k), "reported_on": reportedList,
                     "date": to_iso(noti.visit_date), "creation_date": str(noti.created_at), "ent_group": noti.ent_group,
                     "title": str(noti.title), "allreadyread": str(noti.allreadyread),
                     "description": str(noti.description),
                     "attachments": noti.attachments})
        return jsonify(my_dict), HTTPStatus.OK
        # return jsonify([{'id':str(noti.id),'result': 'success',"apprenticeId":str(noti.apprenticeid),"date":str(noti.date),"timeFromNow":str(noti.timefromnow),"event":str(noti.event),"allreadyread":str(noti.allreadyread)}]), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@reports_form_blueprint.route('/delete', methods=['POST'])
def delete():
    try:
        if correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        data = request.json
        reportIds = data['reportId']
        for id in reportIds:
            res = db.session.query(Report).filter(Report.id == id).delete()
        db.session.commit()
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST
    return jsonify({"result": "success"}), HTTPStatus.OK
    # return jsonify([{'id':str(noti.id),'result': 'success',"apprenticeId":str(noti.apprenticeid),"date":str(noti.date),"timeFromNow":str(noti.timefromnow),"event":str(noti.event),"allreadyread":str(noti.allreadyread)}]), HTTPStatus.OK


@reports_form_blueprint.route("/update", methods=['put'])
def update():
    try:
        if correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        # get tasksAndEvents
        reportId = request.args.get("reportId")
        data = request.json
        updatedEnt = Report.query.get(reportId)
        for key in data:
            setattr(updatedEnt, key, data[key])
        db.session.commit()
        if updatedEnt:
            # TODO: add contact form to DB
            return jsonify({'result': 'success'}), HTTPStatus.OK
        return jsonify({'result': 'error'}), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@reports_form_blueprint.route("/filter_report", methods=['GET'])
def filter_report():
    try:
        if correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        users, apprentice, ent_group_dict = filter_by_request(request)
        types = request.args.get("type").split(",") if request.args.get("type") is not None else None
        if types:
            reports_user = db.session.query(Report.id).filter(Report.user_id.in_(users)).filter(
                Report.title.in_(types)).all()
            reports_apprentice = db.session.query(Report.id).filter(Report.ent_reported.in_(apprentice)).filter(
                Report.title.in_(types)).all()
        else:

            reports_user = db.session.query(Report.id).filter(Report.user_id.in_(users)).all()
            reports_apprentice = db.session.query(Report.id).filter(Report.ent_reported.in_(apprentice)).all()
        ent_group_concat = ", ".join(ent_group_dict.values())
        # reports_ent_group=db.session.query(Visit.id).filter(Visit.ent_group==ent_group.id,ent_group.group_name==ent_group_concat).all()
        users_mess = [str(i[0]) for i in [tuple(row) for row in reports_user]]
        apprentice_mess = [str(i[0]) for i in [tuple(row) for row in reports_apprentice]]
        # ent_group_mess=[str(i[0]) for i in [tuple(row) for row in reports_ent_group]]
        result = set(users_mess + apprentice_mess)
        return jsonify([str(row) for row in result]
                       ), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.BAD_REQUEST


@reports_form_blueprint.route("/filter_to", methods=['GET'])
def filter_to():
    try:
        if correct_auth()==False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        users, apprentice, ent_group_dict = filter_by_request(request)

        ent_group_concat = ""
        if apprentice != [] or users != []:
            ent_group_concat = ", ".join(ent_group_dict.values())
        result = set(users + apprentice)

        return jsonify({"filtered": [str(row) for row in result]

                           ,
                        "ent_group": ent_group_concat
                        }
                       ), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), HTTPStatus.OK


@reports_form_blueprint.route("/add_report_excel", methods=['put'])
def add_report_excel():
    # /home/ubuntu/flaskapp/
    if correct_auth() == False:
        return jsonify({'result': "wrong access token"}), HTTPStatus.OK
    file = request.files['file']

    wb = load_workbook(file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2):
        user_id = row[0].value
        ent_reported = row[1].value
        visit_date = row[2].value
        title = row[3].value
        visit_in_army = row[4].value
        description = row[5].value
        attachments = str(row[6].value).split(",")
        ent_group = row[7].value
        if attachments == ["None"]:
            attachments = []
        rep = Report(

            id=int(str(uuid.uuid4().int)[:5]),
            user_id=user_id,
            ent_reported=ent_reported,
            visit_date=visit_date,
            title=title,
            visit_in_army=visit_in_army,
            description=description,
            attachments=attachments,
            allreadyread=False,
            ent_group=ent_group
        )
        db.session.add(rep)
    try:
        db.session.commit()
    except Exception as e:
        return jsonify({'result': 'error while inserting' + str(e)}), HTTPStatus.BAD_REQUEST

    return jsonify({'result': 'success'}), HTTPStatus.OK
