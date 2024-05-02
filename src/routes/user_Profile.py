import json
import uuid
from enum import Enum

import boto3

from flask import Blueprint, request, jsonify, Response
from http import HTTPStatus

from openpyxl.reader.excel import load_workbook

import config
from app import db, red
from config import AWS_secret_access_key, AWS_access_key_id
from src.models.apprentice_model import Apprentice
from src.models.base_model import Base
from src.models.city_model import City
from src.models.cluster_model import Cluster
from src.models.contact_form_model import ContactForm
from src.models.institution_model import Institution
from src.models.notification_model import notifications
from src.models.user_model import user1, front_end_dict
from src.models.visit_model import Visit
from datetime import datetime,date

from src.routes.apprentice_Profile import visit_gap_color
from src.routes.setEntityDetails_form_routes import validate_email

userProfile_form_blueprint = Blueprint('userProfile_form', __name__, url_prefix='/userProfile_form')
role_name = Enum('Color', ['melave', 'racaz_mosad', 'racaz_eshcol'])
@userProfile_form_blueprint.route('/delete', methods=['post'])
def delete():
    try:
        data = request.json
        userId = data['userId']
        updatedEnt = user1.query.get(userId)
        if updatedEnt :
            db.session.query(ContactForm).filter(ContactForm.created_for_id == userId, ).delete()
            db.session.query(ContactForm).filter(ContactForm.created_by_id == userId, ).delete()
            db.session.query(notifications).filter(notifications.userid == userId, ).delete()
            db.session.query(user1).filter(user1.id == userId).delete()
        else:
            updatedEnt = Apprentice.query.get(userId)
            if updatedEnt:
                res = db.session.query(notifications).filter(notifications.subject == userId, ).delete()
                res = db.session.query(Visit).filter(Visit.ent_reported == userId, ).delete()
                res = db.session.query(Apprentice).filter(Apprentice.id == userId).delete()
            else:
                return jsonify({"result": str("no such id")}), HTTPStatus.BAD_REQUEST

        db.session.commit()
    except Exception as e:
        return jsonify({"result": str(e)}),HTTPStatus.BAD_REQUEST
    return jsonify({"result":"success"}), HTTPStatus.OK

@userProfile_form_blueprint.route("/update", methods=['put'])
def update():
    # get tasksAndEvents
    try:
        userId = request.args.get('userId')
        print(userId)
        data = request.json
        updatedEnt = user1.query.get(userId)
        print("data:",data)
        for key in data:
            if key == "city":
                CityId = db.session.query(City).filter(
                    City.name == str(data[key])).first()
                print("CityId",CityId)
                setattr(updatedEnt, "city_id", CityId.id)
            if key == "region":
                ClusterId = db.session.query(Cluster.id).filter(
                    Cluster.name == str(data[key])).first()
                print("ClusterId",ClusterId.id)
                setattr(updatedEnt, "cluster_id", ClusterId.id)
            elif key == "email" or key == "birthday":
                if validate_email(data[key]):
                    setattr(updatedEnt, key, data[key])
                else:
                    return jsonify({'result': "email or date -wrong format"}), 401
            else:
                setattr(updatedEnt, front_end_dict[key], data[key])

        db.session.commit()
        if updatedEnt:
            # print(f'setWasRead form: subject: [{subject}, notiId: {notiId}]')
            # TODO: add contact form to DB
            return jsonify({'result': 'success'}), HTTPStatus.OK
        return jsonify({'result': 'no updatedEnt'}), 401
    except Exception as e:
        return jsonify({'result': str(e)}), 401



@userProfile_form_blueprint.route('/getProfileAtributes', methods=['GET'])
def getProfileAtributes_form():
    try:
        created_by_id = request.args.get('userId')
        userEnt = user1.query.get(created_by_id)
        if userEnt:
            city = db.session.query(City).filter(City.id == userEnt.city_id).first()
            regionName=db.session.query(Cluster.name).filter(Cluster.id==city.cluster_id).first()
            myApprenticesNamesList=getmyApprenticesNames(created_by_id)
            city = db.session.query(City).filter(City.id == userEnt.city_id).first()
            list = {"id":str(userEnt.id), "firstName":userEnt.name, "lastName":userEnt.last_name, "date_of_birth": toISO(userEnt.birthday), "email":userEnt.email,
                           "city":city.name, "region":str(regionName[0]), "role":str(userEnt.role_id), "institution":str(userEnt.institution_id), "cluster":str(userEnt.eshcol),
                           "apprentices":myApprenticesNamesList, "phone":str(userEnt.id),"teudatZehut":str(userEnt.teudatZehut), "avatar":userEnt.photo_path if userEnt.photo_path is not None else 'https://www.gravatar.com/avatar'}
            return jsonify(list), HTTPStatus.OK
        else:
            return jsonify(results="no such id"), HTTPStatus.OK
    except Exception:
        return jsonify({'result': str(Exception)}), 401


def getmyApprenticesNames(created_by_id):
    try:
        apprenticeList = db.session.query(Apprentice.id,Apprentice.name,Apprentice.last_name).filter(Apprentice.accompany_id == created_by_id).all()
        return [{"id": str(row[0]), "name": row[1], "last_name": row[2]} for row in apprenticeList]
    except Exception as e:
        return jsonify({'result': str(e)}), 401




@userProfile_form_blueprint.route("/add_user_excel", methods=['put'])
def add_user_excel():

    file = request.files['file']
    path = 'data/user_enter.xlsx'
    wb = load_workbook(file)
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2):
        if row[5].value is None:
            continue
        if row[2].value.strip() == "מלווה" :
            role=0
        elif row[2].value.strip() == "רכז" :
            role = 1
        elif row[2].value.strip() == "רכז אשכול":
            role = 2
        elif row[2].value.strip() == "אחראי תוכנית":
            role = 3
        first_name =row[0].value.strip()
        last_name = row[1].value.strip()
        institution_name = row[3].value.strip()
        phone = str(row[5].value).replace("-","").strip()
        #email = row[3].value.strip()
        eshcol = row[4].value.strip()
        try:
            institution_id = db.session.query(Institution.id).filter(
                Institution.name == str(institution_name)).first()
            user = user1(
                id=int(str(phone).replace("-","")),
                name=first_name,
                last_name=last_name,
                role_id=str(role),
                #email=str(email),
                eshcol=eshcol,
                institution_id=institution_id.id,
            )
            print(user)
            db.session.add(user)
            db.session.commit()

        except Exception as e:
            return jsonify({'result': 'error while inserting' + str(e)}), HTTPStatus.BAD_REQUEST

    return jsonify({'result': 'success'}), HTTPStatus.OK

@userProfile_form_blueprint.route("/add_user_manual", methods=['post'])
def add_user_manual():
    data = request.json
    print(data)
    try:
        first_name = data['first_name']
        last_name = data['last_name']
        phone = data['phone']
        institution_id = data['institution_id']
        city_name="עלי"
        try:
            city_name = data['city_name'] or "עלי"
        except:
            print("no city")
        role_id = data['role_id']
        CityId = db.session.query(City).filter(City.name==city_name).first()
        #institution_id = db.session.query(Institution.id).filter(Institution.name==institution_name).first()
        institution_id=institution_id[0] if institution_id else 0
        print(institution_id)
        useEnt = user1(
            id=int(phone[1:]),
            name=first_name,
            last_name=last_name,
            role_id=role_id,
            institution_id=institution_id,
            city_id=CityId.id,
            cluster_id=CityId.cluster_id,
            photo_path="https://www.gravatar.com/avatar"
        )
        db.session.add(useEnt)
        db.session.commit()
    except Exception as e:
        return jsonify({'result': 'error while inserting'+str(e)}), HTTPStatus.BAD_REQUEST

    if useEnt:
        # TODO: add contact form to DB
        return jsonify({'result': 'success'}), HTTPStatus.OK

def toISO(d):
    if d:
        return datetime(d.year, d.month, d.day).isoformat()
    else:
        return None

@userProfile_form_blueprint.route('/myPersonas', methods=['GET'])
def myPersonas():
    try:
        created_by_id = request.args.get('userId')
        print(created_by_id)
        apprenticeList=[]
        user1ent = db.session.query(user1.role_id,user1.institution_id,user1.eshcol).filter(user1.id==created_by_id).first()
        if user1ent.role_id=="0":
            apprenticeList = db.session.query(Apprentice).filter(Apprentice.accompany_id == created_by_id).all()
            userList=[]
        if user1ent.role_id=="1":
            apprenticeList = db.session.query(Apprentice).filter(Apprentice.institution_id == user1ent.institution_id).all()
            userList = db.session.query(user1).filter(user1.institution_id == user1ent.institution_id).all()
        if user1ent.role_id == "2":
            apprenticeList = db.session.query(Apprentice).filter(Apprentice.eshcol == user1ent.eshcol).all()
            userList = db.session.query(user1).filter(user1.institution_id == user1ent.institution_id).all()
        if user1ent.role_id == "3":
            apprenticeList = db.session.query(Apprentice).all()
            userList = db.session.query(user1).all()

        my_dict = []

        for noti in apprenticeList:
            accompany = db.session.query(user1.name,user1.last_name).filter(user1.id == Apprentice.accompany_id).first()
            call_status=visit_gap_color(config.call_report, noti, 30, 15)
            personalMeet_status=visit_gap_color(config.personalMeet_report, noti, 100, 80)
            Horim_status=visit_gap_color(config.HorimCall_report, noti, 365, 350)
            city = db.session.query(City).filter(City.id == noti.city_id).first()
            reportList = db.session.query(Visit.id).filter(Visit.ent_reported == noti.id).all()
            eventlist = db.session.query(notifications.id,notifications.event,notifications.details,notifications.date).filter(
                                                                               notifications.subject == str(noti.id),
                                                                               notifications.numoflinesdisplay == 3).all()
            base_id = db.session.query(Base.id).filter(Base.id == int(noti.base_address)).first()
            base_id = base_id[0] if base_id else 0
            my_dict.append(
                {"Horim_status":Horim_status,
                 "personalMeet_status":personalMeet_status,
                 "call_status":call_status,
                 "highSchoolRavMelamed_phone": noti.high_school_teacher_phone
                         ,"highSchoolRavMelamed_name": noti.high_school_teacher,
                      "highSchoolRavMelamed_email": noti.high_school_teacher_email,

                     "thRavMelamedYearA_name": noti.teacher_grade_a,
                     "thRavMelamedYearA_phone": noti.teacher_grade_a_phone,
                     "thRavMelamedYearA_email": noti.teacher_grade_a_email,

                     "thRavMelamedYearB_name": noti.teacher_grade_b,
                     "thRavMelamedYearB_phone": noti.teacher_grade_b_phone,
                     "thRavMelamedYearB_email": noti.teacher_grade_b_email,
                    "address": {
                        "country": "IL",
                        "city": city.name if city else "",
                        "cityId": str(noti.city_id),
                        "street": noti.address,
                        "houseNumber": "1",
                        "apartment": "1",
                        "region": str(city.cluster_id) if city else "",
                        "entrance": "a",
                        "floor": "1",
                        "postalCode": "12131",
                        "lat": 32.04282620026557,#no need city cord
                        "lng": 34.75186193813887
                    },
                 "contact1_first_name": noti.contact1_first_name,
                 "contact1_last_name": noti.contact1_last_name,
                 "contact1_phone": noti.contact1_phone,
                 "contact1_email": noti.contact1_email,
                 "contact1_relation": noti.contact1_relation,
                 "contact2_first_name": noti.contact2_first_name,
                 "contact2_last_name": noti.contact2_last_name,
                 "contact2_phone": noti.contact2_phone,
                 "contact2_email": noti.contact2_email,
                 "contact2_relation": noti.contact2_relation,
                 "contact3_first_name": noti.contact3_first_name,
                 "contact3_last_name": noti.contact3_last_name,
                 "contact3_phone": noti.contact3_phone,
                 "contact3_email": noti.contact3_email,
                 "contact3_relation": noti.contact3_relation,
                 "activity_score": len(reportList),

                 "reports":
                     [str(i[0]) for i in [tuple(row) for row in reportList]]
                 ,
                 "events":

                      [{"id" :str(row[0]),"title":row[1],"description":row[2],"date" : toISO(row[3])} for row in eventlist]

                    , "id": str(noti.id), "thMentor_name": accompany.name+" "+accompany.last_name,"thMentor_id": str(Apprentice.accompany_id),
                 "militaryPositionNew": str(noti.militaryPositionNew)
                    , "avatar": noti.photo_path if noti.photo_path is not None else 'https://www.gravatar.com/avatar' , "name": str(noti.name), "last_name": str(noti.last_name),
                 "institution_id": str(noti.institution_id), "thPeriod": str(noti.hadar_plan_session),
                 "serve_type": noti.serve_type,
                 "marriage_status": str(noti.marriage_status), "militaryCompoundId": str(base_id),
                 "phone": str(noti.id), "email": noti.email, "teudatZehut": noti.teudatZehut,
                 "birthday": toISO(noti.birthday),  "marriage_date": toISO(noti.marriage_date),
                 "highSchoolInstitution": noti.highSchoolInstitution, "army_role": noti.army_role,
                 "unit_name": noti.unit_name,
                  "matsber": str(noti.spirit_status),
                 "militaryDateOfDischarge": toISO(noti.release_date),
                 "militaryDateOfEnlistment": toISO(noti.recruitment_date)
                    , "militaryUpdatedDateTime": toISO(noti.militaryupdateddatetime),
                 "militaryPositionOld": noti.militaryPositionOld, "educationalInstitution": noti.educationalinstitution
                    , "educationFaculty": noti.educationfaculty, "workOccupation": noti.workoccupation,
                 "workType": noti.worktype, "workPlace": noti.workplace, "workStatus": noti.workstatus, "paying": noti.paying

                 })
            for noti in userList:
                reportList = db.session.query(Visit.id).filter(Visit.user_id == noti.id).all()
                city = db.session.query(City).filter(City.id == noti.city_id).first()
                my_dict.append(
                    {"Horim_status": "",
                     "personalMeet_status": "",
                     "call_status": "",
                     "highSchoolRavMelamed_phone": ""
                        , "highSchoolRavMelamed_name": "",
                     "highSchoolRavMelamed_email": "",

                     "thRavMelamedYearA_name": "",
                     "thRavMelamedYearA_phone": "",
                     "thRavMelamedYearA_email": "",

                     "thRavMelamedYearB_name": "",
                     "thRavMelamedYearB_phone": "",
                     "thRavMelamedYearB_email": "",
                     "address": {
                         "country": "IL",
                         "city": city.name if city else "",
                         "cityId": str(noti.city_id),
                         "street": noti.address,
                         "houseNumber": "1",
                         "apartment": "1",
                         "region": str(city.cluster_id) if city else "",
                         "entrance": "a",
                         "floor": "1",
                         "postalCode": "12131",
                         "lat": 32.04282620026557,  # no need city cord
                         "lng": 34.75186193813887
                     },
                     "contact1_first_name": "",
                     "contact1_last_name": "",
                     "contact1_phone": "",
                     "contact1_email": "",
                     "contact1_relation": "",
                     "contact2_first_name": "",
                     "contact2_last_name": "",
                     "contact2_phone": "",
                     "contact2_email": "",
                     "contact2_relation": "",
                     "contact3_first_name": "",
                     "contact3_last_name": "",
                     "contact3_phone": "",
                     "contact3_email": "",
                     "contact3_relation": "",
                     "activity_score": len(reportList),

                     "reports":
                         ""
                        ,
                     "events":

                         ""

                        , "id": str(noti.id),
                     "thMentor": "",
                     "militaryPositionNew": ""
                        ,
                     "avatar": noti.photo_path if noti.photo_path is not None else 'https://www.gravatar.com/avatar',
                     "name": str(noti.name), "last_name": str(noti.last_name),
                     "institution_id": str(noti.institution_id), "thPeriod": "",
                     "serve_type": "",
                     "marriage_status": "", "militaryCompoundId": "",
                     "phone": str(noti.id),
                     "email": noti.email,
                     "teudatZehut": noti.teudatZehut,
                     "birthday": "",
                     "marriage_date": "",
                     "highSchoolInstitution": "",
                     "army_role": "",
                     "unit_name": "",
                     "matsber": "",
                     "militaryDateOfDischarge": "",
                     "militaryDateOfEnlistment": ""
                        , "militaryUpdatedDateTime": "",
                     "militaryPositionOld": "", "educationalInstitution": ""
                        , "educationFaculty": "", "workOccupation": "",
                     "workType": "", "workPlace": "", "workStatus": "",
                     "paying": ""

                     })
            return  jsonify(my_dict)
    except Exception as e:
        return jsonify({'result': str(e)}), 401