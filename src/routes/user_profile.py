from http import HTTPStatus
from flask import Blueprint, request, jsonify

from openpyxl.reader.excel import load_workbook

import config
from src.models.region_model import Region
from src.services import db, red
from src.models.apprentice_model import Apprentice
from src.models.city_model import City

from src.models.message_model import Message
from src.models.institution_model import Institution
from src.models.task_model import Task
from src.models.user_model import User
from src.models.report_model import Report
from src.routes.set_entity_details_form_routes import validate_email, validate_date
from src.logic import my_personas
import json


userProfile_form_blueprint = Blueprint(
    'userProfile_form', __name__, url_prefix='/userProfile_form')


@userProfile_form_blueprint.route('/delete', methods=['post'])
def delete():
    try:
        if correct_auth() == False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        data = request.json
        userId = data['userId']
        updatedEnt = User.query.get(userId)
        if updatedEnt:
            db.session.query(Message).filter(
                Message.created_for_id == userId, ).delete()
            db.session.query(Message).filter(
                Message.created_by_id == userId, ).delete()
            db.session.query(Task).filter(Task.userid == userId).delete()
            db.session.query(User).filter(User.id == userId).delete()
        else:
            updatedEnt = Apprentice.query.get(userId)
            if updatedEnt:
                res = db.session.query(Task).filter(
                    Task.subject == userId).delete()
                res = db.session.query(Report).filter(
                    Report.ent_reported == userId, ).delete()
                res = db.session.query(Apprentice).filter(
                    Apprentice.id == userId).delete()
            else:
                return jsonify({"result": str("no such id")}), HTTPStatus.BAD_REQUEST

        db.session.commit()
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST
    return jsonify({"result": "success"}), HTTPStatus.OK


@userProfile_form_blueprint.route("/update", methods=['PUT'])
def update():
    try:
        if not correct_auth():
            return jsonify({'result': "wrong access token"}), 401

        apprentice_id = request.args.get('apprenticetId')
        data = request.json

        if not apprentice_id:
            return jsonify({'result': "apprenticetId is required"}), 400

        updatedEnt = User.query.get(apprentice_id)
        if not updatedEnt:
            return jsonify({'result': 'User not found'}), 404

        # Update fields based on the Flutter application structure
        updatedEnt.avatar = data.get('avatar')

        # Military info
        updatedEnt.military_compound_id = data.get('militaryCompoundId')
        updatedEnt.military_service_type = data.get('militaryUnit')
        updatedEnt.military_position_new = data.get('militaryPositionNew')
        updatedEnt.military_position_old = data.get('militaryPositionOld')
        updatedEnt.military_date_of_enlistment = data.get(
            'militaryDateOfEnlistment')
        updatedEnt.military_date_of_discharge = data.get(
            'militaryDateOfDischarge')

        # Personal general info
        updatedEnt.teudat_zehut = data.get('teudatZehut')
        if 'email' in data and validate_email(data['email']):
            updatedEnt.email = data['email']
        updatedEnt.marital_status = data.get('marriage_status')
        updatedEnt.phone = data.get('phone')

        # Personal dates
        if 'birthday' in data and validate_date(data['birthday']):
            updatedEnt.date_of_birth = data['birthday']

        # Personal relationships
        for i in range(1, 4):
            setattr(updatedEnt, f'contact{i}_relationship', data.get(f'contact{i}_relation'))
            setattr(updatedEnt, f'contact{i}_phone', data.get(f'contact{i}_phone'))
            setattr(updatedEnt, f'contact{i}_email', data.get(f'contact{i}_email'))
            setattr(updatedEnt, f'contact{i}_first_name', data.get(f'contact{i}_first_name'))
            setattr(updatedEnt, f'contact{i}_last_name', data.get(f'contact{i}_last_name'))

        # Personal high school
        updatedEnt.high_school_institution = data.get('highSchoolInstitution')
        updatedEnt.high_school_rav_melamed_name = data.get('highSchoolRavMelamed_name')
        updatedEnt.high_school_rav_melamed_phone = data.get('highSchoolRavMelamed_phone')
        updatedEnt.high_school_rav_melamed_email = data.get('highSchoolRavMelamed_email')

        # Personal work
        updatedEnt.work_status = data.get('workStatus')
        updatedEnt.work_occupation = data.get('workOccupation')
        updatedEnt.work_place = data.get('workPlace')
        updatedEnt.work_type = data.get('workType')

        # Tohnit hadar
        updatedEnt.educational_institution = data.get('educationalInstitution')
        updatedEnt.th_period = data.get('thPeriod')
        updatedEnt.th_rav_melamed_year_a_name = data.get('thRavMelamedYearA_name')
        updatedEnt.th_rav_melamed_year_a_phone = data.get('thRavMelamedYearA_phone')
        updatedEnt.th_rav_melamed_year_b_name = data.get('thRavMelamedYearB_name')
        updatedEnt.th_rav_melamed_year_b_phone = data.get('thRavMelamedYearB_phone')
        updatedEnt.paying = data.get('paying')
        updatedEnt.matsbar_status = data.get('matsber')

        db.session.commit()
        return jsonify({'result': 'success'}), 200

    except Exception as e:
        print(f"Error: {str(e)}")  # Log the full error
        return jsonify({'result': 'An error occurred'}), 500


@userProfile_form_blueprint.route('/getProfileAtributes', methods=['GET'])
def getProfileAtributes_form():
    try:
        if correct_auth() == False:
            return jsonify({'result': "wrong access token"}), HTTPStatus.OK
        created_by_id = request.args.get('userId')
        userEnt = User.query.get(created_by_id)
        if userEnt:
            city = db.session.query(City).filter(
                City.id == userEnt.city_id).first()
            regionName = db.session.query(Region.name).filter(
                Region.id == userEnt.region_id).first()
            myApprenticesNamesList = getmyApprenticesNames(created_by_id)
            city = db.session.query(City).filter(
                City.id == userEnt.city_id).first()
            list = userEnt.to_attributes(city.name, str(
                regionName[0]), myApprenticesNamesList)
            return jsonify(list), HTTPStatus.OK
        else:
            return jsonify(results="no such id"), HTTPStatus.OK
    except Exception as e:
        return jsonify({'result': str(e)}), 401


def getmyApprenticesNames(created_by_id):
    try:
        apprenticeList = db.session.query(Apprentice.id, Apprentice.name, Apprentice.last_name).filter(
            Apprentice.accompany_id == created_by_id).all()
        return [{"id": str(row[0]), "name": row[1], "last_name": row[2]} for row in apprenticeList]
    except Exception as e:
        return jsonify({'result': str(e)}), 401


@userProfile_form_blueprint.route("/add_user_excel", methods=['put'])
def add_user_excel():
    if correct_auth() == False:
        return jsonify({'result': "wrong access token"}), HTTPStatus.OK
    file = request.files['file']
    wb = load_workbook(file)
    sheet = wb.active
    uncommited_ids = []
    for row in sheet.iter_rows(min_row=2):
        if row[0].value is None or row[1].value is None or row[2].value is None or row[5].value is None:
            uncommited_ids.append(row[5].value)
            continue
        role_ids = ""
        if "מלווה" in row[2].value.strip():
            role_ids += "0,"
        if "רכז מוסד" in row[2].value.strip():
            role_ids += "1,"
        if "רכז אשכול" in row[2].value.strip():
            role_ids += "2,"
        if "אחראי תוכנית" in row[2].value.strip():
            role_ids += "3,"
        role_ids = role_ids[:-1]
        first_name = row[0].value.strip()
        last_name = row[1].value.strip()
        institution_name = row[3].value.strip(
        ) if not row[3].value is None else "לא ידוע"
        eshcol = row[4].value.strip(
        ) if not row[4].value is None else "" if not row[4].value is None else "לא ידוע"
        phone = str(row[5].value).replace("-", "").strip()
        # email = row[3].value.strip()
        try:
            institution_id = db.session.query(Institution.id).filter(
                Institution.name == str(institution_name)).first()

            user = User(
                id=int(str(phone).replace("-", "")),
                name=first_name,
                last_name=last_name,
                role_ids=role_ids,
                # email=str(email),
                cluster_id=eshcol,
                institution_id=institution_id[0] if institution_id else None,
            )

            db.session.add(user)

            db.session.commit()
        except Exception as e:
            return jsonify({'result': 'error while inserting' + str(e)}), HTTPStatus.BAD_REQUEST
    return jsonify({'result': "success", "uncommited_ids": [x for x in uncommited_ids if x is not None]})


@userProfile_form_blueprint.route("/add_user_manual", methods=['post'])
def add_user_manual():
    # if correct_auth() == False:
    #     return jsonify({'result': "wrong access token"}), HTTPStatus.OK
    data = request.json
    try:
        first_name = data['first_name']
        last_name = data['last_name']
        phone = data['phone']
        institution_id = data['institution_id']
        city_name = "עלי"
        try:
            city_name = data['city_name'] or "עלי"
        except:
            print("no city")
        role_ids = data['role_ids']
        CityId = db.session.query(City).filter(City.name == city_name).first()
        # institution_id = db.session.query(Institution.id).filter(Institution.name==institution_name).first()
        institution_id = institution_id[0] if institution_id else 0
        useEnt = User(
            id=int(phone[1:]),
            name=first_name,
            last_name=last_name,
            role_ids=role_ids,
            institution_id=institution_id,
            city_id=CityId.id,
            cluster_id=CityId.cluster_id,
            photo_path="https://www.gravatar.com/avatar"
        )
        db.session.add(useEnt)
        db.session.commit()
    except Exception as e:
        return jsonify({'result': 'error while inserting' + str(e)}), HTTPStatus.BAD_REQUEST

    if useEnt:
        # TODO: add contact form to DB
        return jsonify({'result': 'success'}), HTTPStatus.OK


@userProfile_form_blueprint.route('/myPersonas', methods=['GET'])
def myPersonas():
    try:
        return my_personas.get_personas(request.args.get('userId'))
    except Exception as e:
        raise e


def correct_auth(external=True):
    if config.Authorization_is_On and external:
        userId = request.args.get("userId")
        accessToken = request.headers.get('Authorization')
        redisaccessToken = red.hget(userId, "accessToken").decode("utf-8")
        print("accessToken: ", accessToken)
        print("redisaccessToken: ", redisaccessToken)
        if redisaccessToken != accessToken:
            return False
        return True
    return True
