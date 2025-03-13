from http import HTTPStatus
from flask import Blueprint, request, jsonify

from openpyxl.reader.excel import load_workbook

from src.models.cluster_model import Cluster
from src.models.region_model import Region
from src.routes.utils.validations import validate_date, validate_email, parse_payload
from src.services import db
from src.models.apprentice_model import Apprentice
from src.models.city_model import City

from src.models.message_model import Message
from src.models.institution_model import Institution
from src.models.task_model_v2 import Task
from src.models.user_model import User
from src.models.report_model import Report
from src.logic import my_personas

import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


userProfile_form_blueprint = Blueprint("users", __name__, url_prefix="/users")


@userProfile_form_blueprint.route("/program_managers", methods=["get"])
def get_program_managers():
    # if not verify_access_token():
    #     return jsonify({"result": "wrong access token"}), HTTPStatus.BAD_REQUEST
    return my_personas.get_program_managers()


@userProfile_form_blueprint.route("/delete", methods=["post"])
def delete():
    try:

        data = request.json
        userId = data["userId"]
        updatedEnt = User.query.get(userId)
        if updatedEnt:
            db.session.query(Message).filter(
                Message.created_for_id == userId,
            ).delete()
            db.session.query(Message).filter(
                Message.created_by_id == userId,
            ).delete()
            db.session.query(Report).filter(
                Report.ent_reported == userId,
            ).delete()
            db.session.query(Report).filter(
                Report.user_id == userId,
            ).delete()
            db.session.query(Task).filter(Task.user_id == userId).delete()
            db.session.query(User).filter(User.id == userId).delete()
        else:
            updatedEnt = Apprentice.query.get(userId)
            if updatedEnt:
                db.session.query(Task).filter(Task.subject_id == userId).delete()
                (
                    db.session.query(Report)
                    .filter(
                        Report.ent_reported == userId,
                    )
                    .delete()
                )
                (db.session.query(Apprentice).filter(Apprentice.id == userId).delete())
            else:
                return jsonify({"result": str("no such id")}), HTTPStatus.BAD_REQUEST

        db.session.commit()
    except Exception as e:
        return jsonify({"result": str(e)}), HTTPStatus.BAD_REQUEST
    return jsonify({"result": "success"}), HTTPStatus.OK


@userProfile_form_blueprint.route("/update", methods=["PUT"])
def update():
    try:

        userId = request.args.get("userId")
        data = request.json
        logger.debug(f"Received update data: {data}")
        updatedEnt = User.query.get(userId)
        if not updatedEnt:
            return jsonify({"result": "User not found"}), HTTPStatus.NOT_FOUND

        logger.debug(f"Current cluster_id: {updatedEnt.cluster_id}")
        for key, value in data.items():
            if key not  in User.__dict__:
                return (
                    jsonify({"result": f"Invalid attribute for {key}"}),
                    HTTPStatus.BAD_REQUEST,
                )
            logger.debug(f"Processing field: {key} with value: {value}")
            if key == "birthday":
                if validate_date(value):
                    setattr(updatedEnt, key, value)
                else:
                    return (
                        jsonify({"result": f"Invalid date format for {key}"}),
                        HTTPStatus.BAD_REQUEST,
                    )
            elif key == "email":
                if validate_email(value):
                    setattr(updatedEnt, key, value)
                else:
                    return (
                        jsonify({"result": f"Invalid email format for {key}"}),
                        HTTPStatus.BAD_REQUEST,
                    )
            elif key == "role_ids":
                if validate_role_ids(value):
                    setattr(updatedEnt, key, ",".join(map(str, value)))
                else:
                    return (
                        jsonify({"result": f"Invalid role_ids format for {key}"}),
                        HTTPStatus.BAD_REQUEST,
                    )
            elif key == "institution":
                institution_id = (
                    db.session.query(Institution.id)
                    .filter(Institution.name == str(data[key]))
                    .first()
                )
                setattr(updatedEnt, "institution_id", institution_id.id)
            elif key == "city":
                city_id = (
                    db.session.query(City.id)
                    .filter(City.name == str(data[key]))
                    .first()
                )
                setattr(updatedEnt, "city_id", city_id.id)
            elif key == "cluster_id":
                logger.debug(f"Attempting to update cluster_id to: {value}")
                setattr(updatedEnt, key, value)
            else:
                setattr(updatedEnt, key, value)

        logger.debug(f"All fields processed. Committing changes.")
        db.session.commit()
        logger.debug(f"Changes committed successfully.")
        logger.debug(f"New cluster_id: {updatedEnt.cluster_id}")
        return jsonify({"result": "success"}), HTTPStatus.OK

    except Exception as e:
        logger.error(f"Error during update: {str(e)}")
        db.session.rollback()
        return jsonify({"result": str(e)}), HTTPStatus.INTERNAL_SERVER_ERROR


def validate_role_ids(role_ids):
    if not isinstance(role_ids, list):
        return False

    # Check if all elements are integers between 0 and 3
    return all(isinstance(role, int) and 0 <= role <= 3 for role in role_ids)


@userProfile_form_blueprint.route("/get_profile_attributes", methods=["GET"])
def get_profile_atributes_form():
    try:

        created_by_id = request.args.get("userId")
        userEnt = User.query.get(created_by_id)
        if userEnt:
            regionName = (
                db.session.query(Region.name)
                .filter(Region.id == userEnt.region_id)
                .first()
            )
            city = db.session.query(City).filter(City.id == userEnt.city_id).first()
            user_to_attributes = userEnt.to_attributes(
                city.name if city else None,
                str(regionName.name) if regionName else None,
            )
            return jsonify(user_to_attributes), HTTPStatus.OK
        else:
            return jsonify(results="no such id"), HTTPStatus.OK
    except Exception as e:
        return jsonify({"result": str(e)}), 401


def get_my_apprentices_names(created_by_id):
    try:
        apprenticeList = (
            db.session.query(Apprentice.id, Apprentice.name, Apprentice.last_name)
            .filter(Apprentice.accompany_id == created_by_id)
            .all()
        )
        return [
            {"id": str(row[0]), "name": row[1], "last_name": row[2]}
            for row in apprenticeList
        ]
    except Exception as e:
        return jsonify({"result": str(e)}), 401


@userProfile_form_blueprint.route("/add_user_excel", methods=["put"])
def add_user_excel():

    file = request.files["file"]
    wb = load_workbook(file)
    sheet = wb.active
    uncommited_ids = []
    index_ = 2
    for row in sheet.iter_rows(min_row=2):
        phone=''.join(c for c in str(row[5].value) if c.isdigit())
        if (
            row[0].value is None
            or row[1].value is None
            or row[2].value is None
            or int(phone) < 500000000
            or row[5].value is None
        ):
            uncommited_ids.append(index_)
            index_ = index_ + 1
            continue
        role_ids = ""
        if "מלווה" in row[2].value.strip():
            role_ids += "0,"
        if "רכז מוסד" in row[2].value.strip():
            role_ids += "1,"
        if "רכז אשכול" in row[2].value.strip():
            role_ids += "2,"
        if "אחראי תוכנית" in row[2].value.strip():
            role_ids += "3,"
        role_ids = role_ids[:-1]
        first_name = row[0].value.strip()
        last_name = row[1].value.strip()
        institution_name = row[3].value.strip() if not row[3].value is None else None
        cluster_name = row[4].value.strip() if not row[4].value is None else None
        phone = str(row[5].value).replace("-", "").strip()
        email = row[6].value.strip() if not row[6].value is None else None
        try:
            institution_id = (
                db.session.query(Institution.id)
                .filter(Institution.name == str(institution_name))
                .first()
            )
            cluster_id = (
                db.session.query(Cluster).filter(Cluster.name == cluster_name).first()
            )
            id_ = int(str(phone)[:9].replace("-", ""))
            User_ = db.session.query(User).filter(User.id == id_).first()
            if User_:
                User_.name = first_name
                User_.last_name = last_name
                User_.role_ids = role_ids
                User_.email = email
                User_.cluster_id = cluster_id.id if cluster_id else None
                User_.institution_id = institution_id.id if institution_id else None
                db.session.add(User_)
                db.session.commit()
                index_ = index_ + 1
                continue
            user = User(
                id=id_,
                name=first_name,
                last_name=last_name,
                role_ids=role_ids,
                email=str(email),
                cluster_id=cluster_id.id if cluster_id else None,
                institution_id=institution_id[0] if institution_id else None,
            )
            db.session.add(user)
            db.session.commit()
            index_ = index_ + 1

        except Exception as e:
            return (
                jsonify({"result": "error while inserting" + str(e)}),
                HTTPStatus.BAD_REQUEST,
            )
    return jsonify(
        {
            "result": "success",
            "not_commited": [x for x in uncommited_ids if x is not None],
        }
    )


@userProfile_form_blueprint.route("/add_user_manual", methods=["post"])
def add_user_manual():
    try:
        data = request.json
        parse_payload_ = parse_payload(data)
        if not isinstance(parse_payload_, dict):
            return (
                jsonify({"result": "error" + str(parse_payload_)}),
                HTTPStatus.BAD_REQUEST,
            )
        first_name = parse_payload_["first_name"]
        last_name = parse_payload_["last_name"]
        phone = parse_payload_["phone"]
        institution_id = parse_payload_["institution_id"]
        role_ids = parse_payload_["role_ids"]
        city_id=None
        if "city_id" in parse_payload_.keys():
            city_id = parse_payload_["city_id"]

        # city = db.session.query(City).filter(City.name == city_name).first()
        # institution_id = db.session.query(Institution.id).filter(Institution.name==institution_name).first()
        useEnt = User(
            id=int(phone[1:]),
            name=first_name,
            last_name=last_name,
            role_ids=role_ids,
            institution_id=institution_id,
            city_id=city_id,
            photo_path="https://www.gravatar.com/avatar",
        )
        db.session.add(useEnt)
        db.session.commit()
    except Exception as e:
        return (
            jsonify({"result": "error while inserting" + str(e)}),
            HTTPStatus.BAD_REQUEST,
        )

    if useEnt:
        # TODO: add contact form to DB
        return jsonify({"result": "success"}), HTTPStatus.OK


@userProfile_form_blueprint.route("/my_personas", methods=["GET"])
def get_my_personas():
    try:
        return my_personas.get_personas(request.args.get("userId"))
    except Exception as e:
        raise e
